from __future__ import annotations

import re
from urllib.parse import quote, urlsplit, urlunsplit

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import OUT, SOURCES, SOURCE_PLAYBOOKS
from .models import CompanyRecord, DiscoveryHit, TODAY, TriggerEvent
from .pipeline import adapter_inventory_label, classify_company, lead_filter_fields, next_action_for_lead, primary_discovery, primary_trigger, score_company_metrics


def excel_safe(value):
    if isinstance(value, str):
        value = re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", value)
        if value and value[0] in ("=", "+", "-", "@"):
            return "'" + value
    return value


def append_excel_row(ws, row):
    ws.append([excel_safe(value) for value in row])


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="Aptos", size=10, color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Aptos", size=10)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

def size_columns(ws, widths=None):
    widths = widths or {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
            continue
        max_len = max((len(str(cell.value)) for cell in ws[letter] if cell.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)

def clean_hyperlink_target(value):
    if not isinstance(value, str):
        return None
    value = value.strip()
    if not value.startswith(("http://", "https://")):
        return None
    if any(ord(char) < 32 for char in value):
        return None
    parts = urlsplit(value)
    if not parts.netloc:
        return None
    path = quote(parts.path, safe="/%:@")
    query = quote(parts.query, safe="=&?/%:@,+;")
    fragment = quote(parts.fragment, safe="=&?/%:@,+;")
    return urlunsplit((parts.scheme, parts.netloc, path, query, fragment))

def add_hyperlinks(ws, cols: list[int]):
    for row in range(2, ws.max_row + 1):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            target = clean_hyperlink_target(cell.value)
            if target:
                cell.value = target
                cell.hyperlink = target
                cell.style = "Hyperlink"


def linkedin_contact_values(contact):
    if contact is None:
        return ["", "", ""]
    return [contact.name, contact.title, contact.url]


def style_linkedin_columns(ws, start_col: int) -> None:
    fill = PatternFill("solid", fgColor="0F6B78")
    for cell in ws[1][start_col - 1:]:
        cell.fill = fill

def write_workbook(companies: dict[str, CompanyRecord], discovery_hits: list[DiscoveryHit], trigger_events: list[TriggerEvent], run_log: list[list[str]]):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Pipeline Summary")
    append_excel_row(ws, ["Metric", "Value"])
    append_excel_row(ws, ["Run date", TODAY])
    append_excel_row(ws, ["Approved sources", len(SOURCES)])
    append_excel_row(ws, ["Fetched/skipped sources", len(run_log)])
    append_excel_row(ws, ["Discovery hits", len(discovery_hits)])
    append_excel_row(ws, ["Companies", len(companies)])
    append_excel_row(ws, ["Trigger events", len(trigger_events)])
    append_excel_row(ws, ["LinkedIn company URLs", sum(bool(record.linkedin.company_url) for record in companies.values())])
    append_excel_row(ws, ["LinkedIn contact targets", sum(record.linkedin.contact_status != "Not targeted" for record in companies.values())])
    append_excel_row(ws, ["LinkedIn contact sets complete", sum(record.linkedin.contact_status.startswith("Complete") for record in companies.values())])
    append_excel_row(ws, ["LinkedIn contact sets partial", sum(record.linkedin.contact_status.startswith("Partial") for record in companies.values())])
    style_sheet(ws)
    size_columns(ws, {"A": 28, "B": 24})

    ws = wb.create_sheet("Pipeline Run Log")
    append_excel_row(ws, ["Source", "Source type", "URL", "Status", "Result"])
    for row in run_log:
        append_excel_row(ws, row)
    style_sheet(ws)
    add_hyperlinks(ws, [3])
    size_columns(ws, {"A": 26, "B": 18, "C": 52, "D": 16, "E": 42})

    ws = wb.create_sheet("Source Inventory")
    append_excel_row(ws, ["Source name", "Type", "URL", "Geography", "Priority", "Update cadence", "Extraction method", "Signal / notes", "Automated adapter"])
    for s in SOURCES:
        append_excel_row(ws, [s.name, s.source_type, s.url, s.geography, s.priority, s.update_cadence, s.extraction_method, s.notes, adapter_inventory_label(s)])
    style_sheet(ws)
    add_hyperlinks(ws, [3])
    size_columns(ws, {"A": 30, "B": 20, "C": 54, "H": 58, "I": 22})

    ws = wb.create_sheet("Source Playbooks")
    append_excel_row(ws, ["Source type", "Search/query terms", "Filters", "Output fields", "Expected evidence", "Best-fit BBT wedge"])
    for row in SOURCE_PLAYBOOKS:
        append_excel_row(ws, row)
    style_sheet(ws)
    size_columns(ws, {"A": 22, "B": 56, "C": 44, "D": 44, "E": 44, "F": 22})

    ws = wb.create_sheet("Discovery Hits")
    append_excel_row(ws, ["Company", "Discovery source", "Source type", "Discovery evidence URL", "Article year", "Discovery rationale", "Matched terms", "Website", "Geography", "Product type", "Accelerator program", "Cohort label", "Cohort year", "Category / track", "Company description", "Captured at"])
    for hit in discovery_hits:
        append_excel_row(ws, [
            hit.company,
            hit.source_name,
            hit.source_type,
            hit.discovery_url,
            hit.article_year,
            hit.discovery_rationale,
            hit.matched_terms,
            hit.website,
            hit.geography,
            hit.product_type,
            hit.accelerator_program,
            hit.cohort_label,
            hit.cohort_year,
            hit.category_or_track,
            hit.company_description,
            hit.captured_at,
        ])
    style_sheet(ws)
    add_hyperlinks(ws, [4, 8])
    size_columns(ws, {"A": 24, "B": 28, "C": 18, "D": 54, "E": 14, "F": 56, "G": 24, "H": 36, "J": 28, "K": 26, "L": 24, "N": 28, "O": 58})

    ws = wb.create_sheet("Lead Intake")
    append_excel_row(ws, ["Company", "Website", "Geography", "Product type", "Accelerator program", "Cohort label", "Cohort year", "Category / track", "Company description", "Discovery source", "Discovery source type", "Discovery evidence URL", "Article year", "Discovery rationale", "Primary trigger event", "Trigger source", "Trigger evidence URL", "Evidence status", "Date captured"])
    for record in sorted(companies.values(), key=lambda r: r.company):
        hit = primary_discovery(record)
        trigger = primary_trigger(record)
        append_excel_row(ws, [
            record.company,
            record.website,
            record.geography,
            record.product_type,
            hit.accelerator_program,
            hit.cohort_label,
            hit.cohort_year,
            hit.category_or_track,
            hit.company_description,
            hit.source_name,
            hit.source_type,
            hit.discovery_url,
            hit.article_year,
            hit.discovery_rationale,
            trigger.trigger_event if trigger else "",
            trigger.trigger_source if trigger else "",
            trigger.evidence_url if trigger else "",
            "Verified trigger" if trigger else "Discovered only",
            TODAY,
        ])
    style_sheet(ws)
    add_hyperlinks(ws, [2, 12, 17])
    size_columns(ws, {"A": 24, "B": 36, "C": 16, "D": 28, "E": 26, "F": 24, "H": 28, "I": 58, "J": 28, "L": 54, "M": 14, "N": 58, "O": 58, "P": 28, "Q": 54})

    ws = wb.create_sheet("Trigger Log")
    append_excel_row(ws, ["Company", "Trigger type", "Trigger event", "Trigger source", "Evidence URL", "Trigger role", "Captured at"])
    for event in trigger_events:
        append_excel_row(ws, [event.company, event.trigger_type, event.trigger_event, event.trigger_source, event.evidence_url, event.trigger_role, event.captured_at])
    style_sheet(ws)
    add_hyperlinks(ws, [5])
    size_columns(ws, {"A": 24, "B": 20, "C": 64, "D": 30, "E": 54, "F": 16})

    ws = wb.create_sheet("Lead Filtering")
    headers = [
        "Company", "Evidence year", "Evidence basis", "Evidence recency", "Trigger type", "Geography",
        "Company type", "Company stage", "Product area", "Hiring signal", "Funding stage",
        "Persona", "Primary BBT quadrant", "Secondary tag", "Pain hypothesis",
        "Value prop", "Outreach angle", "Classification confidence", "Classification method", "LLM used", "Fallback reason",
        "Legacy: Recently funded +3", "Legacy: AI/SaMD/device +3", "Legacy: Hiring QA/reg/V&V +3", "Legacy: Clinical validation +2",
        "FDA/CE/reg language +2", "Grant/public funding +2", "University/grant origin +2",
        "No obvious reg team +2", "Pre-commercial +1", "Large company -1", "Wellness/non-medical -2",
        "Pharma-only -2", "Legacy score", "Legacy priority band", "Evidence status", "Primary evidence URL", "Website",
        "LinkedIn company URL", "LinkedIn company status",
        "Executive contact name", "Executive contact title", "Executive LinkedIn URL",
        "Technical/R&D contact name", "Technical/R&D contact title", "Technical/R&D LinkedIn URL",
        "Quality/QA contact name", "Quality/QA contact title", "Quality/QA LinkedIn URL",
        "LinkedIn contact status",
    ]
    append_excel_row(ws, headers)
    scoring_rows = []
    for record in sorted(companies.values(), key=lambda r: r.company):
        flags, score, band = score_company_metrics(record)
        enrichment = classify_company(record)
        filter_fields = lead_filter_fields(record, enrichment)
        trigger = primary_trigger(record)
        hit = primary_discovery(record)
        evidence_url = trigger.evidence_url if trigger else hit.discovery_url
        status = "Verified trigger" if trigger else "Discovered only"
        append_excel_row(ws, [
            record.company,
            *filter_fields.values(),
            enrichment.persona,
            enrichment.primary_quadrant,
            enrichment.secondary_tag,
            enrichment.pain_hypothesis,
            enrichment.value_prop,
            enrichment.outreach_angle,
            enrichment.confidence,
            enrichment.method,
            "Yes" if enrichment.llm_used else "No",
            enrichment.fallback_reason,
            *flags.values(), score, band, status, evidence_url, record.website,
            record.linkedin.company_url,
            record.linkedin.company_status,
            *linkedin_contact_values(record.linkedin.executive),
            *linkedin_contact_values(record.linkedin.technical),
            *linkedin_contact_values(record.linkedin.quality),
            record.linkedin.contact_status,
        ])
        scoring_rows.append((score, band, status, record, enrichment, evidence_url, trigger))
    style_sheet(ws)
    style_linkedin_columns(ws, 39)
    add_hyperlinks(ws, [37, 38, 39, 43, 46, 49])
    size_columns(ws, {
        "A": 24, "B": 14, "C": 16, "D": 16, "E": 22, "F": 16, "G": 24, "H": 24,
        "I": 24, "J": 14, "K": 28, "L": 28, "M": 22, "N": 22, "O": 58, "P": 54,
        "Q": 54, "R": 14, "S": 18, "T": 12, "U": 22, "AH": 12, "AI": 18, "AJ": 18,
        "AK": 54, "AL": 36,
        "AM": 42, "AN": 24, "AO": 25, "AP": 38, "AQ": 42,
        "AR": 25, "AS": 38, "AT": 42, "AU": 25, "AV": 38, "AW": 42, "AX": 26,
    })

    scoring_rows.sort(key=lambda row: (row[2] == "Verified trigger", row[0]), reverse=True)
    ws = wb.create_sheet("Weekly Review")
    append_excel_row(ws, ["Rank", "Company", "Score", "Priority band", "Evidence status", "Why shortlisted", "Next action", "Owner", "Status", "Review notes", "Evidence URL"])
    for idx, (score, band, status, record, enrichment, evidence_url, trigger) in enumerate(scoring_rows[:15], start=1):
        append_excel_row(ws, [
            idx,
            record.company,
            score,
            band,
            status,
            f"{enrichment.pain_hypothesis} Outreach: {enrichment.outreach_angle}",
            next_action_for_lead(status, enrichment, trigger),
            "",
            "Validate",
            "",
            evidence_url,
        ])
    style_sheet(ws)
    add_hyperlinks(ws, [11])
    size_columns(ws, {"A": 8, "B": 24, "C": 10, "D": 14, "E": 18, "F": 62, "G": 36, "K": 54})

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    try:
        wb.save(OUT)
        return OUT
    except PermissionError:
        fallback = OUT.with_name("BlueBridge_TOFU_BizDev_V1_pipeline.xlsx")
        wb.save(fallback)
        return fallback

