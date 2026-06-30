from __future__ import annotations

import json
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from bbt_bizdev.adapters import linkedin as li


INPUT = ROOT / "BlueBridge_TOFU_BizDev_V1.xlsx"
OUTPUT = Path(__file__).with_name("official_pilot_enrichment.json")
USER_AGENT = "Mozilla/5.0 (compatible; BBT-bizdev-pipeline/1.0)"


def fetch(url: str) -> tuple[str, str | None]:
    try:
        request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "text/html,application/xhtml+xml"})
        return urlopen(request, timeout=18).read().decode("utf-8", "ignore"), None
    except (OSError, HTTPError, URLError, ValueError) as exc:
        return "", str(exc)


def read_pilot() -> list[dict]:
    workbook = load_workbook(INPUT, read_only=True, data_only=True)
    sheet = workbook["Lead Filtering"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    index = {name: headers.index(name) for name in ("Company", "Evidence year", "Website", "Legacy priority band")}
    leads = []
    for row_number, row in enumerate(rows, start=2):
        year = str(row[index["Evidence year"]] or "").strip()
        priority = str(row[index["Legacy priority band"]] or "").strip()
        website = str(row[index["Website"]] or "").strip()
        if year == "2026" and priority in {"Strong", "Good"} and website.startswith(("http://", "https://")):
            leads.append({"row": row_number, "company": str(row[index["Company"]] or "").strip(), "website": website})
    return leads


def contact_dict(contact):
    return None if contact is None else {"name": contact.name, "title": contact.title, "url": contact.url}


def enrich(lead: dict) -> dict:
    result = {
        **lead,
        "company_url": "",
        "company_status": "Not found",
        "executive": None,
        "technical": None,
        "quality": None,
        "contact_status": "No verified matches",
    }
    raw_html, error = fetch(lead["website"])
    if error:
        result["company_status"] = "Search error"
        result["contact_status"] = "Search error"
        return result
    observations = li.extract_page_links(raw_html, lead["website"])
    result["company_url"] = li._company_url_from_observations(observations)
    if result["company_url"]:
        result["company_status"] = "Found - official website"
    for team_url in li._team_page_urls(observations, lead["website"]):
        team_html, _ = fetch(team_url)
        if team_html:
            observations.extend(li.extract_page_links(team_html, team_url))
    candidates = []
    for observation in observations:
        candidate = li.contact_from_official_observation(observation, lead["company"], lead["website"])
        if candidate:
            candidates.append(candidate)
    executive, technical, quality = li.select_contacts(candidates)
    result["executive"] = contact_dict(executive)
    result["technical"] = contact_dict(technical)
    result["quality"] = contact_dict(quality)
    count = sum(contact is not None for contact in (executive, technical, quality))
    result["contact_status"] = "Complete - 3 verified" if count == 3 else f"Partial - {count}/3 verified" if count else "No verified matches"
    return result


def main() -> None:
    leads = read_pilot()
    results = []
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = [executor.submit(enrich, lead) for lead in leads]
        for future in as_completed(futures):
            results.append(future.result())
    results.sort(key=lambda item: item["row"])
    OUTPUT.write_text(json.dumps(results, indent=2), encoding="utf-8")
    complete = sum(item["contact_status"].startswith("Complete") for item in results)
    partial = sum(item["contact_status"].startswith("Partial") for item in results)
    print(f"wrote {OUTPUT} companies={len(results)} complete={complete} partial={partial}")


if __name__ == "__main__":
    main()
