from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
INPUT = HERE / "BlueBridge_TOFU_BizDev_V1_LinkedIn.xlsx"
RESULTS = HERE / "official_pilot_enrichment.json"
OUTPUT = HERE / "BlueBridge_TOFU_BizDev_V1_LinkedIn_Priority_Pilot.xlsx"


def set_url(cell, value: str) -> None:
    cell.value = value
    if value:
        cell.hyperlink = value
        cell.style = "Hyperlink"


def main() -> None:
    results = json.loads(RESULTS.read_text(encoding="utf-8"))
    workbook = load_workbook(INPUT)
    sheet = workbook["Lead Filtering"]
    headers = [cell.value for cell in sheet[1]]
    index = {header: headers.index(header) + 1 for header in headers}
    rows = {str(sheet.cell(row, index["Company"]).value or "").strip(): row for row in range(2, sheet.max_row + 1)}

    contact_fields = [
        ("executive", "Executive contact name", "Executive contact title", "Executive LinkedIn URL"),
        ("technical", "Technical/R&D contact name", "Technical/R&D contact title", "Technical/R&D LinkedIn URL"),
        ("quality", "Quality/QA contact name", "Quality/QA contact title", "Quality/QA LinkedIn URL"),
    ]
    for result in results:
        row = rows[result["company"]]
        if result["company_url"]:
            set_url(sheet.cell(row, index["LinkedIn company URL"]), result["company_url"])
        sheet.cell(row, index["LinkedIn company status"], result["company_status"])
        for key, name_header, title_header, url_header in contact_fields:
            contact = result.get(key)
            sheet.cell(row, index[name_header], contact["name"] if contact else "")
            sheet.cell(row, index[title_header], contact["title"] if contact else "")
            set_url(sheet.cell(row, index[url_header]), contact["url"] if contact else "")
        sheet.cell(row, index["LinkedIn contact status"], result["contact_status"])

    complete = sum(item["contact_status"].startswith("Complete") for item in results)
    partial = sum(item["contact_status"].startswith("Partial") for item in results)
    company_urls = sum(bool(item["company_url"]) for item in results)
    summary = workbook["Pipeline Summary"]
    summary.append(["Priority pilot official sites checked", len(results)])
    summary.append(["Priority pilot company URLs found", company_urls])
    summary.append(["Priority pilot contact sets complete", complete])
    summary.append(["Priority pilot contact sets partial", partial])
    run_log = workbook["Pipeline Run Log"]
    run_log.append([
        "LinkedIn priority pilot", "Enrichment", "Official company websites",
        "Completed - official sites", f"23 checked; {company_urls} company URLs; {complete} complete contact sets; {partial} partial",
    ])
    workbook.save(OUTPUT)
    print(f"{OUTPUT} rows={sheet.max_row - 1} complete={complete} partial={partial} company_urls={company_urls}")


if __name__ == "__main__":
    main()
