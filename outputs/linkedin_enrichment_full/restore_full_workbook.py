from __future__ import annotations

import re
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "BlueBridge_TOFU_BizDev_V1.xlsx"
OUTPUT = Path(__file__).with_name("BlueBridge_TOFU_BizDev_V1_LinkedIn.xlsx")

HEADERS = [
    "LinkedIn company URL", "LinkedIn company status",
    "Executive contact name", "Executive contact title", "Executive LinkedIn URL",
    "Technical/R&D contact name", "Technical/R&D contact title", "Technical/R&D LinkedIn URL",
    "Quality/QA contact name", "Quality/QA contact title", "Quality/QA LinkedIn URL",
    "LinkedIn contact status",
]


def canonical_company_url(value) -> str:
    if not isinstance(value, str):
        return ""
    match = re.search(r"https?://(?:[a-z]{2}\.)?(?:www\.)?linkedin\.com/company/([^/?#]+)", value, flags=re.I)
    return f"https://www.linkedin.com/company/{match.group(1).lower()}" if match else ""


def local_company_links(workbook) -> dict[str, str]:
    links: dict[str, str] = {}
    for sheet in workbook.worksheets:
        headers = [cell.value for cell in sheet[1]]
        if "Company" not in headers:
            continue
        company_col = headers.index("Company") + 1
        for row in range(2, sheet.max_row + 1):
            company = str(sheet.cell(row, company_col).value or "").strip()
            if not company or company in links:
                continue
            for col in range(1, sheet.max_column + 1):
                url = canonical_company_url(sheet.cell(row, col).value)
                if url:
                    links[company] = url
                    break
    links["Qure.ai"] = "https://www.linkedin.com/company/qure.ai"
    return links


def set_url(cell, value: str) -> None:
    cell.value = value
    if value:
        cell.hyperlink = value
        cell.style = "Hyperlink"


def main() -> None:
    workbook = load_workbook(INPUT)
    links = local_company_links(workbook)
    sheet = workbook["Lead Filtering"]
    original_headers = [cell.value for cell in sheet[1]]
    company_col = original_headers.index("Company") + 1
    year_col = original_headers.index("Evidence year") + 1
    start_col = len(original_headers) + 1

    for offset, header in enumerate(HEADERS):
        sheet.cell(1, start_col + offset, header)

    teal = PatternFill("solid", fgColor="0F6B78")
    thin = Side(style="thin", color="D9E2F3")
    widths = [42, 24, 25, 38, 42, 25, 38, 42, 25, 38, 42, 26]
    for offset, width in enumerate(widths):
        cell = sheet.cell(1, start_col + offset)
        cell.fill = teal
        cell.font = Font(name="Aptos", size=10, color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[cell.column_letter].width = width

    target_count = 0
    complete_count = 0
    for row in range(2, sheet.max_row + 1):
        company = str(sheet.cell(row, company_col).value or "").strip()
        year = str(sheet.cell(row, year_col).value or "").strip()
        company_url = links.get(company, "")
        values = [
            company_url,
            "Found - existing/local evidence" if company_url else "Pending external research",
            "", "", "", "", "", "", "", "", "",
            "Not targeted" if year != "2026" else "Pending external research",
        ]
        if year == "2026":
            target_count += 1
        if company == "Qure.ai":
            values = [
                "https://www.linkedin.com/company/qure.ai", "Found - official website",
                "Amit Kakar MD", "Founder & CEO", "https://www.linkedin.com/in/amit-kakar-md-32062165",
                "Pradeep Kumar", "Chief Technology Officer", "https://www.linkedin.com/in/pradeep-kumar-356774b5",
                "Bunty Kundnani", "Chief Regulatory Affairs Officer", "https://www.linkedin.com/in/bunty-kundnani-4379b8142",
                "Complete - 3 verified",
            ]
            complete_count += 1
        for offset, value in enumerate(values):
            cell = sheet.cell(row, start_col + offset)
            if offset in {0, 4, 7, 10}:
                set_url(cell, value)
            else:
                cell.value = value
            cell.font = cell.font.copy(name="Aptos", size=10)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.auto_filter.ref = f"A1:{sheet.cell(1, start_col + len(HEADERS) - 1).column_letter}{sheet.max_row}"

    summary = workbook["Pipeline Summary"]
    for metric, value in [
        ("LinkedIn company URLs (local)", len(links)),
        ("LinkedIn contact targets", target_count),
        ("LinkedIn contact sets complete", complete_count),
        ("LinkedIn research status", "Bulk external lookup requires explicit approval"),
    ]:
        summary.append([metric, value])

    run_log = workbook["Pipeline Run Log"]
    run_log.append([
        "LinkedIn enrichment", "Enrichment", "https://html.duckduckgo.com/html/",
        "Pending approval", "Full 3,014-row dataset preserved; bulk external lookup not run",
    ])
    set_url(run_log.cell(run_log.max_row, 3), run_log.cell(run_log.max_row, 3).value)

    workbook.save(OUTPUT)
    print(f"{OUTPUT} rows={sheet.max_row - 1} target_2026={target_count} local_company_links={len(links)}")


if __name__ == "__main__":
    main()
