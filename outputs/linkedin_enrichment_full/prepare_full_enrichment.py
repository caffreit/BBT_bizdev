from __future__ import annotations

import json
import re
import threading
import time
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import quote_plus
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from bbt_bizdev.adapters import linkedin as li
from bbt_bizdev.models import CompanyRecord


ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "BlueBridge_TOFU_BizDev_V1.xlsx"
OUTPUT = Path(__file__).with_name("full_enrichment.json")
CHECKPOINT = Path(__file__).with_name("full_enrichment.checkpoint.json")
USER_AGENT = "Mozilla/5.0 (compatible; BBT-bizdev-pipeline/1.0)"


def read_leads() -> list[dict]:
    workbook = load_workbook(INPUT, read_only=True, data_only=True)
    sheet = workbook["Lead Filtering"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    indexes = {name: headers.index(name) for name in ("Company", "Evidence year", "Website", "Legacy priority band")}
    leads = [
        {
            "row": row_number,
            "company": str(row[indexes["Company"]] or "").strip(),
            "year": str(row[indexes["Evidence year"]] or "").strip(),
            "website": str(row[indexes["Website"]] or "").strip(),
            "priority": str(row[indexes["Legacy priority band"]] or "").strip(),
        }
        for row_number, row in enumerate(rows, start=2)
    ]
    return [lead for lead in leads if lead["year"] == "2026" and lead["priority"] in {"Strong", "Good"}]


def fetch(url: str, timeout: int = 16) -> tuple[str, str | None]:
    try:
        request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "text/html,application/xhtml+xml"})
        return urlopen(request, timeout=timeout).read().decode("utf-8", "ignore"), None
    except (OSError, HTTPError, URLError, ValueError) as exc:
        return "", str(exc)


def contact_dict(contact):
    return None if contact is None else {
        "name": contact.name,
        "title": contact.title,
        "url": contact.url,
    }


def official_enrichment(lead: dict) -> dict:
    result = {
        **lead,
        "company_url": "",
        "company_status": "Not found",
        "candidates": [],
        "official_error": "",
    }
    website = lead["website"]
    if not website.startswith(("http://", "https://")):
        return result
    raw_html, error = fetch(website)
    if error:
        result["official_error"] = error
        return result
    observations = li.extract_page_links(raw_html, website)
    result["company_url"] = li._company_url_from_observations(observations)
    if result["company_url"]:
        result["company_status"] = "Found - official website"
    if lead["year"] == "2026":
        for team_url in li._team_page_urls(observations, website):
            team_html, _ = fetch(team_url)
            if team_html:
                observations.extend(li.extract_page_links(team_html, team_url))
        candidates = []
        for observation in observations:
            candidate = li.contact_from_official_observation(observation, lead["company"], website)
            if candidate:
                candidates.append(candidate)
        result["candidates"] = candidates
    return result


_search_lock = threading.Lock()
_last_search = 0.0


def bing_search(query: str) -> tuple[list[li.PublicSearchHit], str | None]:
    global _last_search
    with _search_lock:
        delay = 0.22 - (time.monotonic() - _last_search)
        if delay > 0:
            time.sleep(delay)
        _last_search = time.monotonic()
    url = "https://www.bing.com/search?format=rss&q=" + quote_plus(query)
    last_error = ""
    for attempt in range(2):
        raw, error = fetch(url, timeout=18)
        if not error:
            try:
                root = ET.fromstring(raw)
                hits = [
                    li.PublicSearchHit(
                        title=li.clean_space(item.findtext("title", "")),
                        url=li.clean_space(item.findtext("link", "")),
                        snippet=li.clean_space(item.findtext("description", "")),
                    )
                    for item in root.findall(".//item")
                ]
                return hits, None
            except ET.ParseError as exc:
                last_error = str(exc)
        else:
            last_error = error
        time.sleep(1 + attempt)
    return [], last_error or "Search failed"


def find_company_url(item: dict) -> tuple[str, str, str]:
    if item["company_url"]:
        return item["company"], item["company_url"], item["company_status"]
    hits, error = bing_search(f'site:linkedin.com/company "{item["company"]}"')
    for hit in hits:
        url = li.canonicalize_linkedin_url(hit.url, "company")
        if url and li.company_name_matches(item["company"], hit.title, item["website"]):
            return item["company"], url, "Found - public search"
    return item["company"], "", "Search error" if error else "Not found"


def find_contacts(item: dict) -> tuple[str, list, str | None]:
    existing = item["candidates"]
    if all(li.select_contacts(existing)):
        return item["company"], existing, None
    query = f'site:linkedin.com/in "{item["company"]}" CEO CTO quality regulatory R&D engineering'
    hits, error = bing_search(query)
    candidates = list(existing)
    for hit in hits:
        candidate = li._contact_from_search_hit(hit, item["company"], item["website"])
        if candidate:
            candidates.append(candidate)
    return item["company"], candidates, error


def write_checkpoint(items: list[dict]) -> None:
    safe = []
    for item in items:
        copy = dict(item)
        copy["candidates"] = [
            {
                "name": candidate.name,
                "title": candidate.title,
                "url": candidate.url,
                "role_bucket": candidate.role_bucket,
                "source": candidate.source,
                "confidence": candidate.confidence,
            }
            for candidate in item["candidates"]
        ]
        safe.append(copy)
    CHECKPOINT.write_text(json.dumps(safe, indent=2), encoding="utf-8")


def main() -> None:
    leads = read_leads()
    official: list[dict] = []
    with ThreadPoolExecutor(max_workers=24) as executor:
        futures = {executor.submit(official_enrichment, lead): lead for lead in leads}
        for count, future in enumerate(as_completed(futures), start=1):
            official.append(future.result())
            if count % 250 == 0:
                print(f"official {count}/{len(leads)}", flush=True)
    official.sort(key=lambda item: item["row"])
    write_checkpoint(official)

    by_company = {item["company"]: item for item in official}
    unresolved = [item for item in official if not item["company_url"]]
    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = [executor.submit(find_company_url, item) for item in unresolved]
        for count, future in enumerate(as_completed(futures), start=1):
            company, url, status = future.result()
            by_company[company]["company_url"] = url
            by_company[company]["company_status"] = status
            if count % 200 == 0:
                print(f"company search {count}/{len(unresolved)}", flush=True)

    targeted = official
    contact_errors: dict[str, str | None] = {}
    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = [executor.submit(find_contacts, item) for item in targeted]
        for count, future in enumerate(as_completed(futures), start=1):
            company, candidates, error = future.result()
            by_company[company]["candidates"] = candidates
            contact_errors[company] = error
            if count % 100 == 0:
                print(f"contact search {count}/{len(targeted)}", flush=True)

    output = []
    for item in official:
        executive = technical = quality = None
        contact_status = "Not targeted"
        if item["year"] == "2026":
            executive, technical, quality = li.select_contacts(item["candidates"])
            count = sum(contact is not None for contact in (executive, technical, quality))
            if count == 3:
                contact_status = "Complete - 3 verified"
            elif count:
                contact_status = f"Partial - {count}/3 verified"
            elif contact_errors.get(item["company"]):
                contact_status = "Search error"
            else:
                contact_status = "No verified matches"
        output.append({
            "row": item["row"],
            "company": item["company"],
            "company_url": item["company_url"],
            "company_status": item["company_status"],
            "executive": contact_dict(executive),
            "technical": contact_dict(technical),
            "quality": contact_dict(quality),
            "contact_status": contact_status,
        })
    OUTPUT.write_text(json.dumps(output, indent=2), encoding="utf-8")
    print(f"wrote {OUTPUT} rows={len(output)}", flush=True)


if __name__ == "__main__":
    main()
