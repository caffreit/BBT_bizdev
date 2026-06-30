from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
HERE = Path(__file__).resolve().parent
INPUT = ROOT / "outputs" / "linkedin_enrichment_full" / "BlueBridge_TOFU_BizDev_V1_LinkedIn.xlsx"
RESULTS = ROOT / "outputs" / "linkedin_enrichment_full" / "full_enrichment.json"
OUTPUT = HERE / "BlueBridge_TOFU_BizDev_V1_LinkedIn_67_Leads.xlsx"


def set_url(cell, value: str) -> None:
    cell.value = value
    cell.hyperlink = value or None
    if value:
        cell.style = "Hyperlink"


def main() -> None:
    results = json.loads(RESULTS.read_text(encoding="utf-8"))
    if len(results) != 67:
        raise RuntimeError(f"Expected exactly 67 lead results, got {len(results)}")

    workbook = load_workbook(INPUT)
    sheet = workbook["Lead Filtering"]
    headers = [cell.value for cell in sheet[1]]
    index = {header: headers.index(header) + 1 for header in headers}
    contact_fields = [
        ("executive", "Executive contact name", "Executive contact title", "Executive LinkedIn URL"),
        ("technical", "Technical/R&D contact name", "Technical/R&D contact title", "Technical/R&D LinkedIn URL"),
        ("quality", "Quality/QA contact name", "Quality/QA contact title", "Quality/QA LinkedIn URL"),
    ]

    for result in results:
        row = result["row"]
        set_url(sheet.cell(row, index["LinkedIn company URL"]), result["company_url"])
        sheet.cell(row, index["LinkedIn company status"], result["company_status"])
        for key, name_header, title_header, url_header in contact_fields:
            contact = result.get(key)
            sheet.cell(row, index[name_header], contact["name"] if contact else "")
            sheet.cell(row, index[title_header], contact["title"] if contact else "")
            set_url(sheet.cell(row, index[url_header]), contact["url"] if contact else "")
        sheet.cell(row, index["LinkedIn contact status"], result["contact_status"])

    # Keep the filter controls, but clear any applied criteria so all rows open visibly.
    sheet.auto_filter.ref = f"A1:AX{sheet.max_row}"
    sheet.auto_filter.filterColumn = []

    company_urls = sum(bool(row["company_url"]) for row in results)
    complete = sum(row["contact_status"].startswith("Complete") for row in results)
    partial = sum(row["contact_status"].startswith("Partial") for row in results)
    summary = workbook["Pipeline Summary"]
    summary.append(["LinkedIn enriched lead scope", len(results)])
    summary.append(["LinkedIn company URLs found in 67-lead scope", company_urls])
    summary.append(["LinkedIn contact sets complete in 67-lead scope", complete])
    summary.append(["LinkedIn contact sets partial in 67-lead scope", partial])
    run_log = workbook["Pipeline Run Log"]
    run_log.append([
        "LinkedIn 67-lead enrichment",
        "Enrichment",
        "Public search + official company websites",
        "Completed",
        f"67 checked; {company_urls} company URLs; {complete} complete contact sets; {partial} partial",
    ])

    workbook.save(OUTPUT)
    print(f"wrote {OUTPUT} rows=67 urls={company_urls} complete={complete} partial={partial}")


if __name__ == "__main__":
    main()
