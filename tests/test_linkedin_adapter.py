import tempfile
import unittest
import json
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

import build_bbt_bizdev_workbook as pipeline


DDG_FIXTURE = """
<div class="result">
  <a class="result__a" href="//duckduckgo.com/l/?uddg=https%3A%2F%2Fwww.linkedin.com%2Fcompany%2Fnovascan-health%2F">NovaScan Health | LinkedIn</a>
  <a class="result__snippet">NovaScan Health develops medical imaging software.</a>
</div>
<div class="result">
  <a class="result__a" href="https://www.linkedin.com/in/alice-founder/">Alice Founder - Chief Executive Officer at NovaScan Health | LinkedIn</a>
  <a class="result__snippet">Alice leads NovaScan Health.</a>
</div>
"""


class LinkedInAdapterTests(unittest.TestCase):
    def test_brave_search_parses_structured_results(self):
        class Response:
            def read(self):
                return json.dumps({
                    "web": {
                        "results": [
                            {
                                "title": "NovaScan Health | LinkedIn",
                                "url": "https://www.linkedin.com/company/novascan-health/",
                                "description": "NovaScan Health develops medical imaging software.",
                            }
                        ]
                    }
                }).encode("utf-8")

        with patch.dict("os.environ", {"BRAVE_SEARCH_API_KEY": "token"}), patch.object(pipeline, "urlopen", return_value=Response()):
            hits, error = pipeline.brave_search("NovaScan Health")
        self.assertIsNone(error)
        self.assertEqual(hits[0].url, "https://www.linkedin.com/company/novascan-health/")

    def test_canonicalization_and_search_result_parsing(self):
        self.assertEqual(
            pipeline.canonicalize_linkedin_url("https://ie.linkedin.com/company/NovaScan-Health/?trk=abc", "company"),
            "https://www.linkedin.com/company/novascan-health",
        )
        self.assertEqual(
            pipeline.canonicalize_linkedin_url("https://www.linkedin.com/in/Alice-Founder/details/experience/", "person"),
            "https://www.linkedin.com/in/alice-founder",
        )
        self.assertEqual(pipeline.canonicalize_linkedin_url("https://www.linkedin.com/jobs/view/123"), "")

        hits = pipeline.parse_duckduckgo_results(DDG_FIXTURE)
        self.assertEqual(len(hits), 2)
        self.assertEqual(hits[0].url, "https://www.linkedin.com/company/novascan-health/")

    def test_ambiguous_company_names_require_stronger_title_match(self):
        self.assertTrue(pipeline.company_name_matches("AMA", "AMA | LinkedIn", "https://ama.example"))
        self.assertFalse(
            pipeline.company_name_matches(
                "AMA",
                "Amy Smith - Chief Executive Officer at AMA | LinkedIn",
                "https://ama.example",
            )
        )

    def test_contact_selection_prioritizes_roles_and_deduplicates(self):
        candidates = [
            pipeline.LinkedInContact("Alice Founder", "Chief Executive Officer", "https://www.linkedin.com/in/alice", "executive", "fixture", 0.9),
            pipeline.LinkedInContact("Terry Tech", "VP Engineering and R&D", "https://www.linkedin.com/in/terry", "technical", "fixture", 0.9),
            pipeline.LinkedInContact("Quinn Quality", "Director of Quality Assurance", "https://www.linkedin.com/in/quinn", "quality", "fixture", 0.9),
            pipeline.LinkedInContact("Alice Founder", "CEO", "https://www.linkedin.com/in/alice", "executive", "fixture", 0.7),
        ]
        executive, technical, quality = pipeline.select_contacts(candidates)
        self.assertEqual(executive.name, "Alice Founder")
        self.assertEqual(technical.name, "Terry Tech")
        self.assertEqual(quality.name, "Quinn Quality")
        self.assertEqual(len({executive.url, technical.url, quality.url}), 3)

    def test_contact_selection_does_not_use_senior_fallback(self):
        candidates = [
            pipeline.LinkedInContact("Alice Founder", "Chief Executive Officer", "https://www.linkedin.com/in/alice", "executive", "fixture", 0.9),
            pipeline.LinkedInContact("Morgan Ops", "Vice President Operations", "https://www.linkedin.com/in/morgan", "senior", "fixture", 0.9),
        ]
        executive, technical, quality = pipeline.select_contacts(candidates)
        self.assertEqual(executive.name, "Alice Founder")
        self.assertIsNone(technical)
        self.assertIsNone(quality)

    def test_public_search_rejects_company_mentions_without_current_employer_evidence(self):
        record = pipeline.CompanyRecord(company="NovaScan Health")

        def search(query):
            if "/company" in query:
                return [pipeline.PublicSearchHit("NovaScan Health | LinkedIn", "https://www.linkedin.com/company/novascan-health/")], None
            return [
                pipeline.PublicSearchHit(
                    "Alice Founder - Chief Executive Officer | LinkedIn",
                    "https://www.linkedin.com/in/alice-founder/",
                    "Alice advises NovaScan Health and other startups.",
                )
            ], None

        result = pipeline.enrich_company_linkedin(record, True, search_fn=search, fetch_fn=lambda _url: ("", None))
        self.assertEqual(result.company_url, "https://www.linkedin.com/company/novascan-health")
        self.assertEqual(result.contact_status, "No verified matches")
        self.assertIsNone(result.executive)

    def test_public_search_accepts_current_company_at_phrase(self):
        record = pipeline.CompanyRecord(company="NovaScan Health")

        def search(query):
            if "/company" in query:
                return [pipeline.PublicSearchHit("NovaScan Health | LinkedIn", "https://www.linkedin.com/company/novascan-health/")], None
            return [
                pipeline.PublicSearchHit(
                    "Alice Founder - Chief Executive Officer at NovaScan Health | LinkedIn",
                    "https://www.linkedin.com/in/alice-founder/",
                    "",
                )
            ], None

        result = pipeline.enrich_company_linkedin(record, True, search_fn=search, fetch_fn=lambda _url: ("", None))
        self.assertEqual(result.contact_status, "Partial - 1/3 verified")
        self.assertEqual(result.executive.name, "Alice Founder")
        self.assertEqual(result.executive.verification_status, "Verified - current-company search evidence")

    def test_official_site_enrichment_finds_company_and_three_people(self):
        home = """
        <a href="https://www.linkedin.com/company/novascan-health/">LinkedIn</a>
        <a href="/team">Our Team</a>
        """
        team = """
        <h3>Alice Founder</h3><p>Chief Executive Officer</p><a href="https://www.linkedin.com/in/alice-founder/">Alice Founder</a>
        <h3>Terry Tech</h3><p>VP Engineering and R&amp;D</p><a href="https://www.linkedin.com/in/terry-tech/">Terry Tech</a>
        <h3>Quinn Quality</h3><p>Director of Quality Assurance</p><a href="https://www.linkedin.com/in/quinn-quality/">Quinn Quality</a>
        """

        def fetch(url):
            return (team, None) if url.endswith("/team") else (home, None)

        def no_search(_query):
            self.fail("Public search should not run when official links are complete")

        record = pipeline.CompanyRecord(company="NovaScan Health", website="https://novascan.example")
        result = pipeline.enrich_company_linkedin(record, True, search_fn=no_search, fetch_fn=fetch)
        self.assertEqual(result.company_url, "https://www.linkedin.com/company/novascan-health")
        self.assertEqual(result.company_status, "Found - official website")
        self.assertEqual(result.contact_status, "Complete - 3 verified")
        self.assertEqual(result.executive.name, "Alice Founder")
        self.assertEqual(result.technical.name, "Terry Tech")
        self.assertEqual(result.quality.name, "Quinn Quality")

    def test_trusted_source_page_can_resolve_missing_company_website(self):
        source = """
        <h1>NovaScan Health</h1><a href="https://novascan.example">Visit NovaScan Health website</a>
        """
        home = """
        <a href="https://www.linkedin.com/company/novascan-health/">LinkedIn</a>
        <a href="/team">Team</a>
        """
        team = """
        <h3>Alice Founder</h3><p>Chief Executive Officer</p><a href="https://www.linkedin.com/in/alice-founder/">Alice Founder</a>
        """

        def fetch(url):
            if url == "https://trusted.example/profile":
                return source, None
            if url == "https://novascan.example":
                return home, None
            if url == "https://novascan.example/team":
                return team, None
            return "", "not found"

        record = pipeline.CompanyRecord(
            company="NovaScan Health",
            discovery_hits=[pipeline.DiscoveryHit("NovaScan Health", "Trusted", "Accelerator", "https://trusted.example/profile", "Fixture")],
        )
        result = pipeline.enrich_company_linkedin(record, True, search_fn=lambda _query: ([], "blocked"), fetch_fn=fetch)
        self.assertEqual(result.resolved_website, "https://novascan.example")
        self.assertEqual(result.company_url, "https://www.linkedin.com/company/novascan-health")
        self.assertEqual(result.executive.name, "Alice Founder")

    def test_common_about_paths_are_crawled_for_people(self):
        home = '<a href="https://www.linkedin.com/company/novascan-health/">LinkedIn</a>'
        about = """
        <h3>Terry Tech</h3><p>VP Engineering and R&amp;D</p><a href="https://www.linkedin.com/in/terry-tech/">Terry Tech</a>
        """

        def fetch(url):
            if url == "https://novascan.example":
                return home, None
            if url == "https://novascan.example/about":
                return about, None
            return "", "not found"

        record = pipeline.CompanyRecord(company="NovaScan Health", website="https://novascan.example")
        result = pipeline.enrich_company_linkedin(record, True, search_fn=lambda _query: ([], "blocked"), fetch_fn=fetch)
        self.assertEqual(result.technical.name, "Terry Tech")

    def test_icon_only_profile_link_uses_slug_and_nearby_title(self):
        html = """
        <section><h3>Bunty Kundnani</h3><p>Chief Regulatory Affairs Officer</p>
        <a href="https://www.linkedin.com/in/bunty-kundnani-4379b8142"><span>in</span></a></section>
        """
        observations = pipeline.extract_page_links(html, "https://company.example/team")
        candidate = pipeline.contact_from_official_observation(observations[0], "Company", "https://company.example")
        self.assertEqual(candidate.name, "Bunty Kundnani")
        self.assertEqual(candidate.title, "Chief Regulatory Affairs Officer")
        self.assertEqual(candidate.role_bucket, "quality")
        self.assertEqual(candidate.verification_status, "Verified - official company site")

    def test_generic_profile_link_text_uses_slug_name(self):
        html = """
        <section><h3>David McGuire</h3><p>Chief Operating Officer</p>
        <a href="https://www.linkedin.com/in/david-mcguire-73b53490">Connect with me</a></section>
        """
        observations = pipeline.extract_page_links(html, "https://company.example/team")
        candidate = pipeline.contact_from_official_observation(observations[0], "Company", "https://company.example")
        self.assertEqual(candidate.name, "David McGuire")
        self.assertEqual(candidate.title, "Chief Operating Officer")

    def test_official_profile_name_mismatch_is_rejected(self):
        html = """
        <section><h3>Marie Therese Maher</h3><p>Founder</p>
        <a href="https://www.linkedin.com/in/barrymccanngalway">LinkedIn</a></section>
        """
        observations = pipeline.extract_page_links(html, "https://company.example/team")
        candidate = pipeline.contact_from_official_observation(observations[0], "Company", "https://company.example")
        self.assertIsNone(candidate)

    def test_network_errors_are_flagged_without_raising(self):
        record = pipeline.CompanyRecord(company="NovaScan Health", website="https://novascan.example")
        result = pipeline.enrich_company_linkedin(
            record,
            True,
            search_fn=lambda _query: ([], "rate limited"),
            fetch_fn=lambda _url: ("", "blocked"),
        )
        self.assertEqual(result.company_status, "Search error")
        self.assertEqual(result.contact_status, "Search error")

    def test_exact_2026_year_gate_uses_latest_article_or_cohort(self):
        article = pipeline.CompanyRecord(
            company="Article Co",
            discovery_hits=[pipeline.DiscoveryHit("Article Co", "News", "News/search", "https://example.com/a", "Fixture", article_year="2026")],
        )
        cohort = pipeline.CompanyRecord(
            company="Cohort Co",
            discovery_hits=[pipeline.DiscoveryHit("Cohort Co", "Accelerator", "Accelerator", "https://example.com/c", "Fixture", cohort_year="2026")],
        )
        old = pipeline.CompanyRecord(
            company="Old Co",
            discovery_hits=[pipeline.DiscoveryHit("Old Co", "News", "News/search", "https://example.com/o", "Fixture", article_year="2025")],
        )
        self.assertTrue(pipeline.is_linkedin_contact_target(article))
        self.assertTrue(pipeline.is_linkedin_contact_target(cohort))
        self.assertFalse(pipeline.is_linkedin_contact_target(old))

    def test_cache_round_trip_avoids_second_lookup(self):
        record = pipeline.CompanyRecord(company="NovaScan Health", website="https://novascan.example")
        enrichment = pipeline.LinkedInEnrichment(
            company_url="https://www.linkedin.com/company/novascan-health",
            company_status="Found - official website",
            contact_status="Not targeted",
        )
        with tempfile.TemporaryDirectory() as temp_dir, patch.object(pipeline, "LINKEDIN_ENRICHMENT_CACHE_DIR", temp_dir):
            pipeline.save_cached_linkedin_enrichment(record, False, enrichment)
            cached = pipeline.load_cached_linkedin_enrichment(record, False)
        self.assertEqual(cached.company_url, enrichment.company_url)
        self.assertEqual(cached.contact_status, "Not targeted")

    def test_cache_version_change_ignores_stale_files(self):
        record = pipeline.CompanyRecord(company="NovaScan Health", website="https://novascan.example")
        enrichment = pipeline.LinkedInEnrichment(company_url="https://www.linkedin.com/company/novascan-health")
        with tempfile.TemporaryDirectory() as temp_dir, \
             patch.object(pipeline, "LINKEDIN_ENRICHMENT_CACHE_DIR", temp_dir), \
             patch.object(pipeline, "LINKEDIN_ENRICHMENT_CACHE_VERSION", "linkedin_enrichment_old"):
            pipeline.save_cached_linkedin_enrichment(record, False, enrichment)
            self.assertIsNotNone(pipeline.load_cached_linkedin_enrichment(record, False))
            pipeline.LINKEDIN_ENRICHMENT_CACHE_VERSION = "linkedin_enrichment_new"
            self.assertIsNone(pipeline.load_cached_linkedin_enrichment(record, False))

    def test_transient_search_errors_are_not_cached(self):
        record = pipeline.CompanyRecord(company="NovaScan Health", website="https://novascan.example")
        companies = {record.company: record}
        error_result = pipeline.LinkedInEnrichment(company_status="Search error", contact_status="Not targeted")
        with tempfile.TemporaryDirectory() as temp_dir, \
             patch.object(pipeline, "LINKEDIN_ENRICHMENT_CACHE_DIR", temp_dir), \
             patch.object(pipeline, "enrich_company_linkedin", return_value=error_result):
            pipeline.enrich_companies_linkedin(companies, lambda _record: False)
            cached_files = list(Path(temp_dir).glob("*.json"))
        self.assertEqual(cached_files, [])

    def test_workbook_appends_linkedin_fields_and_hyperlinks(self):
        hit_2026 = pipeline.DiscoveryHit(
            "NovaScan Health", "News", "News/search", "https://example.com/a", "Fixture", website="https://novascan.example", article_year="2026"
        )
        hit_2025 = pipeline.DiscoveryHit(
            "Old Co", "News", "News/search", "https://example.com/o", "Fixture", website="https://old.example", article_year="2025"
        )
        companies = pipeline.normalize_companies([hit_2026, hit_2025])
        companies["NovaScan Health"].linkedin = pipeline.LinkedInEnrichment(
            company_url="https://www.linkedin.com/company/novascan-health",
            company_status="Found - official website",
            executive=pipeline.LinkedInContact("Alice Founder", "Chief Executive Officer", "https://www.linkedin.com/in/alice", "executive", "fixture", 0.9),
            technical=pipeline.LinkedInContact("Terry Tech", "VP Engineering", "https://www.linkedin.com/in/terry", "technical", "fixture", 0.9),
            quality=pipeline.LinkedInContact("Quinn Quality", "Director Quality", "https://www.linkedin.com/in/quinn", "quality", "fixture", 0.9),
            contact_status="Complete - 3 verified",
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            original_out = pipeline.OUT
            try:
                pipeline.OUT = Path(temp_dir) / "linkedin.xlsx"
                path = pipeline.write_workbook(companies, [hit_2026, hit_2025], [], [])
                workbook = load_workbook(path)
            finally:
                pipeline.OUT = original_out

        self.assertIn("Leads", workbook.sheetnames)
        self.assertNotIn("Lead Intake", workbook.sheetnames)
        self.assertNotIn("Lead Filtering", workbook.sheetnames)

        sheet = workbook["Leads"]
        headers = [cell.value for cell in sheet[1]]
        rows = {row[0].value: row for row in sheet.iter_rows(min_row=2)}
        self.assertEqual(headers[1], "Company website")
        self.assertEqual(headers[18], "Source URL")
        self.assertEqual(headers[24], "LinkedIn company URL")
        self.assertIn("Executive verification status", headers)
        self.assertIn("Quality/QA email status", headers)
        self.assertNotIn("Legacy score", headers)
        self.assertNotIn("Classification confidence", headers)
        self.assertNotIn("Pain hypothesis", headers)
        contact_status_col = headers.index("LinkedIn contact status")
        executive_url_col = headers.index("Executive LinkedIn URL")
        self.assertEqual(rows["NovaScan Health"][contact_status_col].value, "Complete - 3 verified")
        self.assertEqual(rows["NovaScan Health"][executive_url_col].hyperlink.target, "https://www.linkedin.com/in/alice")
        self.assertEqual(rows["Old Co"][10].value, "N/A")
        self.assertEqual(rows["Old Co"][contact_status_col].value, "Not targeted")


if __name__ == "__main__":
    unittest.main()
