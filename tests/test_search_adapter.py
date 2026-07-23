import json
import html
import re
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

import build_bbt_bizdev_workbook as pipeline


RSS_FIXTURE = """<?xml version="1.0" encoding="UTF-8"?>
<rss version="2.0">
  <channel>
    <item>
      <title>NovaScan Health raises $24M Series A for AI imaging platform - MedTech Dive</title>
      <link>https://news.google.com/rss/articles/novascan</link>
      <description>NovaScan Health raised new funding for FDA-focused medical imaging AI.</description>
      <source>MedTech Dive</source>
      <pubDate>Thu, 11 Jun 2026 12:00:00 GMT</pubDate>
    </item>
    <item>
      <title>NovaScan Health raises $24M Series A for AI imaging platform - MedTech Dive</title>
      <link>https://news.google.com/rss/articles/novascan</link>
      <description>Duplicate result from another query.</description>
      <source>MedTech Dive</source>
      <pubDate>Thu, 11 Jun 2026 12:00:00 GMT</pubDate>
    </item>
    <item>
      <title>PulseDx receives FDA clearance for AI cardiac device - Fierce Biotech</title>
      <link>https://news.google.com/rss/articles/pulsedx</link>
      <description>PulseDx received FDA clearance for a clinical AI medical device.</description>
      <source>Fierce Biotech</source>
      <pubDate>Thu, 11 Jun 2026 13:00:00 GMT</pubDate>
    </item>
    <item>
      <title>ClearPath Medical wins innovation award - Local Health News</title>
      <link>https://news.google.com/rss/articles/clearpath</link>
      <description>ClearPath Medical was highlighted for a medical device prototype.</description>
      <source>Local Health News</source>
      <pubDate>Thu, 11 Jun 2026 14:00:00 GMT</pubDate>
    </item>
  </channel>
</rss>
"""


class SearchAdapterTests(unittest.TestCase):
    def test_add_hyperlinks_trims_excel_relationship_targets(self):
        wb = pipeline.Workbook()
        ws = wb.active
        ws.append(["Website"])
        ws.append(["http://www.delee.co    "])

        pipeline.add_hyperlinks(ws, [1])

        self.assertEqual(ws["A2"].value, "http://www.delee.co")
        self.assertEqual(ws["A2"].hyperlink.target, "http://www.delee.co")

    def test_sources_include_52_vc_portfolio_pages(self):
        vc_sources = [source for source in pipeline.SOURCES if source.source_type == "VC portfolio"]

        self.assertEqual(len(vc_sources), 56)
        self.assertTrue(all(source.adapter for source in vc_sources))
        adapters = {source.name: source.adapter for source in vc_sources}
        self.assertEqual(adapters["Fountain Healthcare Partners portfolio"], "fountain_healthcare")
        self.assertEqual(adapters["Seroba Life Sciences portfolio"], "seroba_life_sciences")
        self.assertEqual(adapters["Atlantic Bridge portfolio"], "atlantic_bridge")

    def test_priority_ireland_sources_use_dedicated_adapters(self):
        adapters = {source.name: source.adapter for source in pipeline.SOURCES}

        self.assertEqual(adapters["BioInnovate Ireland"], "bioinnovate_ireland")
        self.assertEqual(adapters["ARC Hub for HealthTech"], "arc_hub_healthtech")
        self.assertEqual(adapters["Health Innovation Hub Ireland"], "health_innovation_hub_ireland")
        self.assertEqual(adapters["Dogpatch Labs / NDRC"], "dogpatch_ndrc")
        self.assertEqual(adapters["Fountain Healthcare Partners portfolio"], "fountain_healthcare")
        self.assertEqual(adapters["Seroba Life Sciences portfolio"], "seroba_life_sciences")
        self.assertEqual(adapters["Atlantic Bridge portfolio"], "atlantic_bridge")

    def test_priority_eu_accelerator_sources_have_bespoke_split(self):
        sources = {source.name: source for source in pipeline.SOURCES}

        self.assertGreaterEqual(len(pipeline.EU_ACCELERATOR_PRIORITY_SOURCE_NAMES), 50)
        self.assertGreaterEqual(len(pipeline.EU_ACCELERATOR_IMPLEMENTED_BESPOKE_SOURCE_NAMES), 10)
        self.assertIn("HealthTech Nordic", pipeline.EU_ACCELERATOR_BESPOKE_BACKLOG_SOURCE_NAMES)
        for name in pipeline.EU_ACCELERATOR_IMPLEMENTED_BESPOKE_SOURCE_NAMES:
            self.assertIn(name, sources)
            source = sources[name]
            self.assertEqual(source.source_type, "Accelerator")
            self.assertNotEqual(source.adapter, "accelerator_page")
            self.assertNotEqual(pipeline.adapter_inventory_label(source), "Manual/not implemented")
            self.assertNotEqual(source.adapter, "eu_accelerator_directory")
        for name in pipeline.EU_ACCELERATOR_BESPOKE_BACKLOG_SOURCE_NAMES:
            self.assertIn(name, sources)
            source = sources[name]
            self.assertEqual(source.source_type, "Accelerator")
            self.assertEqual(source.adapter, "accelerator_page")
            self.assertEqual(pipeline.adapter_inventory_label(source), "Manual/not implemented")

    def test_mars_health_source_uses_dedicated_adapter(self):
        adapters = {source.name: source.adapter for source in pipeline.SOURCES}

        self.assertEqual(adapters["MaRS Health Sciences"], "mars_health")

    def test_mars_health_parser_extracts_official_api_records(self):
        source = pipeline.Source(
            "MaRS Health Sciences",
            "Accelerator",
            "https://www.marsdd.com/our-sectors/health/",
            "Canada",
            "High",
            "Quarterly",
            "Official health ventures API extraction",
            "Canadian biotech, medtech, diagnostics, and digital health companies.",
            "mars_health",
        )
        payload = {
            "ventureInfo": [
                {
                    "title": "Fluid AI",
                    "permalink": "https://www.marsdd.com/venture/fluid-ai/",
                    "company_url": "https://fluidai.md/",
                    "description": "Postoperative monitoring technology.",
                    "venture_type": "health",
                }
            ]
        }

        hits = pipeline.parse_mars_health_payload(
            source,
            payload,
            "https://www.marsdd.com/wp-json/mars/v1/ventures?sector=health&page=1",
        )

        self.assertEqual([hit.company for hit in hits], ["Fluid AI"])
        self.assertEqual(hits[0].website, "https://fluidai.md/")
        self.assertEqual(hits[0].category_or_track, "Health")
        self.assertIn("mars_health", hits[0].matched_terms)

    def test_mars_health_runner_paginates_and_checks_official_total(self):
        source = pipeline.Source(
            "MaRS Health Sciences",
            "Accelerator",
            "https://www.marsdd.com/our-sectors/health/",
            "Canada",
            "High",
            "Quarterly",
            "Official health ventures API extraction",
            "Canadian health ventures.",
            "mars_health",
        )
        pages = [
            {
                "current_page": "1",
                "max_pages": 2,
                "total_found": 2,
                "ventureInfo": [
                    {
                        "title": "Fluid AI",
                        "permalink": "https://www.marsdd.com/venture/fluid-ai/",
                        "company_url": "https://fluidai.md/",
                    }
                ],
            },
            {
                "current_page": "2",
                "max_pages": 2,
                "total_found": 2,
                "ventureInfo": [
                    {
                        "title": "Retispec",
                        "permalink": "https://www.marsdd.com/venture/retispec/",
                        "company_url": "https://www.retispec.com/",
                    }
                ],
            },
        ]

        with patch.object(
            pipeline,
            "ACCELERATOR_SOURCE_PAGES",
            {"MaRS Health Sciences": ["https://www.marsdd.com/wp-json/mars/v1/ventures?sector=health&page=1"]},
        ), patch.object(pipeline, "fetch_json_url", side_effect=[(pages[0], None), (pages[1], None)]):
            discovery_hits, trigger_events, result = pipeline.run_mars_health(source)

        self.assertEqual([hit.company for hit in discovery_hits], ["Fluid AI", "Retispec"])
        self.assertEqual(len(trigger_events), 2)
        self.assertIn("2/2 public directory ventures", result)
        self.assertNotIn("INCOMPLETE", result)

    def test_mars_ventureconnect_source_uses_snapshot_adapter(self):
        adapters = {source.name: source.adapter for source in pipeline.SOURCES}

        self.assertEqual(
            adapters["MaRS VentureConnect Healthcare & Life Sciences"],
            "mars_ventureconnect",
        )

    def test_mars_ventureconnect_snapshot_runner_checks_reported_total(self):
        source = pipeline.Source(
            "MaRS VentureConnect Healthcare & Life Sciences",
            "Accelerator",
            "https://app.marsdd.com/companies",
            "Canada",
            "High",
            "Quarterly",
            "Verified browser snapshot extraction",
            "Canadian health companies.",
            "mars_ventureconnect",
        )
        payload = {
            "collected_at": "2026-07-23",
            "reported_total": 2,
            "extracted_total": 2,
            "records": [
                {
                    "company": "LIOR Pupillometry",
                    "profile_url": "https://app.marsdd.com/companies/lior-pupillometry",
                    "tags": ["Medical Device", "Diagnostics Imaging and Sensors"],
                    "description": "Portable neurological screening device.",
                },
                {
                    "company": "NodeAI",
                    "profile_url": "https://app.marsdd.com/companies/nodeai",
                    "tags": ["Diagnostics", "AI-Enabled Diagnostics"],
                    "description": "AI diagnostic software for lung cancer.",
                },
            ],
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            snapshot = Path(temp_dir) / "mars.json"
            snapshot.write_text(json.dumps(payload), encoding="utf-8")
            discovery_hits, trigger_events, result = pipeline.run_mars_ventureconnect(source, snapshot)

        self.assertEqual([hit.company for hit in discovery_hits], ["LIOR Pupillometry", "NodeAI"])
        self.assertEqual(discovery_hits[0].category_or_track, "Medical Device; Diagnostics Imaging and Sensors")
        self.assertEqual(len(trigger_events), 2)
        self.assertIn("2/2 VentureConnect", result)
        self.assertNotIn("INCOMPLETE", result)

    def test_tiap_source_uses_dedicated_adapter(self):
        source = next(source for source in pipeline.SOURCES if source.name == "Toronto Innovation Acceleration Partners (TIAP)")

        self.assertEqual(source.adapter, "tiap_portfolio")
        self.assertEqual(source.url, "https://tiap.ca/portfolio/")

    def test_tiap_parser_preserves_status_website_and_description(self):
        source = pipeline.Source(
            "Toronto Innovation Acceleration Partners (TIAP)",
            "Accelerator",
            "https://tiap.ca/portfolio/",
            "Canada",
            "High",
            "Quarterly",
            "Official static portfolio extraction",
            "",
            "tiap_portfolio",
        )
        html = """
        <div class="eael-filterable-gallery-item-wrap eael-cf-active-portfolio">
          <div><a href="https://cohesys.ca/"><h1 class="fg-item-title">Cohesys</h1></a>
          <div class="fg-item-content"><p>Developing surgical tape for fractures.</p></div></div>
        </div>
        <div class="eael-filterable-gallery-item-wrap eael-cf-exited">
          <div><a><h1 class="fg-item-title">CoHealth (Acquired in 2020)</h1></a>
          <div class="fg-item-content"><p>Delivery platform for educational health content.</p></div></div>
        </div>
        """

        hits = pipeline.parse_tiap_portfolio_html(source, html)

        self.assertEqual([hit.company for hit in hits], ["Cohesys", "CoHealth"])
        self.assertEqual(hits[0].website, "https://cohesys.ca/")
        self.assertEqual(hits[0].category_or_track, "Active")
        self.assertEqual(hits[1].category_or_track, "Exited / acquired in 2020")
        self.assertIn("educational health content", hits[1].company_description)

    def test_tiap_runner_checks_active_and_exited_totals(self):
        source = pipeline.Source(
            "Toronto Innovation Acceleration Partners (TIAP)",
            "Accelerator",
            "https://tiap.ca/portfolio/",
            "Canada",
            "High",
            "Quarterly",
            "Official static portfolio extraction",
            "",
            "tiap_portfolio",
        )

        def card(index, status):
            return f"""
            <div class="eael-filterable-gallery-item-wrap eael-cf-{status}">
              <div><a href="https://venture-{index}.example/"><h1 class="fg-item-title">Venture {index:02d}</h1></a>
              <div class="fg-item-content"><p>Health technology company {index}.</p></div></div>
            </div>
            """

        html = "".join(card(index, "active-portfolio") for index in range(41))
        html += "".join(card(index, "exited") for index in range(41, 57))

        hits, triggers, result = pipeline.run_tiap_portfolio(source, fetcher=lambda url: (html, None))

        self.assertEqual(len(hits), 57)
        self.assertEqual(len(triggers), 57)
        self.assertIn("57/57", result)
        self.assertIn("41/41 active", result)
        self.assertIn("16/16 exited", result)
        self.assertNotIn("INCOMPLETE", result)

        _, _, incomplete = pipeline.run_tiap_portfolio(source, fetcher=lambda url: (card(1, "active-portfolio"), None))
        self.assertIn("INCOMPLETE", incomplete)

    def test_admare_source_uses_dedicated_adapter(self):
        source = next(source for source in pipeline.SOURCES if source.name == "adMare BioInnovations")

        self.assertEqual(source.adapter, "admare_portfolio")
        self.assertEqual(source.url, "https://www.admarebio.com/en/companies-weve-helped-build")

    def test_admare_parsers_preserve_cohort_and_detail_fields(self):
        source = next(source for source in pipeline.SOURCES if source.name == "adMare BioInnovations")
        index_html = """
        <div class="list-item-container">
          <h5><a href="/en/companies-weve-helped-build/flosonics-medical-1">Flosonics Medical</a></h5>
        </div>
        """
        records = pipeline.parse_admare_index_html(index_html, source.url, "Companies we've helped build")
        detail_html = """
        <div class="search-object-detail-bloc"><h1>Flosonics Medical</h1>
          <div class="item-website"><a href="//www.flosonicsmedical.com">Website</a></div>
          <div class="item-email"><a href="mailto:hello@example.com"></a></div>
          <div class="item-LinkedIn" data-url="https://linkedin.com/company/flosonics-medical"></div>
          <div class="item-line-of"><span>Medical Devices</span></div>
          <div class="item-detail-right"><h2>Description</h2>
            <p>Non-invasive sensors for critically ill patients.</p>
            <h2>More information</h2><ul><li><a href="https://example.com/news">Funding news</a></li></ul>
            <a class="btn-text-bt-arrow-back">Back</a>
          </div>
        </div>
        """
        records[0].update(pipeline.parse_admare_detail_html(detail_html, records[0]["detail_url"]))
        hits = pipeline.parse_admare_records(source, records)

        self.assertEqual([hit.company for hit in hits], ["Flosonics Medical"])
        self.assertEqual(hits[0].website, "https://www.flosonicsmedical.com")
        self.assertEqual(hits[0].category_or_track, "Medical Devices")
        self.assertIn("critically ill", hits[0].company_description)
        self.assertIn("LinkedIn:", hits[0].matched_terms)

    def test_admare_runner_checks_both_official_cohorts(self):
        source = next(source for source in pipeline.SOURCES if source.name == "adMare BioInnovations")
        pages = {}

        def index_page(prefix, count):
            return "".join(
                f'<div class="list-item-container"><h5><a href="/en/{prefix}/company-{index}">Company {prefix[:1]}{index:02d}</a></h5></div>'
                for index in range(count)
            )

        pages[source.url] = index_page("companies-weve-helped-build", 39)
        pages[pipeline.ADMARE_ACCELERATOR_URL] = index_page("accelerator-companies", 13)
        for url, html in list(pages.items()):
            for href in re.findall(r'href="([^"]+)"', html):
                detail_url = pipeline.urljoin(url, href)
                pages[detail_url] = """
                <div class="search-object-detail-bloc"><h1>Company</h1>
                  <div class="item-website"><a href="https://company.example">Website</a></div>
                  <div class="item-line-of"><span>Therapeutics</span></div>
                  <div class="item-detail-right"><h2>Description</h2><p>Drug development company.</p>
                  <h2>More information</h2><a class="btn-text-bt-arrow-back">Back</a></div>
                </div>
                """

        hits, triggers, result = pipeline.run_admare_portfolio(
            source,
            fetcher=lambda url: (pages.get(url, ""), None if url in pages else "not found"),
        )

        self.assertEqual(len(hits), 52)
        self.assertEqual(len(triggers), 52)
        self.assertIn("52/52 adMare companies", result)
        self.assertIn("39/39 helped-build", result)
        self.assertIn("13/13 accelerator", result)
        self.assertNotIn("INCOMPLETE", result)

    def test_priority_eu_vc_sources_are_configured(self):
        sources = {source.name: source for source in pipeline.SOURCES}

        self.assertEqual(len(pipeline.PRIORITY_EU_VC_SOURCE_NAMES), 20)
        for name in pipeline.PRIORITY_EU_VC_SOURCE_NAMES:
            self.assertIn(name, sources)
            source = sources[name]
            self.assertEqual(source.source_type, "VC portfolio")
            self.assertTrue(source.adapter)

    def test_sources_include_20_jobs_sources_with_dedicated_ats_adapters(self):
        job_sources = [source for source in pipeline.SOURCES if source.source_type == "Jobs"]
        adapters = {source.name: source.adapter for source in job_sources}

        self.assertEqual(len(job_sources), 20)
        self.assertIsNone(adapters["Greenhouse job boards"])
        self.assertIsNone(adapters["Wellfound jobs"])
        self.assertEqual(adapters["Lever job boards"], "lever_jobs")
        self.assertEqual(adapters["Ashby job boards"], "ashby_jobs")
        self.assertEqual(adapters["Workable job boards"], "workable_jobs")
        self.assertEqual(adapters["SmartRecruiters job boards"], "smartrecruiters_jobs")
        self.assertEqual(adapters["Recruitee job boards"], "recruitee_jobs")
        self.assertEqual(adapters["Built In jobs"], "builtin_jobs")
        self.assertEqual(adapters["BioSpace jobs"], "biospace_jobs")
        self.assertEqual(adapters["NHS Jobs"], "nhs_jobs")
        self.assertEqual(sum(1 for source in job_sources if source.adapter == "jobs_page"), 10)

    def test_sources_include_university_spinout_sources_by_geo(self):
        spinout_sources = [source for source in pipeline.SOURCES if source.source_type == "University/spinout"]
        names = {source.name for source in spinout_sources}

        self.assertEqual(len(spinout_sources), 60)
        self.assertTrue(all(pipeline.adapter_inventory_label(source) != "Manual/not implemented" for source in spinout_sources))
        self.assertIn("University/spinout", pipeline.DISCOVERY_TERMS)
        self.assertEqual(pipeline.SOURCE_TRIGGER_TYPES["University/spinout"], "University/spinout origin")
        for name in [
            "Trinity College Dublin spinouts",
            "RCSI spinouts",
            "UCD spinouts",
            "University of Galway spinouts",
            "University of Limerick spinouts",
            "University College Cork spinouts",
            "Queen's University Belfast spinouts",
            "University of Oxford spinouts",
            "University of Cambridge spinouts",
            "Imperial College London spinouts",
            "University of Bristol spinouts",
            "King's College London spinouts",
            "UCL spinouts",
            "University of Edinburgh spinouts",
            "University of Manchester spinouts",
            "University of Leeds spinouts",
            "University of Sheffield spinouts",
            "ETH Zurich spinouts",
            "KU Leuven spinouts",
            "EPFL spinouts",
            "Technical University of Denmark spinouts",
            "TU Delft spinouts",
            "Karolinska Institutet spinouts",
            "Stanford spinouts",
            "MIT spinouts",
            "Harvard spinouts",
            "Mass General Brigham spinouts",
            "Broad Institute ventures",
            "Johns Hopkins spinouts",
            "Mayo Clinic spinouts",
            "UC Berkeley spinouts",
            "UCSF spinouts",
            "University of Pennsylvania spinouts",
            "CHOP spinouts",
            "University of Washington spinouts",
            "Fred Hutch spinouts",
            "University of Michigan spinouts",
            "Duke University spinouts",
            "UNC Chapel Hill spinouts",
            "NC State spinouts",
            "UC San Diego spinouts",
            "Scripps Research spinouts",
            "Georgia Tech spinouts",
            "Emory University spinouts",
            "Columbia University spinouts",
            "Cornell University spinouts",
            "Weill Cornell Medicine spinouts",
            "NYU spinouts",
            "Carnegie Mellon spinouts",
            "University of Pittsburgh spinouts",
            "Rice University spinouts",
            "Baylor College of Medicine spinouts",
            "Texas Medical Center Innovation",
            "UT Austin spinouts",
            "MD Anderson spinouts",
            "Vanderbilt University spinouts",
            "University of Toronto Health & Life Sciences startups",
        ]:
            self.assertIn(name, names)

    def test_sources_include_us_prioritization_sources(self):
        sources = {source.name: source for source in pipeline.SOURCES}

        self.assertEqual(sources["AUTM Licensing Survey / STATT"].source_type, "Institution prioritization")
        self.assertEqual(sources["BRIMR NIH rankings"].source_type, "Institution prioritization")
        self.assertIsNone(sources["AUTM Licensing Survey / STATT"].adapter)
        self.assertIsNone(sources["BRIMR NIH rankings"].adapter)

    def test_google_news_queries_quarantine_university_spinout_terms(self):
        expected_quarantined_count = len(pipeline.UNIVERSITY_SPINOUT_SEARCH_UNIVERSITIES) * len(pipeline.UNIVERSITY_SPINOUT_SEARCH_PATTERNS)

        self.assertEqual(pipeline.SEARCH_QUERIES, pipeline.CORE_SEARCH_QUERIES)
        self.assertEqual(len(pipeline.UNIVERSITY_SPINOUT_SEARCH_QUERIES), expected_quarantined_count)
        self.assertNotIn('"University of Oxford" spinout digital health', pipeline.SEARCH_QUERIES)
        self.assertIn('"University of Oxford" spinout digital health', pipeline.UNIVERSITY_SPINOUT_SEARCH_QUERIES)

    def test_university_spinout_adapter_extracts_candidate_company_links(self):
        source = pipeline.Source("Trinity College Dublin spinouts", "University/spinout", "https://www.tcd.ie/innovation/", "Ireland", "High", "Quarterly", "Spinout extraction", "AI health and medtech spinouts.", "tcd_spinouts")
        html = """
        <html><body>
          <article><a href="https://retinaai.example/">RetinaAI Health</a><p>AI imaging medical device spinout.</p></article>
          <a href="/innovation/news/2026/new-healthtech-start-up-launches/">New Healthtech Start-up Launches</a>
          <a href="/innovation/contact/">Contact</a>
        </body></html>
        """

        discovery_hits, trigger_events = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["RetinaAI Health"])

    def test_uoft_health_source_uses_dedicated_adapter(self):
        source = next(source for source in pipeline.SOURCES if source.name == "University of Toronto Health & Life Sciences startups")

        self.assertEqual(source.adapter, "uoft_health_startups")
        self.assertEqual(source.source_type, "University/spinout")
        self.assertIn("startups-directory", source.url)

    def test_uoft_health_parser_preserves_directory_metadata(self):
        source = pipeline.Source(
            "University of Toronto Health & Life Sciences startups",
            "University/spinout",
            "https://entrepreneurs.utoronto.ca/our-startups/startups-directory/",
            "Canada",
            "High",
            "Quarterly",
            "Official REST directory extraction",
            "",
            "uoft_health_startups",
        )
        payload = {
            "structured_posts": [
                {
                    "title": "Cohesys",
                    "link": "https://entrepreneurs.utoronto.ca/startup/cohesys/",
                    "excerpt": "Medical-device company developing a flexible surgical adhesive.",
                    "terms": {
                        "category": [
                            {"id": 58, "name": "Health &amp; Life Sciences"},
                            {"id": 59, "name": "Medical Devices"},
                        ]
                    },
                    "postmeta": {
                        "location": ["Toronto, ON"],
                        "size": ["1-10"],
                        "additional_links_0_link": [
                            'a:3:{s:5:"title";s:23:"Visit Cohesys Website";s:3:"url";s:22:"http://www.cohesys.ca/";s:6:"target";s:6:"_blank";}'
                        ],
                    },
                    "accelerator_post_tags": '<a title="accelerator: UTEST">UTEST</a>',
                }
            ]
        }

        hits = pipeline.parse_uoft_health_startup_payload(source, payload)

        self.assertEqual(len(hits), 1)
        self.assertEqual(hits[0].company, "Cohesys")
        self.assertEqual(hits[0].website, "http://www.cohesys.ca/")
        self.assertEqual(hits[0].geography, "Toronto, ON")
        self.assertIn("Medical Devices", hits[0].matched_terms)
        self.assertIn("UTEST", hits[0].matched_terms)
        self.assertIn("listed size: 1-10", hits[0].matched_terms)

    def test_uoft_health_runner_paginates_and_checks_total(self):
        source = pipeline.Source(
            "University of Toronto Health & Life Sciences startups",
            "University/spinout",
            "https://entrepreneurs.utoronto.ca/our-startups/startups-directory/",
            "Canada",
            "High",
            "Quarterly",
            "Official REST directory extraction",
            "",
            "uoft_health_startups",
        )
        records = [
            {
                "title": f"Health Venture {index:03d}",
                "link": f"https://entrepreneurs.utoronto.ca/startup/health-venture-{index:03d}/",
                "excerpt": "Health technology company.",
                "terms": {"category": [{"id": 58, "name": "Health & Life Sciences"}]},
                "postmeta": {"location": ["Toronto, ON"]},
            }
            for index in range(171)
        ]
        requested_pages = []

        def fake_fetcher(url, payload):
            requested_pages.append(int(payload["pageNum"]))
            self.assertEqual(json.loads(payload["taxQueries"]), {"category": ["58"]})
            start = (int(payload["pageNum"]) - 1) * 16
            return {
                "total_posts": 171,
                "structured_posts": records[start : start + 16],
            }, None

        hits, triggers, result = pipeline.run_uoft_health_startups(source, fetcher=fake_fetcher)

        self.assertEqual(len(hits), 171)
        self.assertEqual(len(triggers), 171)
        self.assertEqual(requested_pages, list(range(1, 12)))
        self.assertIn("171/171", result)
        self.assertNotIn("INCOMPLETE", result)

    def test_university_spinout_adapter_uses_curated_official_fallbacks(self):
        source = pipeline.Source("EPFL spinouts", "University/spinout", "https://www.epfl.ch/innovation/startup/", "Europe", "High", "Quarterly", "Spinout extraction", "EPFL medical startups.", "university_spinout_directory")

        with patch.object(pipeline, "fetch_raw_text", return_value=("<html><body></body></html>", None)):
            discovery_hits, trigger_events, result = pipeline.run_university_spinout_pages(source)

        names = {hit.company for hit in discovery_hits}
        self.assertIn("MoleSense", names)
        self.assertIn("SwissIonics", names)
        self.assertTrue(all(hit.source_type == "University/spinout" for hit in discovery_hits))
        self.assertTrue(all("curated official source fallback" in hit.matched_terms for hit in discovery_hits))
        self.assertEqual(discovery_hits[0].source_type, "University/spinout")
        self.assertEqual(trigger_events[0].trigger_type, "University/spinout origin")
        self.assertIn("curated", result)

    def test_university_spinout_adapter_extracts_yesdelft_health_api(self):
        source = pipeline.Source("TU Delft spinouts", "University/spinout", "https://yesdelft.com/wp-json/wp/v2/startups?sectors=49&per_page=100", "Europe", "Medium", "Quarterly", "Spinout extraction", "YES!Delft Health and Pharma startups.", "university_spinout_directory")
        payload = """
        [
          {
            "title": {"rendered": "Corbotics"},
            "link": "https://yesdelft.com/startups/corbotics/",
            "content": {"rendered": "<p>Autonomous cardiac echo robot.</p><a href=\\"https://www.corbotics.com\\">Website</a>"},
            "excerpt": {"rendered": "Health & Pharma startup"}
          },
          {
            "title": {"rendered": "Nonmedical Sensor"},
            "link": "https://yesdelft.com/startups/nonmedical-sensor/",
            "excerpt": {"rendered": "Industrial startup"}
          }
        ]
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, payload)

        self.assertEqual([hit.company for hit in discovery_hits], ["Corbotics", "Nonmedical Sensor"])
        self.assertEqual(discovery_hits[0].website, "https://www.corbotics.com")
        self.assertEqual(discovery_hits[0].discovery_url, "https://yesdelft.com/startups/corbotics/")

    def test_university_spinout_adapter_extracts_kuleuven_heading_directory(self):
        source = pipeline.Source("KU Leuven spinouts", "University/spinout", "https://lrd.kuleuven.be/en/spinoff/spin-off-companies", "Europe", "High", "Quarterly", "Spinout extraction", "KU Leuven spin-off portfolio.", "university_spinout_directory")
        html = """
        <html><body>
          <h4>ADx NeuroSciences</h4>
          <p><a href="https://www.adxneurosciences.com">ADx NeuroSciences</a> focuses on biomarkers for dementia diagnosis.</p>
          <h4>Adinex</h4>
          <p><a href="https://www.adinex.be">Adinex NV</a> offers industrial explosion safety services.</p>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["ADx NeuroSciences", "Adinex"])
        self.assertEqual(discovery_hits[0].website, "https://www.adxneurosciences.com")

    def test_university_spinout_adapter_extracts_rcsi_repeated_card_text(self):
        source = pipeline.Source("RCSI spinouts", "University/spinout", "https://www.rcsi.com/dublin/research-and-innovation/innovation/investors-entrepreneurs-and-spin-outs", "Ireland", "High", "Quarterly", "Spinout extraction", "RCSI spin-out companies.", "university_spinout_directory")
        html = """
        <html><body>
          <a href="/dublin/research-and-innovation/innovation/investors-entrepreneurs-and-spin-outs/inthelia-therapeutics">
            Inthelia Therapeutics Inthelia Therapeutics is developing new treatments for sepsis.
          </a>
          <a href="/kelada">KelAda Pharmachem KelAda’s mission is to improve pharmaceutical manufacturing.</a>
          <a href="/docleaf">DocLeaf The DocLeaf Project is developing wound repair device technology.</a>
          <a href="/professional">Professional exams</a>
          <a href="/contact">Contact us Contact us</a>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["Inthelia Therapeutics", "KelAda Pharmachem", "DocLeaf"])
        self.assertIn("sepsis", discovery_hits[0].company_description)

    def test_university_spinout_adapter_rejects_generic_navigation(self):
        source = pipeline.Source("Imperial College London spinouts", "University/spinout", "https://www.imperial.ac.uk/enterprise/", "UK", "High", "Quarterly", "Spinout extraction", "Medtech spinouts.", "imperial_spinouts")
        html = """
        <html><body>
          <a href="/admin-services/enterprise/about/data-and-reporting/metrics-2017-18/4-commercialisation/">4. Commercialisation</a>
          <a href="/innovation/fast-ip/fast-ip-faqs/">FAQs</a>
          <a href="/innovation/location/">Location</a>
          <a href="/innovation/about/our-leadership/">Our leadership</a>
          <a href="/innovation/licensing/">Licensing</a>
          <a href="/licensing-opportunities/browse-innovations">Browse innovations</a>
          <a href="/explore-consultancy/be-consultant">Be a consultant</a>
          <article><a href="https://retinaai.example/">RetinaAI Health</a><p>Medtech spinout building diagnostic imaging software.</p></article>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["RetinaAI Health"])

    def test_university_spinout_adapter_extracts_official_directory_shapes(self):
        source = pipeline.Source("University of Oxford spinouts", "University/spinout", "https://innovation.ox.ac.uk/portfolio/", "UK", "High", "Quarterly", "Spinout extraction", "Medtech spinouts.", "oxford_spinouts")
        html = """
        <html><body>
          <div class="portfolio-card">
            <a href="https://example-oxford-spinout.com">NeuroVista</a>
            <p>Medical imaging AI platform spun out of Oxford research.</p>
          </div>
          <div class="portfolio-card">
            <a href="/portfolio/quantum-battery/">Quantum Battery</a>
            <p>Energy storage materials.</p>
          </div>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["NeuroVista"])
        self.assertEqual(discovery_hits[0].website, "https://example-oxford-spinout.com")
        self.assertEqual(discovery_hits[0].discovery_url, "https://innovation.ox.ac.uk/portfolio/")
        self.assertIn("official university directory", discovery_hits[0].matched_terms)

    def test_university_spinout_adapter_extracts_oxford_finance_loop_cards(self):
        source = pipeline.Source("University of Oxford spinouts", "University/spinout", "https://www.oxfordinnovationfinance.co.uk/portfolio/", "UK", "High", "Quarterly", "Spinout extraction", "Health and science portfolio companies.", "oxford_spinouts")
        html = """
        <html><body>
          <div data-elementor-type="loop-item" class="portfolio type-portfolio category-health-science">
            <a class="loop-card" href="https://crainio.com">
              <div class="company-name"><p>Crainio</p></div>
              <div data-widget_type="theme-post-excerpt.default">Non-invasive measurement of intracranial pressure, a vital indicator of brain health.</div>
              <p><span>Health &amp; Science</span></p>
            </a>
          </div>
          <div data-elementor-type="loop-item" class="portfolio type-portfolio category-health-science">
            <a class="loop-card" href="https://perkier.co.uk">
              <div class="company-name"><p>Perkier Foods</p></div>
              <div data-widget_type="theme-post-excerpt.default">A challenger food brand leveraging healthy, free-from and plant based trends.</div>
              <p><span>Health &amp; Science</span></p>
            </a>
          </div>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["Crainio"])
        self.assertEqual(discovery_hits[0].website, "https://crainio.com")
        self.assertEqual(discovery_hits[0].discovery_url, "https://www.oxfordinnovationfinance.co.uk/portfolio/")

    def test_university_spinout_adapter_extracts_cambridge_modal_cards(self):
        source = pipeline.Source("University of Cambridge spinouts", "University/spinout", "https://www.enterprise.cam.ac.uk/portfolio/", "UK", "High", "Quarterly", "Spinout extraction", "Healthcare and life sciences spinouts.", "cambridge_spinouts")
        html = """
        <html><body>
          <div id="company-52north" class="mfp-hide">
            <h2>52 North</h2>
            <p>52 North is reinventing the healthcare journey with medical device technology.</p>
            <a href="https://www.linkedin.com/in/founder/">Founder</a>
            <a class="button" href="https://www.52north.health">Visit the website</a>
          </div>
          <div id="company-joltsynsor" class="mfp-hide">
            <h2>JoltSynSor</h2>
            <p>AI-driven structural health monitoring for infrastructure.</p>
            <a class="button" href="https://www.joltsynsor.com/">Visit the website</a>
          </div>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["52 North"])
        self.assertEqual(discovery_hits[0].website, "https://www.52north.health")
        self.assertEqual(discovery_hits[0].discovery_url, "https://www.enterprise.cam.ac.uk/portfolio/")

    def test_university_spinout_adapter_extracts_imperial_link_sources(self):
        source = pipeline.Source("Imperial College London spinouts", "University/spinout", "https://www.imperial.ac.uk/news/alumni/", "UK", "High", "Quarterly", "Spinout extraction", "Imperial health and medtech founders.", "imperial_spinouts")
        html = """
        <html><body>
          <a href="/research-groups/">Research groups</a>
          <p>Dr Max Munford, founder of <a href="https://osstec.uk/">OSSTEC</a>, which has developed 3D-printed replacement joints.</p>
          <p><a href="https://www.pulpatronics.com/">PulpaTronics</a> is making RFID tags for retail.</p>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["OSSTEC"])
        self.assertEqual(discovery_hits[0].website, "https://osstec.uk/")
        self.assertEqual(discovery_hits[0].discovery_url, "https://www.imperial.ac.uk/news/alumni/")

    def test_university_spinout_adapter_extracts_qubis_internal_profiles(self):
        source = pipeline.Source("Queen's University Belfast spinouts", "University/spinout", "https://www.qubis.co.uk/portfolio/all", "Ireland/UK", "Medium", "Quarterly", "Spinout extraction", "QUBIS portfolio.", "qubis_spinouts")
        html = """
        <html><body>
          <a href="/portfolio/aflo-respiratory-analytics-ltd">Life Science, Scientific Aflo - Respiratory Analytics Ltd A health tech company transforming respiratory care with a smart inhaler platform.</a>
          <a href="/portfolio/nuada">Environmental Nuada Carbon capture technology.</a>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["Aflo Respiratory Analytics"])
        self.assertEqual(discovery_hits[0].discovery_url, "https://www.qubis.co.uk/portfolio/aflo-respiratory-analytics-ltd")

    def test_university_spinout_adapter_extracts_mit_startup_profiles(self):
        source = pipeline.Source("MIT spinouts", "University/spinout", "https://tlo.mit.edu/industry-entrepreneurs/startups/", "US", "High", "Quarterly", "Spinout extraction", "MIT startup directory.", "mit_spinouts")
        html = """
        <html><body>
          <div class="startup-card">
            <h3><a href="/industry-entrepreneurs/startups/musclemetrix">MuscleMetrix</a></h3>
            <p>Biomaterials, bioelectronics, biotechnology, sensing, imaging, and healthy living.</p>
          </div>
          <div class="startup-card">
            <h3><a href="/industry-entrepreneurs/startups/found-energy">Found Energy</a></h3>
            <p>Clean energy metals recycling.</p>
          </div>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["MuscleMetrix"])
        self.assertEqual(discovery_hits[0].discovery_url, "https://tlo.mit.edu/industry-entrepreneurs/startups/musclemetrix")

    def test_university_spinout_adapter_extracts_harvard_venture_cards(self):
        source = pipeline.Source("Harvard spinouts", "University/spinout", "https://innovationlabs.harvard.edu/ventures/", "US", "High", "Quarterly", "Spinout extraction", "Harvard venture directory.", "harvard_ventures")
        html = """
        <html><body>
          <a href="https://innovationlabs.harvard.edu/venture/aidra-health" class="venture-card student-i-lab">
            <h3>Aidra Health</h3>
            <p>Democratizing access to life-saving medical equipment in emerging markets.</p>
          </a>
          <a href="https://innovationlabs.harvard.edu/venture/air-right-exchange" class="venture-card student-i-lab">
            <h3>Air Right Exchange of America</h3>
            <p>Simplifying zoning and air rights transfer.</p>
          </a>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["Aidra Health"])
        self.assertEqual(discovery_hits[0].discovery_url, "https://innovationlabs.harvard.edu/venture/aidra-health")

    def test_university_spinout_adapter_extracts_bayes_cohort_profiles(self):
        source = pipeline.Source("University of Edinburgh spinouts", "University/spinout", "https://bayes-centre.ed.ac.uk/programmes/vbi/cohorts/6.0", "UK", "Medium", "Quarterly", "Spinout extraction", "Bayes Centre cohort.", "edinburgh_spinouts")
        html = """
        <html><body>
          <h2><a href="/accelerating-entrepreneurship/venture-builder-incubator/vbi/cohort-6/cadencedx">CadenceDx</a></h2>
          <p>Meet CadenceDx.</p>
          <h2><a href="/accelerating-entrepreneurship/venture-builder-incubator/vbi/cohort-6/ecomatter">EcoMatter</a></h2>
          <p>Meet EcoMatter.</p>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["CadenceDx"])
        self.assertEqual(discovery_hits[0].discovery_url, "https://bayes-centre.ed.ac.uk/accelerating-entrepreneurship/venture-builder-incubator/vbi/cohort-6/cadencedx")

    def test_priority_university_spinout_adapter_scans_configured_pages(self):
        source = pipeline.Source("UCD spinouts", "University/spinout", "https://www.ucd.ie/innovation/", "Ireland", "High", "Quarterly", "Spinout extraction", "Medtech and digital health spinouts.", "ucd_spinouts")
        pages = [
            ('<article><a href="https://clinicflow.example/">ClinicFlow Health</a><p>Digital health workflow software.</p></article>', None),
            ('<article><a href="https://pulsedx.example/">PulseDx</a><p>Medical device diagnostics startup.</p></article>', None),
            ("", "HTTP Error 404: Not Found"),
        ]

        page_urls = ["https://example.com/community", "https://example.com/alumni", "https://example.com/missing"]
        with patch.dict(pipeline.UNIVERSITY_SPINOUT_SOURCE_PAGES, {"UCD spinouts": page_urls}):
            with patch.object(pipeline, "fetch_raw_text", side_effect=pages):
                discovery_hits, trigger_events, result = pipeline.run_university_spinout_pages(source)

        self.assertEqual([hit.company for hit in discovery_hits], ["ClinicFlow Health", "PulseDx"])
        self.assertEqual([hit.discovery_url for hit in discovery_hits], ["https://example.com/community", "https://example.com/alumni"])
        self.assertEqual(discovery_hits[0].website, "https://clinicflow.example/")
        self.assertEqual(len(trigger_events), 2)
        self.assertIn("3 configured university directory pages", result)
        self.assertIn("HTTP Error 404", result)

    def test_university_spinout_sources_without_directories_skip_fail_closed(self):
        source = pipeline.Source("Unknown University spinouts", "University/spinout", "https://example.edu/", "UK", "Low", "Quarterly", "Spinout extraction", "Medtech spinouts.")

        discovery_hits, trigger_events, result = pipeline.run_university_spinout_pages(source)

        self.assertEqual(discovery_hits, [])
        self.assertEqual(trigger_events, [])
        self.assertIn("skipped fail-closed", result)

    def test_university_workbook_qa_rejects_old_bad_company_titles(self):
        bad_titles = {"FAQs", "Location", "Our leadership", "Licensing", "Browse innovations", "Be a consultant"}
        source = pipeline.Source("UCD spinouts", "University/spinout", "https://www.ucd.ie/innovation/", "Ireland", "High", "Quarterly", "Spinout extraction", "Medtech spinouts.", "ucd_spinouts")
        html = "".join(f'<a href="/innovation/{title.lower().replace(" ", "-")}/">{title}</a>' for title in bad_titles)

        discovery_hits, _ = pipeline.build_university_spinout_evidence(source, html)

        self.assertEqual(discovery_hits, [])
        self.assertTrue(all("university spinout page scan" not in hit.matched_terms for hit in discovery_hits))

    def test_job_board_parsers_cover_supported_ats_shapes(self):
        greenhouse = pipeline.parse_greenhouse_jobs({"jobs": [{"title": "Quality Engineer", "absolute_url": "https://job/gh", "content": "<p>Medical device QA</p>"}]})
        lever = pipeline.parse_lever_jobs([{"text": "Regulatory Affairs Lead", "hostedUrl": "https://job/lever", "description": "FDA submissions"}])
        ashby = pipeline.parse_ashby_jobs({"jobs": [{"title": "Clinical AI Product Manager", "jobUrl": "https://job/ashby", "descriptionHtml": "<p>Digital health</p>"}]})
        workable = pipeline.parse_workable_jobs({"jobs": [{"title": "Design Assurance Engineer", "url": "https://job/workable", "description": "V&V"}]})
        smartrecruiters = pipeline.parse_smartrecruiters_jobs({"content": [{"name": "Medical Device QA Manager", "url": "https://job/sr", "description": "Quality systems"}]})
        recruitee = pipeline.parse_recruitee_jobs({"offers": [{"title": "Clinical Validation Lead", "careers_url": "https://job/recruitee", "description": "Diagnostics"}]})

        self.assertEqual([jobs[0].title for jobs in [greenhouse, lever, ashby, workable, smartrecruiters, recruitee]], [
            "Quality Engineer",
            "Regulatory Affairs Lead",
            "Clinical AI Product Manager",
            "Design Assurance Engineer",
            "Medical Device QA Manager",
            "Clinical Validation Lead",
        ])

    def test_greenhouse_search_parser_extracts_job_urls_and_tokens(self):
        html = """
        <html><body>
          <a href="https://boards.greenhouse.io/novascan/jobs/123">Regulatory Affairs Manager</a>
          <a href="/url?q=https%3A%2F%2Fboards.greenhouse.io%2Fpulsedx%2Fjobs%2F456">Quality Engineer</a>
          https://job-boards.greenhouse.io/clearpath/jobs/789
        </body></html>
        """

        urls = pipeline.extract_greenhouse_job_urls(html)

        self.assertEqual(
            [pipeline.greenhouse_board_token_from_url(url) for url in urls],
            ["novascan", "pulsedx", "clearpath"],
        )

    def test_greenhouse_discovery_creates_company_level_hiring_signal(self):
        source = pipeline.Source("Greenhouse job boards", "Jobs", "https://www.greenhouse.com/", "Global", "Medium", "Weekly", "Company careers page search", "Fixture", "greenhouse_jobs")
        search_html = """
        <html><body>
          <a href="https://boards.greenhouse.io/novascan/jobs/123">Regulatory Affairs Manager</a>
          <a href="https://boards.greenhouse.io/novascan/jobs/456">Quality Engineer</a>
        </body></html>
        """
        board_fixture = {"name": "NovaScan Health", "content": "<p>AI imaging company</p>"}
        jobs_fixture = {
            "jobs": [
                {"title": "Regulatory Affairs Manager", "absolute_url": "https://job/reg", "content": "<p>FDA medical device submissions</p>"},
                {"title": "Quality Engineer", "absolute_url": "https://job/qa", "content": "<p>Design controls and V&V</p>"},
            ]
        }

        with patch.object(pipeline, "fetch_raw_text", return_value=(search_html, None)), patch.object(pipeline, "fetch_json_url", side_effect=[(board_fixture, None), (jobs_fixture, None)]):
            discovery_hits, trigger_events, result = pipeline.run_greenhouse_discovery(source, ["fixture query"])

        self.assertEqual([hit.company for hit in discovery_hits], ["NovaScan Health"])
        self.assertEqual([event.trigger_type for event in trigger_events], ["Hiring signal"])
        self.assertIn("Regulatory Affairs Manager", discovery_hits[0].discovery_rationale)
        self.assertIn("Quality Engineer", trigger_events[0].trigger_event)
        self.assertIn("1 search queries", result)
        self.assertIn("1 board tokens fetched", result)

    def test_greenhouse_discovery_ignores_irrelevant_jobs(self):
        source = pipeline.Source("Greenhouse job boards", "Jobs", "https://www.greenhouse.com/", "Global", "Medium", "Weekly", "Company careers page search", "Fixture", "greenhouse_jobs")
        search_html = '<a href="https://boards.greenhouse.io/genericco/jobs/123">Software Engineer</a>'
        board_fixture = {"name": "GenericCo"}
        jobs_fixture = {"jobs": [{"title": "Software Engineer", "absolute_url": "https://job/software", "content": "<p>Build internal tools.</p>"}]}

        with patch.object(pipeline, "fetch_raw_text", return_value=(search_html, None)), patch.object(pipeline, "fetch_json_url", side_effect=[(board_fixture, None), (jobs_fixture, None)]):
            discovery_hits, trigger_events, result = pipeline.run_greenhouse_discovery(source, ["fixture query"])

        self.assertEqual(discovery_hits, [])
        self.assertEqual(trigger_events, [])
        self.assertIn("1 boards with no matching jobs", result)

    def test_job_board_adapter_ignores_irrelevant_engineering_without_health_context(self):
        source = pipeline.Source("Lever job boards", "Jobs", "https://www.lever.co/", "Global", "Medium", "Weekly", "Company careers page search", "Fixture", "lever_jobs")
        registry = {
            "GenericCo": {
                "aliases": ["GenericCo"],
                "website": "https://generic.example",
                "geography": "US",
                "product_type": "Workflow software",
                "job_boards": [{"platform": "lever", "account": "genericco"}],
            }
        }
        fixture = [{"text": "Software Engineer", "hostedUrl": "https://job/software", "description": "Build internal tools."}]

        with patch.object(pipeline, "COMPANY_REGISTRY", registry), patch.object(pipeline, "fetch_json_url", return_value=(fixture, None)):
            discovery_hits, trigger_events, result = pipeline.run_job_board_adapter(source, "lever")

        self.assertEqual(discovery_hits, [])
        self.assertEqual(trigger_events, [])
        self.assertIn("1 boards with no matching jobs", result)

    def test_job_board_adapter_reports_no_config(self):
        source = pipeline.Source("Ashby job boards", "Jobs", "https://www.ashbyhq.com/", "Global", "Medium", "Weekly", "Company careers page search", "Fixture", "ashby_jobs")

        with patch.object(pipeline, "COMPANY_REGISTRY", {}):
            discovery_hits, trigger_events, result = pipeline.run_job_board_adapter(source, "ashby")

        self.assertEqual(discovery_hits, [])
        self.assertEqual(trigger_events, [])
        self.assertIn("No registry companies configured", result)

    def test_parse_biospace_jobs_extracts_company_role_and_url(self):
        html = """
        <ul id="listing" class="lister cf block">
          <li class="lister__item cf" id="item-3052525">
            <div class="lister__details cf js-clickable">
              <h3 class="lister__header"><a href="/job/3052525/manager-regulatory-affairs/"><span>Manager, Regulatory Affairs</span></a></h3>
              <ul class="lister__meta">
                <li class="lister__meta-item lister__meta-item--location">Boca Raton, FL</li>
                <li class="lister__meta-item lister__meta-item--recruiter">ADMA Biologics</li>
              </ul>
              <p class="lister__description js-clamp-2">FDA medical device regulatory submissions.</p>
            </div>
          </li>
        </ul>
        """

        leads = pipeline.parse_biospace_jobs(html, "https://jobs.biospace.com/jobs/?keywords=regulatory+affairs", "regulatory affairs")

        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0].company, "ADMA Biologics")
        self.assertEqual(leads[0].posting.title, "Manager, Regulatory Affairs")
        self.assertEqual(leads[0].posting.url, "https://jobs.biospace.com/job/3052525/manager-regulatory-affairs/")

    def test_biospace_adapter_discovers_companies_from_role_search(self):
        source = pipeline.Source("BioSpace jobs", "Jobs", "https://www.biospace.com/jobs/", "US/global", "High", "Weekly", "BioSpace role search", "Fixture", "biospace_jobs")
        html = """
        <ul id="listing" class="lister cf block">
          <li class="lister__item cf" id="item-1">
            <div class="lister__details cf js-clickable">
              <h3 class="lister__header"><a href="/job/1/regulatory-affairs-manager/"><span>Regulatory Affairs Manager</span></a></h3>
              <ul class="lister__meta">
                <li class="lister__meta-item lister__meta-item--location">Boston, MA</li>
                <li class="lister__meta-item lister__meta-item--recruiter">NovaScan Health</li>
              </ul>
              <p class="lister__description js-clamp-2">FDA medical device and diagnostic submissions.</p>
            </div>
          </li>
          <li class="lister__item cf" id="item-2">
            <div class="lister__details cf js-clickable">
              <h3 class="lister__header"><a href="/job/2/software-engineer/"><span>Software Engineer</span></a></h3>
              <ul class="lister__meta">
                <li class="lister__meta-item lister__meta-item--location">Remote</li>
                <li class="lister__meta-item lister__meta-item--recruiter">GenericCo</li>
              </ul>
              <p class="lister__description js-clamp-2">Build internal tools.</p>
            </div>
          </li>
        </ul>
        """

        with patch.object(pipeline, "fetch_raw_text", return_value=(html, None)):
            discovery_hits, trigger_events, result = pipeline.run_biospace_jobs(source, ["regulatory affairs"])

        self.assertEqual([hit.company for hit in discovery_hits], ["NovaScan Health"])
        self.assertEqual([event.trigger_type for event in trigger_events], ["Hiring signal"])
        self.assertIn("Regulatory Affairs Manager", discovery_hits[0].discovery_rationale)
        self.assertIn("1 search queries", result)

    def test_parse_builtin_jobs_extracts_company_role_and_url(self):
        html = """
        <div id="job-card-9668948" data-id="job-card">
          <a href="/company/optum" data-id="company-title"><span>Optum</span></a>
          <h2><a href="/job/senior-director-actuarial-regulatory-affairs-pricing-underwriting/9668948" data-id="job-card-title">Senior Director, Actuarial &amp; Regulatory Affairs</a></h2>
          <i class="fa-regular fa-location-dot"></i></div><div><span class="font-barlow text-gray-04">Dublin, IRL</span></div>
        </div>
        """

        leads = pipeline.parse_builtin_jobs(html, "https://builtin.com/jobs?search=regulatory+affairs", "regulatory affairs")

        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0].company, "Optum")
        self.assertEqual(leads[0].posting.title, "Senior Director, Actuarial & Regulatory Affairs")
        self.assertEqual(leads[0].posting.url, "https://builtin.com/job/senior-director-actuarial-regulatory-affairs-pricing-underwriting/9668948")

    def test_builtin_adapter_discovers_companies_from_role_search(self):
        source = pipeline.Source("Built In jobs", "Jobs", "https://builtin.com/jobs", "US", "Medium", "Weekly", "Built In role search", "Technology startup hiring signals, including healthtech and AI companies.", "builtin_jobs")
        html = """
        <div id="job-card-1" data-id="job-card">
          <a href="/company/pulsedx" data-id="company-title"><span>PulseDx</span></a>
          <h2><a href="/job/quality-engineer-healthcare/1" data-id="job-card-title">Quality Engineer, Healthcare AI</a></h2>
          <i class="fa-regular fa-location-dot"></i></div><div><span class="font-barlow text-gray-04">Remote</span></div>
        </div>
        <div id="job-card-2" data-id="job-card">
          <a href="/company/genericco" data-id="company-title"><span>GenericCo</span></a>
          <h2><a href="/job/account-executive/2" data-id="job-card-title">Account Executive</a></h2>
        </div>
        """

        with patch.object(pipeline, "fetch_raw_text", return_value=(html, None)):
            discovery_hits, trigger_events, result = pipeline.run_builtin_jobs(source, ["quality engineer"])

        self.assertEqual([hit.company for hit in discovery_hits], ["PulseDx"])
        self.assertEqual([event.trigger_type for event in trigger_events], ["Hiring signal"])
        self.assertIn("Quality Engineer", discovery_hits[0].discovery_rationale)
        self.assertIn("1 search queries", result)

    def test_parse_nhs_jobs_extracts_employer_role_and_url(self):
        html = """
        <ul class="nhsuk-list search-results">
          <li class="nhsuk-list-panel search-result" data-test="search-result">
            <h2><a href="/candidate/jobadvert/C9444-26-0330?keyword=clinical%20safety&amp;language=en" data-test="search-result-job-title">Clinical Safety Officer</a></h2>
            <div class="nhsuk-u-margin-bottom-4" data-test="search-result-location">
              <h3>Coventry and Warwickshire Partnership Trust
                <div class="location-font-size">Coventry CV6 6NY</div>
              </h3>
            </div>
            <li data-test="search-result-jobType">Contract type: <strong>Fixed-Term</strong></li>
          </li>
        </ul>
        """

        leads = pipeline.parse_nhs_jobs(html, "https://www.jobs.nhs.uk/candidate/search/results?keyword=clinical+safety", "clinical safety")

        self.assertEqual(len(leads), 1)
        self.assertEqual(leads[0].company, "Coventry and Warwickshire Partnership Trust")
        self.assertEqual(leads[0].posting.title, "Clinical Safety Officer")
        self.assertEqual(leads[0].posting.url, "https://www.jobs.nhs.uk/candidate/jobadvert/C9444-26-0330?keyword=clinical%20safety&language=en")

    def test_nhs_adapter_discovers_provider_organisations_from_role_search(self):
        source = pipeline.Source("NHS Jobs", "Jobs", "https://www.jobs.nhs.uk/", "UK", "Medium", "Weekly", "NHS role search", "Fixture", "nhs_jobs")
        html = """
        <ul class="nhsuk-list search-results">
          <li class="nhsuk-list-panel search-result" data-test="search-result">
            <h2><a href="/candidate/jobadvert/C1" data-test="search-result-job-title">Clinical Safety Officer</a></h2>
            <div data-test="search-result-location"><h3>North Example NHS Trust<div class="location-font-size">London</div></h3></div>
          </li>
          <li class="nhsuk-list-panel search-result" data-test="search-result">
            <h2><a href="/candidate/jobadvert/C2" data-test="search-result-job-title">Catering Assistant</a></h2>
            <div data-test="search-result-location"><h3>Generic Hospital<div class="location-font-size">Leeds</div></h3></div>
          </li>
        </ul>
        """

        with patch.object(pipeline, "fetch_raw_text", return_value=(html, None)):
            discovery_hits, trigger_events, result = pipeline.run_nhs_jobs(source, ["clinical safety"])

        self.assertEqual([hit.company for hit in discovery_hits], ["North Example NHS Trust"])
        self.assertEqual([event.trigger_type for event in trigger_events], ["Hiring signal"])
        self.assertIn("Clinical Safety Officer", discovery_hits[0].discovery_rationale)
        self.assertIn("1 search queries", result)

    def test_parse_google_news_rss(self):
        results = pipeline.parse_google_news_rss(RSS_FIXTURE, "MedTech AI Funding")

        self.assertEqual(len(results), 4)
        self.assertEqual(results[0].query, "MedTech AI Funding")
        self.assertEqual(results[0].publisher, "MedTech Dive")
        self.assertEqual(results[0].link, "https://news.google.com/rss/articles/novascan")
        self.assertEqual(pipeline.article_year_from_pubdate(results[0].published_at), "2026")

    def test_search_evidence_extraction_classification_and_dedupe(self):
        source = pipeline.Source("Google News / web funding search", "News/search", "https://news.google.com/search", "US/EU/global", "High", "Weekly", "Google News RSS query", "Search fixture", "google_news_search")
        results = pipeline.parse_google_news_rss(RSS_FIXTURE, "MedTech AI Funding")
        discovery_hits, trigger_events = pipeline.build_google_news_evidence(source, results)

        self.assertEqual([hit.company for hit in discovery_hits], ["NovaScan Health", "PulseDx", "ClearPath Medical"])
        self.assertEqual([event.company for event in trigger_events], ["NovaScan Health", "PulseDx"])
        self.assertEqual(trigger_events[0].trigger_type, "Funding")
        self.assertEqual(trigger_events[1].trigger_type, "Regulatory clearance")
        self.assertIn("query: MedTech AI Funding", discovery_hits[0].matched_terms)
        self.assertEqual(discovery_hits[0].article_year, "2026")

    def test_workbook_preserves_search_traceability(self):
        source = pipeline.Source("Google News / web funding search", "News/search", "https://news.google.com/search", "US/EU/global", "High", "Weekly", "Google News RSS query", "Search fixture", "google_news_search")
        results = pipeline.parse_google_news_rss(RSS_FIXTURE, "MedTech AI Funding")
        discovery_hits, search_triggers = pipeline.build_google_news_evidence(source, results)
        companies = pipeline.normalize_companies(discovery_hits)
        trigger_events = pipeline.attach_trigger_events(companies, search_triggers)
        pipeline.mark_primary_triggers(companies)

        with tempfile.TemporaryDirectory() as temp_dir:
            original_out = pipeline.OUT
            try:
                pipeline.OUT = Path(temp_dir) / "traceability.xlsx"
                workbook_path = pipeline.write_workbook(companies, discovery_hits, trigger_events, [["Google News / web funding search", "News/search", source.url, "Fetched", "fixture"]])
                wb = load_workbook(workbook_path)
            finally:
                pipeline.OUT = original_out

        discovery_row = list(wb["Discovery Hits"].iter_rows(min_row=2, values_only=True))[0]
        trigger_row = list(wb["Trigger Log"].iter_rows(min_row=2, values_only=True))[0]
        lead_rows = list(wb["Leads"].iter_rows(min_row=2, values_only=True))
        novascan_lead = [row for row in lead_rows if row[0] == "NovaScan Health"][0]

        self.assertEqual(discovery_row[1], "Google News / web funding search: MedTech AI Funding")
        self.assertEqual(discovery_row[3], "https://news.google.com/rss/articles/novascan")
        self.assertEqual(discovery_row[4], "2026")
        self.assertEqual(trigger_row[4], "https://news.google.com/rss/articles/novascan")
        self.assertEqual(novascan_lead[14], "2026")
        self.assertEqual(novascan_lead[15], "Verified trigger")

    def test_workbook_uses_expected_sheet_order_without_weekly_review(self):
        hit = pipeline.DiscoveryHit(
            "NovaScan Health",
            "Fixture Accelerator",
            "Accelerator",
            "https://example.com/novascan",
            "Fixture accelerator extraction.",
            website="https://novascan.example",
        )
        companies = pipeline.normalize_companies([hit])

        with tempfile.TemporaryDirectory() as temp_dir:
            original_out = pipeline.OUT
            try:
                pipeline.OUT = Path(temp_dir) / "sheets.xlsx"
                workbook_path = pipeline.write_workbook(companies, [hit], [], [])
                wb = load_workbook(workbook_path)
            finally:
                pipeline.OUT = original_out

        self.assertEqual(wb.sheetnames, pipeline.EXPECTED_WORKBOOK_SHEETS)
        self.assertNotIn("Weekly Review", wb.sheetnames)

    def test_workbook_includes_accelerator_metadata_columns(self):
        hit = pipeline.DiscoveryHit(
            company="NovaScan Health",
            source_name="Fixture Accelerator",
            source_type="Accelerator",
            discovery_url="https://example.com/novascan",
            discovery_rationale="Fixture accelerator extraction.",
            product_type="Diagnostics / imaging",
            accelerator_program="Fixture Accelerator",
            cohort_label="Fixture 2026 cohort",
            cohort_year="2026",
            category_or_track="Diagnostics",
            company_description="AI imaging triage for hospitals.",
        )
        companies = pipeline.normalize_companies([hit])

        with tempfile.TemporaryDirectory() as temp_dir:
            original_out = pipeline.OUT
            try:
                pipeline.OUT = Path(temp_dir) / "metadata.xlsx"
                workbook_path = pipeline.write_workbook(companies, [hit], [], [["Fixture Accelerator", "Accelerator", hit.discovery_url, "Fetched", "fixture"]])
                wb = load_workbook(workbook_path)
            finally:
                pipeline.OUT = original_out

        discovery_headers = [cell.value for cell in wb["Discovery Hits"][1]]
        lead_headers = [cell.value for cell in wb["Leads"][1]]
        lead_row = list(wb["Leads"].iter_rows(min_row=2, values_only=True))[0]
        discovery_row = list(wb["Discovery Hits"].iter_rows(min_row=2, values_only=True))[0]

        self.assertIn("Accelerator program", discovery_headers)
        self.assertIn("Article year", discovery_headers)
        self.assertIn("Company description", lead_headers)
        self.assertIn("Source URL", lead_headers)
        self.assertEqual(discovery_row[10], "Fixture Accelerator")
        self.assertEqual(discovery_row[12], "2026")
        self.assertEqual(lead_row[10], "Fixture Accelerator")
        self.assertEqual(lead_row[12], "2026")

    def test_primary_discovery_prefers_richer_cohort_metadata(self):
        generic_hit = pipeline.DiscoveryHit(
            company="Rosalind Dx",
            source_name="MedTech Innovator",
            source_type="Accelerator",
            discovery_url="https://example.com/generic",
            discovery_rationale="Generic portfolio row.",
            cohort_label="MedTech Innovator portfolio",
        )
        cohort_hit = pipeline.DiscoveryHit(
            company="Rosalind Dx",
            source_name="MedTech Innovator",
            source_type="Accelerator",
            discovery_url="https://example.com/apac",
            discovery_rationale="Specific APAC cohort row.",
            website="https://www.rosalinddx.com",
            cohort_label="MedTech Innovator APAC 2025",
            cohort_year="2025",
            category_or_track="Diagnostics",
            company_description="Accessible prenatal testing.",
        )
        record = pipeline.CompanyRecord(company="Rosalind Dx", discovery_hits=[generic_hit, cohort_hit])

        self.assertEqual(pipeline.primary_discovery(record).discovery_url, "https://example.com/apac")

    def test_rule_classification_covers_core_personas(self):
        cases = [
            (
                "AcceleratorCo",
                pipeline.DiscoveryHit("AcceleratorCo", "Fixture Accelerator", "Accelerator", "https://example.com/a", "Current accelerator cohort.", product_type="AI medical device"),
                [],
                "Early startup",
                "Accelerator/cohort",
            ),
            (
                "FundedCo",
                pipeline.DiscoveryHit("FundedCo", "Funding News", "News/search", "https://example.com/f", "Raised a Series A for diagnostic AI.", product_type="Diagnostics"),
                [pipeline.TriggerEvent("FundedCo", "Funding", "Raised a Series A.", "Funding News", "https://example.com/f")],
                "Funded startup",
                "Funding trigger",
            ),
            (
                "JobsCo",
                pipeline.DiscoveryHit("JobsCo", "Jobs", "Jobs", "https://example.com/j", "Hiring a regulatory affairs lead.", product_type="SaMD"),
                [pipeline.TriggerEvent("JobsCo", "Hiring signal", "Hiring regulatory affairs and QA.", "Jobs", "https://example.com/j")],
                "Jobs-led capability gap",
                "Hiring gap",
            ),
            (
                "RegCo",
                pipeline.DiscoveryHit("RegCo", "FDA", "Regulatory database", "https://example.com/r", "FDA clearance listing.", product_type="Medical device"),
                [pipeline.TriggerEvent("RegCo", "Regulatory clearance", "Received FDA clearance.", "FDA", "https://example.com/r")],
                "Regulatory-led opportunity",
                "Regulatory pathway",
            ),
            (
                "SpinoutCo",
                pipeline.DiscoveryHit("SpinoutCo", "University", "University/spinout", "https://example.com/u", "University spinout.", product_type="Medical device"),
                [],
                "University/spinout",
                "Medical device",
            ),
        ]

        for company, hit, triggers, persona, secondary_tag in cases:
            record = pipeline.CompanyRecord(company=company, product_type=hit.product_type, discovery_hits=[hit], triggers=triggers)
            enrichment = pipeline.classify_company_rules(record)

            self.assertEqual(enrichment.persona, persona)
            self.assertEqual(enrichment.secondary_tag, secondary_tag)
            self.assertFalse(enrichment.llm_used)
            self.assertEqual(enrichment.method, "rules")

    def test_classification_marks_missing_llm_fallback(self):
        hit = pipeline.DiscoveryHit("NovaScan Health", "Fixture", "Accelerator", "https://example.com", "AI imaging accelerator company.")
        record = pipeline.CompanyRecord(company="NovaScan Health", discovery_hits=[hit])

        with patch.dict("os.environ", {}, clear=True):
            enrichment = pipeline.classify_company(record)

        self.assertFalse(enrichment.llm_used)
        self.assertEqual(enrichment.fallback_reason, "llm_not_configured")

    def test_classification_falls_back_for_llm_errors_and_invalid_outputs(self):
        hit = pipeline.DiscoveryHit("NovaScan Health", "Fixture", "Accelerator", "https://example.com", "AI imaging accelerator company.")
        record = pipeline.CompanyRecord(company="NovaScan Health", discovery_hits=[hit])

        with patch.dict("os.environ", {"BBT_LEAD_ENRICHMENT_API_KEY": "fixture"}, clear=True), patch.object(pipeline, "load_cached_llm_enrichment", return_value=None), patch("bbt_bizdev.pipeline._call_lead_enrichment_llm", side_effect=RuntimeError("boom")):
            error_fallback = pipeline.classify_company(record)

        with patch.dict("os.environ", {"BBT_LEAD_ENRICHMENT_API_KEY": "fixture"}, clear=True), patch.object(pipeline, "load_cached_llm_enrichment", return_value=None), patch("bbt_bizdev.pipeline._call_lead_enrichment_llm", side_effect=ValueError("invalid_json")):
            json_fallback = pipeline.classify_company(record)

        with patch.dict("os.environ", {"BBT_LEAD_ENRICHMENT_API_KEY": "fixture"}, clear=True), patch.object(pipeline, "load_cached_llm_enrichment", return_value={"persona": "Bad", "primary_quadrant": "Advisory", "secondary_tag": "SaMD/AI", "pain_hypothesis": "x", "value_prop": "x", "outreach_angle": "x", "confidence": 0.5, "rationale": "x"}):
            taxonomy_fallback = pipeline.classify_company(record)

        self.assertEqual(error_fallback.fallback_reason, "llm_error")
        self.assertEqual(json_fallback.fallback_reason, "invalid_json")
        self.assertEqual(taxonomy_fallback.fallback_reason, "invalid_taxonomy")
        self.assertFalse(error_fallback.llm_used)

    def test_workbook_includes_enrichment_columns_and_varied_personas(self):
        accelerator_hit = pipeline.DiscoveryHit("NovaScan Health", "Fixture Accelerator", "Accelerator", "https://example.com/novascan", "Current accelerator cohort.", product_type="AI medical device", website="https://novascan.example")
        jobs_hit = pipeline.DiscoveryHit("PulseDx", "Jobs", "Jobs", "https://example.com/pulsedx", "Hiring regulatory affairs and QA.", product_type="SaMD", website="https://pulsedx.example")
        companies = pipeline.normalize_companies([accelerator_hit, jobs_hit])
        trigger_events = pipeline.attach_trigger_events(
            companies,
            [pipeline.TriggerEvent("PulseDx", "Hiring signal", "Hiring regulatory affairs and QA.", "Jobs", "https://example.com/pulsedx")],
        )
        pipeline.mark_primary_triggers(companies)

        with tempfile.TemporaryDirectory() as temp_dir, patch.dict("os.environ", {}, clear=True):
            original_out = pipeline.OUT
            try:
                pipeline.OUT = Path(temp_dir) / "enrichment.xlsx"
                workbook_path = pipeline.write_workbook(companies, [accelerator_hit, jobs_hit], trigger_events, [["Fixture", "Jobs", jobs_hit.discovery_url, "Fetched", "fixture"]])
                wb = load_workbook(workbook_path)
            finally:
                pipeline.OUT = original_out

        headers = [cell.value for cell in wb["Leads"][1]]
        rows = list(wb["Leads"].iter_rows(min_row=2, values_only=True))
        personas = {row[22] for row in rows}

        self.assertNotIn("Value prop", headers)
        self.assertNotIn("Outreach angle", headers)
        self.assertNotIn("LLM used", headers)
        self.assertNotIn("Fallback reason", headers)
        self.assertNotIn("Evidence recency", headers)
        self.assertNotIn("Legacy score", headers)
        self.assertNotIn("Classification confidence", headers)
        self.assertIn("Company website", headers)
        self.assertIn("Evidence year", headers)
        self.assertIn("Trigger type", headers)
        self.assertIn("Geography", headers)
        self.assertIn("Company stage", headers)
        self.assertIn("Product area", headers)
        self.assertIn("Hiring signal", headers)
        self.assertIn("Funding stage", headers)
        self.assertIn("Early startup", personas)
        self.assertIn("Jobs-led capability gap", personas)
        self.assertNotIn("AI/SaMD or healthtech company from approved source", personas)
        self.assertEqual(headers[18], "Source URL")
        self.assertEqual(headers[24], "LinkedIn company URL")
        self.assertEqual({row[0]: row[1] for row in rows}, {"NovaScan Health": "https://novascan.example", "PulseDx": "https://pulsedx.example"})

    def test_lead_filter_fields_use_latest_evidence_and_explicit_signals(self):
        older = pipeline.DiscoveryHit("NovaScan", "Accelerator", "Accelerator", "https://example.com/2023", "Accelerator cohort.", cohort_year="2023", geography="Ireland")
        latest = pipeline.DiscoveryHit("NovaScan", "Funding News", "News/search", "https://example.com/2025", "Raised a Series A for diagnostic imaging AI.", article_year="2025", geography="Ireland")
        trigger = pipeline.TriggerEvent("NovaScan", "Funding", "Raised a Series A.", "Funding News", latest.discovery_url)
        record = pipeline.CompanyRecord(company="NovaScan", geography="Ireland", discovery_hits=[older, latest], triggers=[trigger])
        trigger.trigger_role = "Primary"

        fields = pipeline.lead_filter_fields(record, pipeline.classify_company_rules(record))

        self.assertEqual(fields["Evidence year"], "2025")
        self.assertEqual(fields["Evidence basis"], "Article year")
        self.assertEqual(fields["Trigger type"], "Funding")
        self.assertEqual(fields["Funding stage"], "Series A")
        self.assertEqual(fields["Product area"], "AI / SaMD")
        self.assertEqual(fields["Hiring signal"], "No")

    def test_geography_is_normalized_to_filter_regions(self):
        cases = {
            "San Francisco, CA, USA": "US",
            "Toronto, ON, Canada": "Canada",
            "London, England, United Kingdom": "UK",
            "Galway, County Galway, Ireland": "Ireland",
            "Paris, Ile-de-France, France": "Europe",
            "Singapore, Singapore": "Asia-Pacific",
            "Sydney, NSW, Australia": "Australia/New Zealand",
            "Israel/US": "Middle East",
            "EU/US": "Europe",
            "US/EU/global": "US",
            "Remote": "Unknown",
        }
        for raw, expected in cases.items():
            self.assertEqual(pipeline.normalize_geography_region(raw), expected)

    def test_digitalhealth_london_adapter_paginates_directory_cards(self):
        source = pipeline.Source("DigitalHealth.London Accelerator", "Accelerator", "https://digitalhealth.london/programmes/accelerator/", "UK", "High", "Annual", "Cohort extraction", "NHS-facing digital health.", "digitalhealth_london")
        page_1 = """
        <html><body>
          <a href="/innovation-directory/profile/alpha-care">Alpha Care Company Remote monitoring platform for NHS pathways.</a>
          <a href="/innovation-directory/profile/beta-dx">BetaDx Company Diagnostic decision support for urgent care.</a>
          <a href="https://digitalhealth.london/innovation-directory/companies/page/2">Older posts</a>
        </body></html>
        """
        page_2 = '<html><body><a href="/innovation-directory/profile/gamma-ehr">Gamma EHR Company EHR workflow automation for hospital teams.</a></body></html>'
        alpha_profile = "<html><body>Cohort: 2025 Sector: Digital health</body></html>"
        beta_profile = "<html><body>2024 Accelerator Technology: Diagnostics</body></html>"
        gamma_profile = "<html><body>Launchpad - 2023</body></html>"

        with patch.object(pipeline, "fetch_raw_text", side_effect=[(page_1, None), (alpha_profile, None), (beta_profile, None), (page_2, None), (gamma_profile, None)]):
            discovery_hits, trigger_events, result = pipeline.run_digitalhealth_london(source)

        self.assertEqual([hit.company for hit in discovery_hits], ["Alpha Care", "BetaDx", "Gamma EHR"])
        self.assertEqual([hit.cohort_year for hit in discovery_hits], ["2025", "2024", "2023"])
        self.assertIn("Remote monitoring", discovery_hits[0].company_description)
        self.assertEqual(len(trigger_events), 3)
        self.assertIn("2 directory pages", result)
        self.assertIn("3 profiles fetched", result)

    def test_digitalhealth_london_parser_can_enrich_from_profile_html(self):
        source = pipeline.Source("DigitalHealth.London Accelerator", "Accelerator", "https://digitalhealth.london/programmes/accelerator/", "UK", "High", "Annual", "Cohort extraction", "NHS-facing digital health.", "digitalhealth_london")
        page = '<html><body><a href="/innovation-directory/profile/alpha-care">Alpha Care Company Remote monitoring.</a></body></html>'
        profile_url = "https://digitalhealth.london/innovation-directory/profile/alpha-care"
        profile_html = '<html><head><meta property="og:description" content="Remote monitoring platform for NHS pathways."></head><body>Sector: Digital health Technology: AI Cohort 2026</body></html>'

        hits = pipeline.parse_digitalhealth_london_page(source, page, "https://digitalhealth.london/innovation-directory/companies", {profile_url: profile_html})

        self.assertEqual(hits[0].cohort_year, "2026")
        self.assertIn("Digital health", hits[0].category_or_track)

    def test_medtech_innovator_adapter_flags_incomplete_current_cohort(self):
        source = pipeline.Source("MedTech Innovator", "Accelerator", "https://medtechinnovator.org/2026cohort/", "US/EU/global", "High", "Quarterly", "Cohort extraction", "Medtech source.", "medtech_innovator")
        cohort_html = """
        <html><body>
          <p>65 companies selected for the 2026 cohort.</p>
          <a href="https://pro.innovator.org/showcase/2026cohort">Showcase</a>
        </body></html>
        """
        showcase_html = """
        <html><body>
          <h2>Diagnostics</h2>
          <a href="https://example.com/heartscan">HeartScan</a>
          <a href="https://example.com/neuroflow">NeuroFlowx</a>
        </body></html>
        """

        with patch.object(pipeline, "fetch_raw_text", side_effect=[(cohort_html, None), (showcase_html, None)]), patch.object(pipeline, "fetch_medtech_innovator_pory_records", return_value=([], [])):
            discovery_hits, _, result = pipeline.run_medtech_innovator(source)

        self.assertEqual([hit.company for hit in discovery_hits], ["HeartScan", "NeuroFlowx"])
        self.assertEqual(discovery_hits[0].cohort_year, "2026")
        self.assertIn("Diagnostics", discovery_hits[0].category_or_track)
        self.assertIn("INCOMPLETE current-cohort extraction", result)

    def test_medtech_innovator_adapter_extracts_pory_portfolio_records(self):
        source = pipeline.Source("MedTech Innovator", "Accelerator", "https://medtechinnovator.org/2026cohort/", "US/EU/global", "High", "Quarterly", "Cohort extraction", "Medtech source.", "medtech_innovator")
        records = [
            {
                "id": "rec1",
                "fields": {
                    "Company": "2morrow",
                    "Website": "https://www.2morrowinc.com/",
                    "Year.": "2017",
                    "Program.": "Accelerator-US",
                    "Product Short Description": "Clinically-tested mobile behavior change platform.",
                    "Thematic Categories": ["Digital Therapeutics", "Chronic Disease Management"],
                    "Company Country/Territory (Old Field)": "United States",
                },
            }
        ]

        hits = pipeline.parse_medtech_innovator_pory_records(source, records)

        self.assertEqual(hits[0].company, "2morrow")
        self.assertEqual(hits[0].discovery_url, "https://app.pory.dev/data/66eb41bc87c0d05ea2b410b8/records/rec1")
        self.assertEqual(hits[0].cohort_year, "2017")
        self.assertEqual(hits[0].website, "https://www.2morrowinc.com/")
        self.assertIn("Digital Therapeutics", hits[0].category_or_track)

    def test_mayo_accelerate_adapter_extracts_descriptions_from_headings(self):
        source = pipeline.Source("Mayo Clinic Platform Accelerate", "Accelerator", "https://www.mayoclinicplatform.org/accelerate/", "US/global", "High", "Quarterly", "Cohort extraction", "AI digital health.", "mayo_accelerate")
        html = """
        <html><body>
          <h2>ClinicAI</h2><p>AI platform for clinical workflow automation and patient risk triage.</p>
          <h2>VitalsCloud</h2><p>Remote monitoring software for chronic care teams.</p>
        </body></html>
        """

        hits = pipeline.parse_mayo_accelerate_page(source, html, "https://example.com/accelerate-2026")

        self.assertEqual([hit.company for hit in hits], ["ClinicAI", "VitalsCloud"])
        self.assertEqual(hits[0].cohort_year, "2026")
        self.assertIn("clinical workflow", hits[0].company_description)

    def test_mayo_accelerate_adapter_uses_live_reader_when_direct_fetch_is_blocked(self):
        source = pipeline.Source("Mayo Clinic Platform Accelerate", "Accelerator", "https://www.mayoclinicplatform.org/accelerate/", "US/global", "High", "Quarterly", "Cohort extraction", "AI digital health.", "mayo_accelerate")
        reader_markdown = """
        Title: Accelerate Cohort Landing Page - Mayo Clinic Platform
        URL Source: https://www.mayoclinicplatform.org/focus-areas/digital-health/accelerate/accelerate-cohort-landing-page/
        Markdown Content:
        February 2026
        ## Meet the Newest Cohort of Innovative Health Tech Startups

        [![Image 1](https://cdn.example.com/100ms.jpg)](https://100ms.ai/)
        **100ms** builds AI agents for patient access, helping patients automate intake and scheduling for specialty practices.

        ![Image 2](https://cdn.example.com/wfh.jpg)
        "WFH: Wellness from Home" is a continuous health monitoring platform for elderly remote patient monitoring.
        """

        with patch.object(pipeline, "fetch_raw_text", side_effect=[("", "HTTP Error 403: Forbidden"), (reader_markdown, None)]):
            discovery_hits, trigger_events, result = pipeline.run_mayo_accelerate(source)

        self.assertEqual([hit.company for hit in discovery_hits], ["100ms", "WFH: Wellness from Home"])
        self.assertEqual(discovery_hits[0].website, "https://100ms.ai/")
        self.assertIn("live reader page", discovery_hits[0].matched_terms)
        self.assertEqual(len(trigger_events), len(discovery_hits))
        self.assertIn("HTTP Error 403", result)

    def test_mayo_accelerate_adapter_reports_incomplete_when_all_live_fetches_fail(self):
        source = pipeline.Source("Mayo Clinic Platform Accelerate", "Accelerator", "https://www.mayoclinicplatform.org/accelerate/", "US/global", "High", "Quarterly", "Cohort extraction", "AI digital health.", "mayo_accelerate")

        with patch.object(pipeline, "fetch_raw_text", return_value=("", "HTTP Error 403: Forbidden")):
            discovery_hits, trigger_events, result = pipeline.run_mayo_accelerate(source)

        self.assertEqual(discovery_hits, [])
        self.assertEqual(trigger_events, [])
        self.assertIn("INCOMPLETE Mayo extraction", result)

    def test_eit_health_catapult_adapter_extracts_winners_and_tracks(self):
        source = pipeline.Source("EIT Health Catapult", "Accelerator", "https://eithealth.eu/programmes/catapult/", "EU", "High", "Annual", "Finalist extraction", "European health startups.", "eit_health_catapult")
        html = """
        <html><body>
          <h2>Digital Health winners</h2>
          <img alt="DeepEye" src="/deepeye.jpg">
          <h2>MedTech winners</h2>
          <a href="https://example.com/acorai">Acorai</a>
        </body></html>
        """

        hits = pipeline.parse_eit_health_catapult_page(source, html, "https://eithealth.eu/programmes/catapult/")

        self.assertEqual([hit.company for hit in hits], ["DeepEye", "Acorai"])
        self.assertIn("Digital Health", hits[0].category_or_track)
        self.assertIn("MedTech", hits[1].category_or_track)

    def test_priority_ireland_accelerator_parsers_extract_company_links(self):
        fixtures = [
            (
                pipeline.Source("BioInnovate Ireland", "Accelerator", "https://www.bioinnovate.ie/", "Ireland", "High", "Annual", "Fellowship/company extraction", "Medtech programme.", "bioinnovate_ireland"),
                '<a href="/our-companies/proverum/">ProVerum</a><p>Medical device company for urology.</p>',
                "ProVerum",
            ),
            (
                pipeline.Source("ARC Hub for HealthTech", "Accelerator", "https://www.universityofgalway.ie/arc-healthtech/", "Ireland", "High", "Quarterly", "Commercialisation extraction", "Healthtech commercialisation.", "arc_hub_healthtech"),
                '<a href="/arc-healthtech/projects/feeltect/">FeelTect</a><p>Connected health compression monitoring device.</p>',
                "FeelTect",
            ),
            (
                pipeline.Source("Health Innovation Hub Ireland", "Accelerator", "https://www.hih.ie/", "Ireland", "High", "Quarterly", "Innovation extraction", "Clinical validation hub.", "health_innovation_hub_ireland"),
                '<a href="/case-studies/patientmpower/">patientMpower</a><p>Digital health respiratory monitoring platform.</p>',
                "PatientMpower",
            ),
            (
                pipeline.Source("Dogpatch Labs / NDRC", "Accelerator", "https://www.ndrc.ie/", "Ireland", "Medium", "Quarterly", "Portfolio extraction", "National accelerator.", "dogpatch_ndrc"),
                '<a href="/portfolio/silvercloud-health/">SilverCloud Health</a><p>Mental health digital therapeutics platform.</p>',
                "SilverCloud Health",
            ),
        ]

        for source, html, expected_company in fixtures:
            with self.subTest(source=source.name):
                hits = pipeline.parse_priority_accelerator_page(source, html, source.url)

                self.assertEqual([hit.company for hit in hits], [expected_company])
                self.assertEqual(hits[0].source_type, "Accelerator")
                self.assertIn(source.adapter, hits[0].matched_terms)

    def test_priority_ireland_accelerator_runners_emit_triggers(self):
        source = pipeline.Source("BioInnovate Ireland", "Accelerator", "https://www.bioinnovate.ie/", "Ireland", "High", "Annual", "Fellowship/company extraction", "Medtech programme.", "bioinnovate_ireland")
        html = '<script src="https://stories.universityofgalway.ie/bioinnovate/start-ups/embed.js"></script><script>fetch("https://data.shorthand.com/erKSumnd3Q/collections/Xta6wmZajc/items.json")</script>'
        payload = {
            "title": "BioInnovate Alumni Companies",
            "items": [
                {
                    "title": "Luma Vision",
                    "description": "Cardiac imaging medical device company.",
                    "url": "https://lumavision.com/",
                }
            ],
        }

        with patch.object(pipeline, "ACCELERATOR_SOURCE_PAGES", {"BioInnovate Ireland": ["https://www.bioinnovate.ie/bioinnovate/alumni/"]}), patch.object(pipeline, "fetch_raw_text", return_value=(html, None)), patch.object(pipeline, "fetch_json_url", return_value=(payload, None)):
            discovery_hits, trigger_events, result = pipeline.run_bioinnovate_ireland(source)

        self.assertEqual([hit.company for hit in discovery_hits], ["Luma Vision"])
        self.assertEqual(discovery_hits[0].website, "https://lumavision.com/")
        self.assertIn("shorthand alumni collection", discovery_hits[0].matched_terms)
        self.assertEqual([event.trigger_type for event in trigger_events], ["Accelerator/cohort"])
        self.assertIn("alumni collections scanned", result)

    def test_bioinnovate_extracts_all_plausible_alumni_without_health_keyword_filter(self):
        source = pipeline.Source("BioInnovate Ireland", "Accelerator", "https://www.bioinnovate.ie/", "Ireland", "High", "Annual", "Fellowship/company extraction", "Medtech programme.", "bioinnovate_ireland")
        html = """
        <a href="/our-companies/luma-vision/">Luma Vision</a><p>Alumni company.</p>
        <a href="/our-companies/galenband/">Galenband</a>
        <a href="/our-companies/proverum/">ProVerum</a><p>Venture profile.</p>
        """

        hits = pipeline.parse_priority_accelerator_page(source, html, source.url)

        self.assertEqual([hit.company for hit in hits], ["Luma Vision", "Galenband", "ProVerum"])

    def test_bioinnovate_skips_alumni_navigation_links(self):
        source = pipeline.Source("BioInnovate Ireland", "Accelerator", "https://www.bioinnovate.ie/bioinnovate/alumni/", "Ireland", "High", "Annual", "Fellowship/company extraction", "Medtech programme.", "bioinnovate_ireland")
        html = """
        <a href="/bioinnovate/alumni/">Alumni</a>
        <a href="/bioinnovate/alumni/directory/">BioInnovate Alumni</a>
        <a href="/bioinnovate/alumni/directory/">Alumni Directory</a>
        """

        hits = pipeline.parse_priority_accelerator_page(source, html, source.url)

        self.assertEqual(hits, [])

    def test_ndrc_filters_portfolio_links_to_healthcare_keyword_matches(self):
        source = pipeline.Source("Dogpatch Labs / NDRC", "Accelerator", "https://www.ndrc.ie/", "Ireland", "Medium", "Quarterly", "Portfolio extraction", "National accelerator.", "dogpatch_ndrc")
        html = """
        <a href="/portfolio/silvercloud-health/">SilverCloud Health</a>
        <p>Mental health digital therapeutics platform for patient care.</p>
        <a href="/portfolio/payrollflow/">PayrollFlow</a>
        <p>Payroll automation for small businesses.</p>
        """

        hits = pipeline.parse_priority_accelerator_page(source, html, source.url)

        self.assertEqual([hit.company for hit in hits], ["SilverCloud Health"])
        self.assertIn("healthcare keywords:", hits[0].matched_terms)
        self.assertIn("mental health", hits[0].matched_terms)

    def test_ndrc_extracts_healthcare_matches_from_current_cohort_domain_links(self):
        source = pipeline.Source("Dogpatch Labs / NDRC", "Accelerator", "https://www.ndrc.ie/accelerator-cohort-2024-h1", "Ireland", "Medium", "Quarterly", "Portfolio extraction", "National accelerator.", "dogpatch_ndrc")
        html = """
        <p>Blynksolve enables pharmaceutical drug substance manufacturers to build a digital knowledge twin.</p>
        <a href="https://www.blynksolve.com">blynksolve.com</a>
        <p>Vesta Insights serves the mortgage industry.</p>
        <a href="https://vesta-insights.example">vesta's website</a>
        """

        hits = pipeline.parse_priority_accelerator_page(source, html, source.url)

        self.assertEqual([hit.company for hit in hits], ["Blynksolve"])
        self.assertIn("pharma", hits[0].matched_terms)

    def test_nhs_innovation_accelerator_extracts_innovation_links(self):
        source = pipeline.Source("NHS Innovation Accelerator", "Accelerator", "https://nhsaccelerator.com/innovations/", "UK", "High", "Annual", "Innovation extraction", "NHS-backed health innovations.", "nhs_innovation_accelerator")
        html = """
        <a href="/innovations/ai-dimension/">AI Dimension AI-powered imaging workflow for clinicians.</a>
        <a href="/faqs/">Frequently asked questions</a>
        """

        hits = pipeline.parse_nhs_innovation_accelerator_page(source, html, source.url)

        self.assertEqual([hit.company for hit in hits], ["AI Dimension"])
        self.assertEqual(hits[0].source_type, "Accelerator")
        self.assertIn("nhs_innovation_accelerator", hits[0].matched_terms)

    def test_nlc_health_extracts_external_portfolio_websites(self):
        source = pipeline.Source("NLC Health", "Accelerator", "https://nlc.health/portfolio", "EU", "High", "Quarterly", "Venture extraction", "European health venture builder.", "nlc_health")
        html = """
        <article><h3>PEP Health</h3><p>Patient experience analytics for healthcare providers.</p><a href="https://www.pephealth.ai/">Visit website</a></article>
        <button>Accept All</button><a href="/ventures/">Ventures</a>
        """

        hits = pipeline.parse_nlc_health_page(source, html, source.url)

        self.assertEqual([hit.company for hit in hits], ["PEP Health"])
        self.assertEqual(hits[0].website, "https://www.pephealth.ai/")
        self.assertIn("nlc_health", hits[0].matched_terms)

    def test_yesdelft_medtech_extracts_health_sector_json_records(self):
        source = pipeline.Source("YES!Delft MedTech", "Accelerator", "https://yesdelft.com/wp-json/wp/v2/startups", "EU", "Medium", "Quarterly", "Startup extraction", "Medtech and health startups.", "yesdelft_medtech")
        payload = [
          {
            "title": {"rendered": "EchoPatch"},
            "link": "https://yesdelft.com/startups/echopatch/",
            "content": {"rendered": "<p>Wearable medical device for cardiac diagnostics.</p>"},
          }
        ]

        hits = pipeline.parse_yesdelft_medtech_payload(source, payload, source.url)

        self.assertEqual([hit.company for hit in hits], ["EchoPatch"])
        self.assertEqual(hits[0].discovery_url, "https://yesdelft.com/startups/echopatch/")

    def test_bioinnovation_institute_extracts_human_health_projects(self):
        source = pipeline.Source("BioInnovation Institute", "Accelerator", "https://bii.dk/community/start-ups-projects/", "EU", "High", "Quarterly", "Project extraction", "BII human health startups.", "bioinnovation_institute")
        payload = {
          "projectItems": [
            {
              "title": "Equilibrium Diagnostics",
              "summary": "Non-invasive kidney diagnostics for chronic kidney disease.",
              "url": "https://www.linkedin.com/company/equilibrium-diagnostics/",
              "externalLink": "https://www.linkedin.com/company/equilibrium-diagnostics/",
              "focusAreas": [{"name": "Human health"}],
              "programs": [{"name": "Venture Lab"}],
              "subAreas": [{"name": "Diagnostics"}],
            },
            {
              "title": "Planet Biofuel",
              "summary": "Industrial enzymes for agriculture.",
              "url": "https://planet.example",
              "focusAreas": [{"name": "Planetary health"}],
              "programs": [{"name": "Venture Lab"}],
              "subAreas": [],
            },
          ]
        }

        hits = pipeline.parse_bioinnovation_institute_payload(source, payload, source.url)

        self.assertEqual([hit.company for hit in hits], ["Equilibrium Diagnostics"])
        self.assertEqual(hits[0].category_or_track, "Diagnostics")

    def test_yesdelft_medtech_runner_emits_triggers(self):
        source = pipeline.Source("YES!Delft MedTech", "Accelerator", "https://yesdelft.com/wp-json/wp/v2/startups", "EU", "Medium", "Quarterly", "Startup extraction", "Medtech and health startups.", "yesdelft_medtech")
        payload = [
          {
            "title": {"rendered": "ClinicFlow"},
            "link": "https://yesdelft.com/startups/clinicflow/",
            "content": {"rendered": "<p>Clinical workflow software for hospitals.</p>"},
          }
        ]

        with patch.object(pipeline, "ACCELERATOR_SOURCE_PAGES", {"YES!Delft MedTech": ["https://yesdelft.com/wp-json/wp/v2/startups?sectors=49"]}), patch.object(pipeline, "fetch_json_url", return_value=(payload, None)):
            discovery_hits, trigger_events, result = pipeline.run_yesdelft_medtech(source)

        self.assertEqual([hit.company for hit in discovery_hits], ["ClinicFlow"])
        self.assertEqual([event.trigger_type for event in trigger_events], ["Accelerator/cohort"])
        self.assertIn("YES!Delft pages/endpoints scanned", result)

    def test_nucleate_activator_extracts_health_records_from_reader_text(self):
        source = pipeline.Source("Nucleate Activator", "Accelerator", "https://nucleate.org/companies", "EU/US/global", "High", "Annual", "Company extraction", "Academic biotech accelerator.", "nucleate_activator")
        markdown = """
        Markdown Content:
        3DDiagnostix San diego

         Launched 2022

        Diagnostics 3D DiagnostiX develops diagnostic tools for early detection of Alzheimer's disease.
        Aephoris Boston

         Launched 2022

        Eco Tolerance-enhanced yeast for cellulosic diesel.
        Biosens8 Boston

         Launched 2022

        Medical Devices Biosensor platform targeting ovulation confirmation in fertility medicine.
        """

        hits = pipeline.parse_nucleate_activator_page(source, markdown, source.url)

        self.assertEqual([hit.company for hit in hits], ["3DDiagnostix", "Biosens8"])
        self.assertEqual(hits[0].cohort_year, "2022")
        self.assertIn("nucleate_activator", hits[0].matched_terms)

    def test_cdl_health_extracts_only_health_stream_cards(self):
        source = pipeline.Source("Creative Destruction Lab Health", "Accelerator", "https://creativedestructionlab.com/companies/", "EU/UK/Canada", "High", "Annual", "Company extraction", "CDL health streams.", "cdl_health")
        html = """
        <a href="/companies/correlia-biosystems/" class="js-companybio-link company-link--noscale">
          <p class="companybio-location">Correlia Biosystems</p>
          <p class="companybio-stream">Health</p>
        </a>
        <a href="/companies/energy-grid/" class="js-companybio-link company-link--noscale">
          <p class="companybio-location">Energy Grid</p>
          <p class="companybio-stream">Energy</p>
        </a>
        <a href="/companies/voxcell/" class="js-companybio-link company-link--noscale">
          <p class="companybio-location">VoxCell BioInnovation</p>
          <p class="companybio-stream">Biomedical Engineering, Cancer</p>
        </a>
        """

        hits = pipeline.parse_cdl_health_page(source, html, source.url)

        self.assertEqual([hit.company for hit in hits], ["Correlia Biosystems", "VoxCell BioInnovation"])
        self.assertEqual(hits[0].category_or_track, "Health")
        self.assertEqual(hits[1].category_or_track, "Biomedical Engineering, Cancer")

    def test_cdl_health_runner_audits_canadian_sites_and_detail_pages(self):
        source = next(source for source in pipeline.SOURCES if source.name == "Creative Destruction Lab Health")
        pages = {}

        def card(index, location, stream="Health"):
            return f"""
            <a href="/companies/company-{index}/" class="js-companybio-link company-link--noscale">
              <p class="companybio-location">Company {index:03d}</p>
              <p class="companybio-stream">{stream}</p>
            </a>
            """

        toronto_indexes = range(139)
        vancouver_indexes = range(109, 226)
        pages[pipeline.CDL_CANADA_DIRECTORY_URLS["CDL-Toronto"]] = "".join(
            card(index, "CDL-Toronto", "Biomedical Engineering, Cancer" if index == 0 else "Health")
            for index in toronto_indexes
        )
        pages[pipeline.CDL_CANADA_DIRECTORY_URLS["CDL-Vancouver"]] = "".join(
            card(index, "CDL-Vancouver")
            for index in vancouver_indexes
        )
        for index in range(226):
            pages[f"https://creativedestructionlab.com/companies/company-{index}/"] = f"""
            <meta property="og:description" content="Health company {index} description." />
            <div class="company-category">
              <p><strong>Site:</strong> CDL-{"Toronto" if index < 139 else "Vancouver"}</p>
              <p><strong>Cohort Year:</strong> 2024/25</p>
              <p><strong>Stream:</strong> Health</p>
            </div>
            <a class="btn company-website-link" href="https://company-{index}.example/">Website</a>
            <h2 class="c-primary">Company {index:03d}</h2>
            """

        hits, triggers, result = pipeline.run_cdl_health(
            source,
            fetcher=lambda url: (pages.get(url, ""), None if url in pages else "not found"),
            max_workers=4,
        )

        self.assertEqual(len(hits), 226)
        self.assertEqual(len(triggers), 226)
        self.assertIn("226/226 unique Canadian-site", result)
        self.assertIn("139/139 Toronto", result)
        self.assertIn("117/117 Vancouver", result)
        self.assertNotIn("INCOMPLETE", result)
        self.assertEqual(hits[0].website, "https://company-0.example/")
        self.assertEqual(hits[0].cohort_year, "2024/25")

    def test_centech_health_source_uses_dedicated_adapter(self):
        source = next(source for source in pipeline.SOURCES if source.name == "Centech – Medtech")

        self.assertEqual(source.adapter, "centech_health")

    def test_centech_health_parser_preserves_api_metadata(self):
        source = next(source for source in pipeline.SOURCES if source.name == "Centech – Medtech")
        payload = [
            {
                "id": 1,
                "link": "https://centech.co/en/startups/aion-healthtech/",
                "title": {"rendered": "Aion Healthtech"},
                "categories": [5, 1094],
                "acf": {
                    "startups_cohorte": "Hiver 2026",
                    "startups_radio": "radio1",
                    "startups_website": {"url": "https://aion.example/"},
                    "startups_link": {"url": "https://linkedin.com/company/aion"},
                    "startups_description_en": "AI-assisted mental-health data analysis.",
                },
            }
        ]

        hits = pipeline.parse_centech_health_payload(source, payload)

        self.assertEqual([hit.company for hit in hits], ["Aion Healthtech"])
        self.assertEqual(hits[0].website, "https://aion.example/")
        self.assertEqual(hits[0].cohort_year, "2026")
        self.assertIn("Digital Health", hits[0].category_or_track)
        self.assertIn("Propulsé", hits[0].category_or_track)
        self.assertIn("LinkedIn:", hits[0].matched_terms)

    def test_centech_health_runner_checks_records_and_unique_companies(self):
        source = next(source for source in pipeline.SOURCES if source.name == "Centech – Medtech")
        payload = []
        for index in range(95):
            company_index = index if index < 92 else index - 92
            payload.append(
                {
                    "id": index,
                    "link": f"https://centech.co/startups/company-{index}/",
                    "title": {"rendered": f"Company {company_index:02d}"},
                    "categories": [32],
                    "acf": {"startups_radio": "radio2"},
                }
            )

        hits, triggers, result = pipeline.run_centech_health(
            source,
            fetcher=lambda url: (payload, None),
        )

        self.assertEqual(len(hits), 95)
        self.assertEqual(len(triggers), 95)
        self.assertIn("95/95", result)
        self.assertIn("92/92 unique", result)
        self.assertNotIn("INCOMPLETE", result)

    def test_remaining_canada_tier_a_sources_use_dedicated_adapters(self):
        expected = {
            "CTS Santé": "cts_sante_portfolio",
            "District 3 – Bio and Health": "district3_health",
            "OBIO": "obio_cohorts",
            "MEDTEQ+": "medteq_portfolio",
            "Innovate Calgary / UCeed Health": "uceed_health",
            "University of Alberta Health Innovation Hub": "ualberta_health_hub",
            "Innovation UBC Human Health portfolio": "innovation_ubc_health",
        }
        actual = {source.name: source.adapter for source in pipeline.SOURCES if source.name in expected}
        self.assertEqual(actual, expected)

    def test_cts_sante_runner_traverses_complete_portfolio_sitemap(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "cts_sante_portfolio")
        urls = [f"https://ctssante.com/portfolio/company-{index}/" for index in range(21)]
        pages = {
            pipeline.CTS_SITEMAP_URL: "".join(f"<loc>{url}</loc>" for url in urls),
            **{
                url: (
                    f"<h1>CTS Company {index}</h1>"
                    f'<meta name="description" content="Medical technology company {index}.">'
                    f'<a href="https://cts-{index}.example/">Website</a>'
                )
                for index, url in enumerate(urls)
            },
        }
        hits, triggers, result = pipeline.run_cts_sante(
            source, fetcher=lambda url: (pages.get(url, ""), None if url in pages else "missing")
        )
        self.assertEqual(len(hits), 21)
        self.assertEqual(len(triggers), 21)
        self.assertNotIn("INCOMPLETE", result)
        self.assertEqual(hits[-1].website, "https://cts-20.example/")

    def test_district3_runner_filters_all_paginated_cards(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "district3_health")

        def card(index, stream):
            return (
                '<div class="startup_grid-card">'
                f'<a href="https://d3-{index}.example/">D3 Company {index}</a>'
                f'<div class="news_grid-card-text">Company description {index}</div>'
                f'<div class="tag" fs-cmsfilter-field="stream">{stream}</div></div>'
            )

        pages = {
            source.url: "".join(card(i, "Healthcare" if i < 10 else "High Tech") for i in range(20)),
            f"{source.url}?37e06a8e_page=2": "".join(
                card(i, "Biotech" if i < 22 else "Social Innovation") for i in range(20, 25)
            ),
            f"{source.url}?37e06a8e_page=3": "",
        }
        hits, _, result = pipeline.run_district3(
            source, fetcher=lambda url: (pages.get(url, ""), None)
        )
        self.assertEqual(len(hits), 12)
        self.assertEqual({hit.category_or_track for hit in hits}, {"Healthcare", "Biotech"})
        self.assertNotIn("INCOMPLETE", result)

    def test_obio_runner_preserves_program_records_and_repeat_company(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "obio_cohorts")
        pages = {
            url: "<html>" + " | ".join(companies) + "</html>"
            for _, _, url, companies in pipeline.OBIO_COHORTS
        }
        hits, _, result = pipeline.run_obio(source, fetcher=lambda url: (pages[url], None))
        self.assertEqual(len(hits), 25)
        self.assertEqual(len({hit.company.lower() for hit in hits}), 24)
        self.assertNotIn("INCOMPLETE", result)

    def test_medteq_runner_requires_all_investment_portfolio_cards(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "medteq_portfolio")
        boxes = "".join(
            f'<div class="box"><p>{company} develops medical technology.</p></div>'
            for company in pipeline.MEDTEQ_COMPANIES
        )
        html = f'<section><div class="mosaique-logos">{boxes}</div></section>'
        hits, _, result = pipeline.run_medteq(source, fetcher=lambda url: (html, None))
        self.assertEqual(len(hits), 17)
        self.assertEqual(hits[1].company, "Gray Oncology Solutions")
        self.assertNotIn("INCOMPLETE", result)

    def test_uceed_runner_collects_both_health_funds(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "uceed_health")
        pages = {}
        for cohort, url, expected in pipeline.UCEED_PAGES:
            pages[url] = "<h2>Health Fund Portfolios</h2>" + "".join(
                f'<h3>{cohort} Company {index}</h3><p>Health venture.</p>'
                f'<a href="https://uceed-{expected}-{index}.example/">Visit website</a>'
                for index in range(expected)
            ) + "<h2>Advisors</h2>"
        hits, _, result = pipeline.run_uceed(source, fetcher=lambda url: (pages[url], None))
        self.assertEqual(len(hits), 42)
        self.assertTrue(all(hit.website for hit in hits))
        self.assertNotIn("INCOMPLETE", result)

    def test_ualberta_runner_uses_complete_company_card_directory(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "ualberta_health_hub")
        html = "".join(
            '<div class="card"><div class="card-header">'
            f'<span>Alberta Company {index}</span></div>'
            f'<div class="card-body"><p>Health product {index}</p>'
            f'<a href="https://alberta-{index}.example/">Company</a></div></div>'
            for index in range(44)
        )
        hits, _, result = pipeline.run_ualberta(source, fetcher=lambda url: (html, None))
        self.assertEqual(len(hits), 44)
        self.assertEqual(hits[0].geography, "Alberta, Canada")
        self.assertNotIn("INCOMPLETE", result)

    def test_innovation_ubc_runner_traverses_seven_pages_and_exact_health_filter(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "innovation_ubc_health")
        pages = {}
        counts = [60, 60, 60, 60, 60, 60, 25]
        health_counts = [22, 22, 22, 22, 22, 22, 20]
        company_index = 0
        for page, (count, health_count) in enumerate(zip(counts, health_counts)):
            rows = []
            for index in range(count):
                impact = "Human Health" if index < health_count else "Planetary Health"
                rows.append(
                    f"<tr><td><h5>UBC Company {company_index}</h5></td>"
                    f"<td>Vancouver, BC</td><td>{impact}</td><td>Spin-off</td>"
                    f'<td><a href="https://ubc-{company_index}.example/"><i></i></a>'
                    f'<a href="https://linkedin.com/company/ubc-{company_index}"><i></i></a></td></tr>'
                )
                company_index += 1
            url = source.url if page == 0 else f"{source.url}?page={page}"
            pages[url] = "<table><tbody>" + "".join(rows) + "</tbody></table>"
        pages[f"{source.url}?page=7"] = "<table></table>"
        hits, _, result = pipeline.run_innovation_ubc(
            source, fetcher=lambda url: (pages.get(url, ""), None)
        )
        self.assertEqual(len(hits), 152)
        self.assertTrue(all(hit.website for hit in hits))
        self.assertNotIn("INCOMPLETE", result)

    def test_tier_a_canada_investors_use_dedicated_adapters(self):
        expected = {
            "Lumira Ventures portfolio": "lumira_portfolio",
            "Genesys Capital portfolio": "genesys_portfolio",
            "Amplitude Ventures portfolio": "amplitude_portfolio",
            "BDC current health and life-sciences portfolio": "bdc_health_portfolio",
            "FACIT oncology investment portfolio": "facit_portfolio",
        }
        actual = {source.name: source.adapter for source in pipeline.SOURCES if source.name in expected}
        self.assertEqual(actual, expected)

    def test_lumira_runner_checks_current_and_exited_denominators(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "lumira_portfolio")
        cards, modals = [], []
        for index in range(59):
            status = "exits" if index < 25 else "current"
            cards.append(
                f'<div class="col grid-item {status}"><a class="stretched-link" '
                f'data-bs-target="#portfolio-popup-{index}"></a></div>'
            )
            modals.append(
                f'<div id="portfolio-popup-{index}" class="modal portfolio-madal">'
                f'<div class="member-img-modal"><img alt="Lumira Company {index}"></div>'
                f'<div class="modal-member-desc">Medical technology. Industry: Medical Device | '
                f'Headquarters: Toronto, Ontario <a href="https://lumira-{index}.example/">Website</a></div></div>'
            )
        hits, _, result = pipeline.run_lumira(source, fetcher=lambda url: ("".join(cards + modals), None))
        self.assertEqual(len(hits), 59)
        self.assertIn("34/34 current", result)
        self.assertIn("25/25 exited", result)
        self.assertNotIn("INCOMPLETE", result)

    def test_genesys_runner_traverses_all_active_investment_details(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "genesys_portfolio")
        links = "".join(f'<a href="/investments/company-{index}"><img></a>' for index in range(12))

        def fetcher(url):
            if url == source.url:
                return links + '<a href="/investments/coming-soon"></a>', None
            index = url.rsplit("-", 1)[-1]
            return (
                f'<h1>Genesys Company {index}</h1><p>Healthcare company developing a clinical product.</p>'
                f'<a href="https://genesys-{index}.example/">Company website</a>',
                None,
            )

        hits, _, result = pipeline.run_genesys(source, fetcher=fetcher)
        self.assertEqual(len(hits), 12)
        self.assertTrue(all(hit.website for hit in hits))
        self.assertNotIn("INCOMPLETE", result)

    def test_amplitude_runner_decodes_complete_embedded_portfolio(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "amplitude_portfolio")
        records = []
        for index in range(24):
            status = "exited" if index < 4 else "active"
            records.append({
                "title": f"Amplitude Company {index}",
                "portfolio_status": {"value": status},
                "categories": ["oncology"],
                "content": "Canadian precision medicine company.",
                "website": f"https://amplitude-{index}.example/",
            })
        raw_html = f'<portfolio-page :portfolios="{html.escape(json.dumps(records), quote=True)}"></portfolio-page>'
        hits, _, result = pipeline.run_amplitude(source, fetcher=lambda url: (raw_html, None))
        self.assertEqual(len(hits), 24)
        self.assertIn("20/20 active", result)
        self.assertIn("4/4 exited", result)
        self.assertNotIn("INCOMPLETE", result)

    def test_bdc_runner_unions_six_health_filters_and_company_details(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "bdc_health_portfolio")
        counts = dict(zip(pipeline.BDC_HEALTH_SECTORS, [9, 3, 2, 3, 6, 8]))
        detail_pages = {}

        def fetcher(url):
            if "FilterByFund" in url:
                sector = re.search(r"industrySector=([^&]+)", url).group(1)
                cards = []
                offset = sum(value for key, value in counts.items() if pipeline.BDC_HEALTH_SECTORS.index(key) < pipeline.BDC_HEALTH_SECTORS.index(sector))
                for index in range(counts[sector]):
                    company_index = offset + index
                    detail_url = f"https://www.bdc.ca/en/bdc-capital/venture-capital/portfolio/company-{company_index}"
                    detail_pages[detail_url] = (
                        f'<meta name="description" content="Health company {company_index}.">'
                        f'Region Ontario Industry sector {sector} Investment year 2025 Fund Seed Venture Fund '
                        f'<a href="https://bdc-{company_index}.example/">Website</a>'
                    )
                    cards.append(
                        f'<a title="BDC Company {company_index}" '
                        f'href="/en/bdc-capital/venture-capital/portfolio/company-{company_index}"></a>'
                    )
                return "".join(cards), None
            return detail_pages[url], None

        hits, _, result = pipeline.run_bdc_health(source, fetcher=fetcher, max_workers=2)
        self.assertEqual(len(hits), 31)
        self.assertNotIn("INCOMPLETE", result)

    def test_facit_runner_excludes_institution_owned_pre_company_assets(self):
        source = next(source for source in pipeline.SOURCES if source.adapter == "facit_portfolio")
        articles = []
        for index in range(81):
            name = f"OICR (Asset {index})" if index < 25 else f"FACIT Company {index}"
            articles.append(
                f'<article id="portfolio-{index}"><a href="/portfolio/company-{index}"><h1>{name}</h1></a>'
                '<div class="term"><div class="label">Innovation Type:</div><div class="value">Medical Technologies</div></div>'
                '<div class="term"><div class="label">Funding Stage:</div><div class="value">Seed</div></div>'
                f'<div class="innovation">Oncology product company {index}.</div></article>'
            )
        hits, _, result = pipeline.run_facit(source, fetcher=lambda url: ("".join(articles), None))
        self.assertEqual(len(hits), 56)
        self.assertTrue(all(hit.company_description for hit in hits))
        self.assertIn("25 institution-owned", result)
        self.assertNotIn("INCOMPLETE", result)

    def test_masschallenge_healthcare_extracts_cohort_links_from_article_json(self):
        source = pipeline.Source("MassChallenge HealthTech", "Accelerator", "https://masschallenge.org/articles/healthcare-life-sciences-traction-cohort-2026/", "US/global", "Medium", "Annual", "Cohort extraction", "Healthcare cohort.", "masschallenge_healthcare")
        payload = {
            "title": {"rendered": "MassChallenge Announces the 2026 Healthcare &amp; Life Sciences Traction Cohort"},
            "link": "https://masschallenge.org/articles/healthcare-life-sciences-traction-cohort-2026/",
            "content": {"rendered": """
                <h2>MEET THE COHORT</h2>
                <p><strong>MEDTECH</strong></p>
                <p><a href="https://latde-dx.com/">Latde Diagnostics</a> | <a href="https://safebvm.com/">SafeBVM Corp.</a></p>
                <p><strong>DIGITAL HEALTH &amp; AI</strong></p>
                <p><a href="https://mindmuscle.health/">MindMuscle</a></p>
                <h2>BUILDING THE FUTURE OF HEALTHCARE</h2>
            """},
        }

        hits = pipeline.parse_masschallenge_healthcare_article(source, payload, source.url)

        self.assertEqual([hit.company for hit in hits], ["Latde Diagnostics", "SafeBVM", "MindMuscle"])
        self.assertEqual(hits[0].category_or_track, "Medtech")
        self.assertEqual(hits[-1].category_or_track, "Digital Health & AI")

    def test_rockstart_health_extracts_healthcare_vertical_json_records(self):
        source = pipeline.Source("Rockstart Health", "Accelerator", "https://rockstart.com/portfolio/vertical-healthcare/", "EU", "Medium", "Quarterly", "Healthcare portfolio extraction", "Rockstart healthcare vertical.", "rockstart_health")
        payload = [
            {
                "title": {"rendered": "Aisel Health"},
                "link": "https://rockstart.com/portfolio-company/aisel-health/",
                "date": "2025-04-13T11:16:32",
            }
        ]

        hits = pipeline.parse_rockstart_health_payload(source, payload, source.url)

        self.assertEqual([hit.company for hit in hits], ["Aisel Health"])
        self.assertEqual(hits[0].category_or_track, "Healthcare")
        self.assertEqual(hits[0].cohort_year, "2025")

    def test_rockstart_health_extracts_healthcare_vertical_html_cards(self):
        source = pipeline.Source("Rockstart Health", "Accelerator", "https://rockstart.com/portfolio/vertical-healthcare/", "EU", "Medium", "Quarterly", "Healthcare portfolio extraction", "Rockstart healthcare vertical.", "rockstart_health")
        html = """
        <h5 class="elementor-heading-title"><a href="https://www.aisel.co/" target="_blank">Aisel Health</a></h5>
        <p>AI-driven innovation in psychiatry, streamlining workflows with pre-assessments.</p>
        """

        hits = pipeline.parse_rockstart_health_page(source, html, source.url)

        self.assertEqual([hit.company for hit in hits], ["Aisel Health"])
        self.assertEqual(hits[0].website, "https://www.aisel.co/")

    def test_sbri_healthcare_extracts_directory_items(self):
        source = pipeline.Source("SBRI Healthcare", "Grant/funding", "https://sbrihealthcare.co.uk/impact-case-studies/company-directory", "UK", "High", "Quarterly", "Portfolio extraction", "NHS-facing innovation portfolio.", "sbri_healthcare")
        html = """
        <div class="directory-item">
          <h3>Sensixa</h3>
          <div class="tableWrap">
            <table>
              <tr><th>Project</th><td>Care for adaptive living with interactive sensing</td></tr>
              <tr><th>Description</th><td><p>Miniaturised sensor and app used to predict adverse events.</p></td></tr>
              <tr><th>Health Innovation Network Partner</th><td>Health Innovation Network South London</td></tr>
              <tr><th>Website</th><td><a href="http://sensixa.com/">http://sensixa.com/</a></td></tr>
            </table>
          </div>
        </div>
        """

        hits = pipeline.parse_sbri_healthcare_page(source, html, source.url)

        self.assertEqual([hit.company for hit in hits], ["Sensixa"])
        self.assertEqual(hits[0].website, "http://sensixa.com/")
        self.assertIn("sbri_healthcare", hits[0].matched_terms)

    def test_sifted_ranking_parses_public_health_entries(self):
        source = pipeline.Source("Sifted AI 100 2025", "Public ranking", "https://sifted.eu/rankings/ai-100-2025", "Europe", "Medium", "Annual", "Sifted ranking scan", "AI ranking with health and biotech entries.", "sifted_ranking")
        html = """
        <a href="https://cardia.example">CardiaAI</a>
        Location Cambridge UK Founded 2024 Stage Seed CardiaAI builds clinical AI diagnostic software for hospitals.
        <a href="https://defense.example">DroneForge</a>
        Location Munich Germany Founded 2023 Stage Seed DroneForge builds defence robotics.
        """

        hits, anonymized = pipeline.parse_sifted_ranking_page(source, html, source.url)

        self.assertFalse(anonymized)
        self.assertEqual([hit.company for hit in hits], ["CardiaAI"])
        self.assertIn("sifted_ranking", hits[0].matched_terms)

    def test_sifted_ranking_reports_anonymized_prioritization_only(self):
        source = pipeline.Source("Sifted Healthtech Investors 2025", "Investor prioritization", "https://sifted.eu/rankings/europes-most-active-investors-in-2025-healthtech", "Europe", "High", "Annual", "Sifted ranking scan", "Healthtech investor ranking.", "sifted_ranking")
        html = "<p>Anonymized Company</p><p>Anonymized City, Anonymized Country</p>"

        with patch.object(pipeline, "fetch_raw_text", return_value=(html, None)):
            discovery_hits, trigger_events, result = pipeline.run_sifted_ranking(source)

        self.assertEqual(discovery_hits, [])
        self.assertEqual(trigger_events, [])
        self.assertIn("prioritization only", result)

    def test_priority_vc_parsers_extract_portfolio_company_links(self):
        fixtures = [
            (
                pipeline.Source("Fountain Healthcare Partners portfolio", "VC portfolio", "https://www.fountainhealthcare.com/portfolio/", "Ireland/EU/US", "High", "Monthly", "Portfolio extraction", "Life sciences investor.", "fountain_healthcare"),
                '<a href="/portfolio/neurovalve/">NeuroValve</a><p>Medical device company for structural heart disease.</p>',
                "NeuroValve",
            ),
            (
                pipeline.Source("Seroba Life Sciences portfolio", "VC portfolio", "https://seroba-lifesciences.com/portfolio/", "EU/Ireland", "High", "Monthly", "Portfolio extraction", "Life sciences investor.", "seroba_life_sciences"),
                '<a href="/portfolio/atlanti-dx/">AtlantiDx</a><p>Diagnostics platform for clinical labs.</p>',
                "AtlantiDx",
            ),
            (
                pipeline.Source("Atlantic Bridge portfolio", "VC portfolio", "https://www.abven.com/portfolio/", "Ireland/EU/US", "High", "Monthly", "Portfolio extraction", "Deeptech investor.", "atlantic_bridge"),
                '<a href="/portfolio/clinic-ai/">ClinicAI</a><p>AI health workflow spinout.</p>',
                "ClinicAI",
            ),
        ]

        for source, html, expected_company in fixtures:
            with self.subTest(source=source.name):
                hits = pipeline.parse_vc_portfolio_page(source, html, source.url)

                self.assertEqual([hit.company for hit in hits], [expected_company])
                self.assertEqual(hits[0].source_type, "VC portfolio")
                self.assertIn(source.adapter, hits[0].matched_terms)

    def test_priority_vc_runners_emit_investor_triggers(self):
        source = pipeline.Source("Seroba Life Sciences portfolio", "VC portfolio", "https://seroba-lifesciences.com/portfolio/", "EU/Ireland", "High", "Monthly", "Portfolio extraction", "Life sciences investor.", "seroba_life_sciences")
        html = '<a href="/portfolio/medbridge/">MedBridge</a><p>Digital health platform for regulated care pathways.</p>'

        with patch.object(pipeline, "fetch_raw_text", return_value=(html, None)):
            discovery_hits, trigger_events, result = pipeline.run_seroba_life_sciences(source)

        self.assertEqual([hit.company for hit in discovery_hits], ["MedBridge"])
        self.assertEqual([event.trigger_type for event in trigger_events], ["Investor backing"])
        self.assertIn("1 VC portfolio page", result)

    def test_generic_accelerator_adapter_is_skipped(self):
        source = pipeline.Source("Illumina Accelerator", "Accelerator", "https://www.illumina.com/science/accelerator.html", "US/UK", "High", "Annual", "Portfolio extraction", "Genomics startups.", "accelerator_page")

        with patch.object(pipeline, "fetch_raw_text") as fetch_raw_text:
            discovery_hits, trigger_events, run_log = pipeline.run_discovery([source])

        fetch_raw_text.assert_not_called()
        self.assertEqual(discovery_hits, [])
        self.assertEqual(trigger_events, [])
        self.assertEqual(run_log[0][3], "Skipped")
        self.assertIn("No source-specific accelerator adapter", run_log[0][4])
        self.assertEqual(pipeline.adapter_inventory_label(source), "Manual/not implemented")

    def test_source_page_adapter_extracts_candidates_and_triggers(self):
        source = pipeline.Source("Fixture Accelerator", "Accelerator", "https://example.com/cohort", "US/global", "High", "Annual", "Cohort extraction", "AI medical device startups.", "accelerator_page")
        html = """
        <html><body>
          <a href="/companies/nova-scan">NovaScan Health</a>
          <a href="/about">About</a>
          <a href="/companies/pulse-dx">PulseDx</a>
        </body></html>
        """

        discovery_hits, trigger_events = pipeline.build_source_page_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["NovaScan Health", "PulseDx"])
        self.assertEqual([event.trigger_type for event in trigger_events], ["Accelerator/cohort", "Accelerator/cohort"])
        self.assertEqual(discovery_hits[0].discovery_url, "https://example.com/companies/nova-scan")

    def test_source_page_adapter_uses_known_registry_metadata(self):
        source = pipeline.Source("Fixture VC", "VC portfolio", "https://example.com/portfolio", "US", "High", "Monthly", "Portfolio extraction", "Medtech portfolio.", "vc_portfolio_page")
        html = "<html><body><p>Portfolio includes Aidoc and other clinical AI companies.</p></body></html>"

        discovery_hits, trigger_events = pipeline.build_source_page_evidence(source, html)

        self.assertEqual(discovery_hits[0].company, "Aidoc")
        self.assertEqual(discovery_hits[0].website, "https://www.aidoc.com/")
        self.assertEqual(trigger_events[0].trigger_type, "Investor backing")

    def test_regulatory_adapter_does_not_extract_navigation_links(self):
        source = pipeline.Source("Fixture FDA", "Regulatory database", "https://example.com/fda", "US", "High", "Monthly", "Regulatory extraction", "FDA device database.", "regulatory_page")
        html = """
        <html><body>
          <a href="#search_form">Skip to Search</a>
          <a href="/medical-devices">Medical Devices</a>
          <p>Aidoc appears in this regulatory listing.</p>
        </body></html>
        """

        discovery_hits, trigger_events = pipeline.build_source_page_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["Aidoc"])
        self.assertEqual(trigger_events[0].trigger_type, "Regulatory listing")

    def test_source_page_adapter_rejects_directory_action_links(self):
        source = pipeline.Source("Fixture Conference", "Conference", "https://example.com", "Global", "High", "Annual", "Exhibitor extraction", "Medtech conference.", "conference_page")
        html = """
        <html><body>
          <a href="/company-list">Company List</a>
          <a href="/become-an-exhibitor">Become an Exhibitor</a>
          <a href="/exhibitors/nova-scan">NovaScan Health</a>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_source_page_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["NovaScan Health"])

    def test_source_page_adapter_rejects_team_links(self):
        source = pipeline.Source("Fixture VC", "VC portfolio", "https://example.com", "US", "High", "Monthly", "Portfolio extraction", "Medtech portfolio.", "vc_portfolio_page")
        html = """
        <html><body>
          <a href="/team/jane-founder">Core Jane Founder</a>
          <a href="/portfolio/nova-scan">NovaScan Health</a>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_source_page_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["NovaScan Health"])

    def test_source_page_adapter_rejects_person_names_on_portfolio_paths(self):
        source = pipeline.Source("Fixture VC", "VC portfolio", "https://example.com", "US", "High", "Monthly", "Portfolio extraction", "Medtech portfolio.", "vc_portfolio_page")
        html = """
        <html><body>
          <a href="/portfolio/andrew-kress">Andrew Kress</a>
          <a href="/portfolio/brightheart">BrightHeart</a>
        </body></html>
        """

        discovery_hits, _ = pipeline.build_source_page_evidence(source, html)

        self.assertEqual([hit.company for hit in discovery_hits], ["BrightHeart"])

    def test_yc_healthcare_adapter_paginates_and_sorts_by_launch_date(self):
        source = pipeline.Source("Y Combinator Healthcare", "Accelerator", "https://www.ycombinator.com/companies", "US/global", "Medium", "Quarterly", "YC Algolia company directory query", "Healthcare startups.", "yc_healthcare")
        pages = [
            {
                "nbHits": 3,
                "nbPages": 2,
                "hits": [
                    {"name": "Older Health", "slug": "older-health", "website": "https://older.example", "one_liner": "Healthcare workflow", "all_locations": "US", "batch": "Winter 2024", "tags": ["Healthcare"], "launched_at": 100},
                    {"name": "Newest Health", "slug": "newest-health", "website": "https://newest.example", "one_liner": "AI clinic ops", "all_locations": "UK", "batch": "Summer 2026", "tags": ["Healthcare", "AI"], "launched_at": 300},
                ],
            },
            {
                "nbHits": 3,
                "nbPages": 2,
                "hits": [
                    {"name": "Middle Health", "slug": "middle-health", "website": "https://middle.example", "one_liner": "Digital health", "all_locations": "EU", "batch": "Summer 2025", "tags": ["Healthcare"], "launched_at": 200},
                ],
            },
        ]

        with patch.object(pipeline, "fetch_json", side_effect=[(pages[0], None), (pages[1], None)]):
            discovery_hits, trigger_events, result = pipeline.run_yc_healthcare(source)

        self.assertEqual([hit.company for hit in discovery_hits], ["Newest Health", "Middle Health", "Older Health"])
        self.assertEqual(discovery_hits[0].discovery_url, "https://www.ycombinator.com/companies/newest-health")
        self.assertEqual([hit.cohort_year for hit in discovery_hits], ["2026", "2025", "2024"])
        self.assertEqual(discovery_hits[0].cohort_label, "Y Combinator Summer 2026")
        self.assertEqual(len(trigger_events), 3)
        self.assertIn("3 matches", result)

    def test_yc_batch_year_handles_short_batch_codes(self):
        self.assertEqual(pipeline.infer_yc_batch_year("W24"), "2024")
        self.assertEqual(pipeline.infer_yc_batch_year("S25"), "2025")


if __name__ == "__main__":
    unittest.main()
