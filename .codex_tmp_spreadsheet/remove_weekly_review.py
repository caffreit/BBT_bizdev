from pathlib import Path

from openpyxl import load_workbook


root = Path(__file__).resolve().parents[1]
input_path = root / "BlueBridge_TOFU_BizDev_V1.xlsx"
output_dir = root / "outputs" / "weekly_review_removed"
output_path = output_dir / "BlueBridge_TOFU_BizDev_V1_no_weekly_review.xlsx"

wb = load_workbook(input_path)
print("Before:", wb.sheetnames)

if "Weekly Review" not in wb.sheetnames:
    raise SystemExit("Weekly Review sheet was not found.")

del wb["Weekly Review"]
print("After:", wb.sheetnames)

output_dir.mkdir(parents=True, exist_ok=True)
wb.save(output_path)
print(output_path)
