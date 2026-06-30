from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(__file__).resolve().parents[1] / "BlueBridge_TOFU_BizDev_V1.xlsx"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "leads-data.js"

FIELD_MAP = {
    "Company": "company",
    "Company website": "website",
    "Company description": "description",
    "Product type": "productType",
    "Product area": "productArea",
    "Company type": "companyType",
    "Company stage": "companyStage",
    "Hiring signal": "hiringSignal",
    "Geography": "geography",
    "Funding stage": "fundingStage",
    "Accelerator program": "acceleratorProgram",
    "Cohort label": "cohortLabel",
    "Cohort year": "cohortYear",
    "Category / track": "categoryTrack",
    "Evidence year": "evidenceYear",
    "Evidence status": "evidenceStatus",
    "Source name": "sourceName",
    "Source type": "sourceType",
    "Source URL": "sourceUrl",
    "Discovery rationale": "discoveryRationale",
    "Matched terms": "matchedTerms",
    "Trigger type": "triggerType",
    "Persona": "persona",
    "BBT quadrant": "bbtQuadrant",
    "LinkedIn company URL": "linkedinCompanyUrl",
    "LinkedIn company status": "linkedinCompanyStatus",
    "Executive contact name": "executiveContactName",
    "Executive contact title": "executiveContactTitle",
    "Executive LinkedIn URL": "executiveLinkedinUrl",
    "Technical/R&D contact name": "technicalContactName",
    "Technical/R&D contact title": "technicalContactTitle",
    "Technical/R&D LinkedIn URL": "technicalLinkedinUrl",
    "Quality/QA contact name": "qualityContactName",
    "Quality/QA contact title": "qualityContactTitle",
    "Quality/QA LinkedIn URL": "qualityLinkedinUrl",
    "LinkedIn contact status": "linkedinContactStatus",
    "Date captured": "dateCaptured",
}

FILTER_FIELDS = [
    "evidenceYear",
    "geography",
    "productArea",
    "companyType",
    "companyStage",
    "fundingStage",
    "hiringSignal",
    "triggerType",
    "persona",
    "bbtQuadrant",
    "sourceType",
]


def clean_value(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def export_leads(workbook_path: Path, output_path: Path) -> dict:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Leads" not in wb.sheetnames:
        raise ValueError(f"{workbook_path} does not contain a Leads sheet")

    ws = wb["Leads"]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    header_index = {header: index for index, header in enumerate(headers)}

    missing = [header for header in FIELD_MAP if header not in header_index]
    if missing:
        raise ValueError(f"Missing expected Leads columns: {', '.join(missing)}")

    leads = []
    options = defaultdict(set)
    for row_number, row in enumerate(rows, start=2):
        lead = {"id": f"lead-{row_number - 1}"}
        for workbook_field, js_field in FIELD_MAP.items():
            lead[js_field] = clean_value(row[header_index[workbook_field]])
        lead["searchText"] = " ".join(
            lead[key]
            for key in (
                "company",
                "description",
                "discoveryRationale",
                "matchedTerms",
                "sourceName",
                "sourceType",
                "website",
                "productType",
            )
            if lead[key]
        ).lower()
        for field in FILTER_FIELDS:
            if lead[field]:
                options[field].add(lead[field])
        leads.append(lead)

    payload = {
        "meta": {
            "sourceWorkbook": workbook_path.name,
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "leadCount": len(leads),
        },
        "filterFields": FILTER_FIELDS,
        "options": {field: sorted(values) for field, values in options.items()},
        "leads": leads,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        "window.BBT_LEADS_DATA = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Export workbook Leads sheet for the static frontend.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    payload = export_leads(args.workbook, args.out)
    print(f"Exported {payload['meta']['leadCount']} leads to {args.out}")


if __name__ == "__main__":
    main()
