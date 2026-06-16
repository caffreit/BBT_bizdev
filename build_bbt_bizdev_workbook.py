from __future__ import annotations

import html
import json
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urlencode, urljoin, urlparse
from urllib.error import URLError
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUT = Path("BlueBridge_TOFU_BizDev_V1.xlsx")
TODAY = date.today().isoformat()
USER_AGENT = "Mozilla/5.0 (compatible; BBT-bizdev-pipeline/1.0)"

SEARCH_QUERIES = [
    "MedTech AI Funding",
    "SaMD Series A",
    "medical device",
    "FDA clearance AI",
    "digital health regulatory clearance",
]

SOURCE_TYPE_ADAPTERS = {
    "Conference": "conference_page",
    "VC portfolio": "vc_portfolio_page",
    "Grant/funding": "grant_funding_page",
    "Regulatory database": "regulatory_page",
    "Jobs": "jobs_page",
}

DISCOVERY_TERMS = {
    "Accelerator": ["accelerator", "cohort", "startup", "health", "medtech", "medical device", "digital health", "diagnostic", "ai"],
    "Conference": ["exhibitor", "sponsor", "startup", "medtech", "medical device", "digital health", "diagnostic", "imaging", "ai"],
    "VC portfolio": ["portfolio", "investment", "company", "health", "medtech", "medical device", "digital health", "diagnostic", "ai"],
    "Grant/funding": ["grant", "award", "funding", "project", "health", "medtech", "medical device", "digital health", "diagnostic", "ai"],
    "Regulatory database": ["fda", "clearance", "cleared", "510(k)", "de novo", "medical device", "software", "ai", "diagnostic"],
    "Jobs": ["regulatory", "quality", "design assurance", "v&v", "clinical", "medical device", "digital health", "ai"],
}

ADAPTER_STATUS_NAMES = {
    "yc_healthcare": "YC Healthcare directory adapter",
    "medtech_innovator": "MedTech Innovator adapter",
    "digitalhealth_london": "DigitalHealth.London adapter",
    "mayo_accelerate": "Mayo Clinic Platform Accelerate adapter",
    "eit_health_catapult": "EIT Health Catapult adapter",
    "conference_page": "Conference page adapter",
    "vc_portfolio_page": "VC portfolio adapter",
    "grant_funding_page": "Grant/funding adapter",
    "regulatory_page": "Regulatory adapter",
    "jobs_page": "Jobs page adapter",
    "greenhouse_jobs": "Greenhouse jobs adapter",
    "lever_jobs": "Lever jobs adapter",
    "ashby_jobs": "Ashby jobs adapter",
    "workable_jobs": "Workable jobs adapter",
    "smartrecruiters_jobs": "SmartRecruiters jobs adapter",
    "recruitee_jobs": "Recruitee jobs adapter",
    "generic_page_scan": "Generic page scan",
}

SOURCE_TRIGGER_TYPES = {
    "Accelerator": "Accelerator/cohort",
    "Conference": "Conference presence",
    "VC portfolio": "Investor backing",
    "Grant/funding": "Grant/public funding",
    "Regulatory database": "Regulatory listing",
    "Jobs": "Hiring signal",
}

YC_ALGOLIA_APP_ID = "45BWZJ1SGC"
YC_ALGOLIA_API_KEY = "NzllNTY5MzJiZGM2OTY2ZTQwMDEzOTNhYWZiZGRjODlhYzVkNjBmOGRjNzJiMWM4ZTU0ZDlhYTZjOTJiMjlhMWFuYWx5dGljc1RhZ3M9eWNkYyZyZXN0cmljdEluZGljZXM9WUNDb21wYW55X3Byb2R1Y3Rpb24lMkNZQ0NvbXBhbnlfQnlfTGF1bmNoX0RhdGVfcHJvZHVjdGlvbiZ0YWdGaWx0ZXJzPSU1QiUyMnljZGNfcHVibGljJTIyJTVE"
YC_HEALTHCARE_QUERY = "Healthcare"
MEDTECH_INNOVATOR_PORY_CONFIG_ID = "66eb41bc87c0d05ea2b410b8"
MEDTECH_INNOVATOR_PORY_APP_URL = "https://medtechinnovator-portfolio.pory.app"
MEDTECH_INNOVATOR_PORY_RECORDS_URL = f"https://app.pory.dev/data/{MEDTECH_INNOVATOR_PORY_CONFIG_ID}/records"
MAYO_READER_PREFIX = "https://r.jina.ai/http://"
JOB_BOARD_ADAPTERS = {
    "greenhouse_jobs": "greenhouse",
    "lever_jobs": "lever",
    "ashby_jobs": "ashby",
    "workable_jobs": "workable",
    "smartrecruiters_jobs": "smartrecruiters",
    "recruitee_jobs": "recruitee",
}
MAX_MATCHED_JOBS_PER_COMPANY = 5

ACCELERATOR_SOURCE_PAGES = {
    "MedTech Innovator": ["https://medtechinnovator.org/2026cohort/", "https://medtechinnovator.org/portfolio/"],
    "DigitalHealth.London Accelerator": ["https://digitalhealth.london/innovation-directory/companies"],
    "Mayo Clinic Platform Accelerate": ["https://www.mayoclinicplatform.org/focus-areas/digital-health/accelerate/accelerate-cohort-landing-page/"],
    "EIT Health Catapult": ["https://eithealth.eu/programmes/catapult/"],
    "TMC Innovation": ["https://www.tmc.edu/innovation/accelerator-healthtech/"],
    "Techstars Healthcare": ["https://www.techstars.com/portfolio"],
    "StartX Med": ["https://startx.com/companies"],
    "MassChallenge HealthTech": ["https://masschallenge.org/startups/"],
    "Plug and Play Health": ["https://www.plugandplaytechcenter.com/health/"],
}


@dataclass(frozen=True)
class Source:
    name: str
    source_type: str
    url: str
    geography: str
    priority: str
    update_cadence: str
    extraction_method: str
    notes: str
    adapter: str | None = None


@dataclass
class DiscoveryHit:
    company: str
    source_name: str
    source_type: str
    discovery_url: str
    discovery_rationale: str
    product_type: str = ""
    geography: str = ""
    website: str = ""
    matched_terms: str = ""
    captured_at: str = TODAY
    accelerator_program: str = ""
    cohort_label: str = ""
    cohort_year: str = ""
    category_or_track: str = ""
    company_description: str = ""


@dataclass
class TriggerEvent:
    company: str
    trigger_type: str
    trigger_event: str
    trigger_source: str
    evidence_url: str
    trigger_role: str = "Secondary"
    captured_at: str = TODAY


@dataclass(frozen=True)
class SearchResult:
    query: str
    title: str
    link: str
    summary: str = ""
    publisher: str = ""
    published_at: str = ""


@dataclass(frozen=True)
class JobPosting:
    title: str
    url: str
    description: str = ""
    location: str = ""
    department: str = ""


@dataclass
class CompanyRecord:
    company: str
    website: str = ""
    geography: str = ""
    product_type: str = ""
    discovery_hits: list[DiscoveryHit] = field(default_factory=list)
    triggers: list[TriggerEvent] = field(default_factory=list)


SOURCES: list[Source] = [
    Source("TIME HealthTech 2025", "Public ranking", "https://time.com/7318020/worlds-top-healthtech-companies-2025/", "Global", "High", "Annual", "Page scan", "Public ranking used as a discovery source for AI, diagnostics, wearables, and digital health companies.", "time_healthtech"),
    Source("Google News / web funding search", "News/search", "https://news.google.com/search?q=medtech%20AI%20medical%20device%20funding", "US/EU/global", "High", "Weekly", "Google News RSS query", "General public search for recent medtech, AI health, SaMD, device, diagnostics, and digital health funding announcements.", "google_news_search"),
    Source("FDA 510(k) database", "Regulatory database", "https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpmn/pmn.cfm", "US", "High", "Monthly", "Database query/export", "Recent clearances and predicate/category intelligence for medical device and SaMD companies.", "regulatory_page"),
    Source("FDA De Novo database", "Regulatory database", "https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfpmn/denovo.cfm", "US", "High", "Monthly", "Database query/export", "Recent De Novo decisions and novel device/software categories.", "regulatory_page"),
    Source("FDA AI/ML-enabled medical devices list", "Regulatory database", "https://www.fda.gov/medical-devices/software-medical-device-samd/artificial-intelligence-and-machine-learning-aiml-enabled-medical-devices", "US", "High", "Monthly", "Page/table extraction", "FDA-maintained list of AI/ML-enabled device authorizations.", "regulatory_page"),
    Source("MedTech Innovator", "Accelerator", "https://medtechinnovator.org/2026cohort/", "US/EU/global", "High", "Quarterly", "Cohort/company page extraction", "Medical device, diagnostic, and digital health startup source.", "medtech_innovator"),
    Source("Mayo Clinic Platform Accelerate", "Accelerator", "https://www.mayoclinicplatform.org/accelerate/", "US/global", "High", "Quarterly", "Cohort/company page extraction", "AI/digital health companies needing validation.", "mayo_accelerate"),
    Source("TMC Innovation", "Accelerator", "https://www.tmc.edu/innovation/accelerator-healthtech/", "US/global", "Medium", "Semiannual", "Cohort/company page extraction", "Texas Medical Center healthtech/medtech startups.", "accelerator_page"),
    Source("EIT Health Catapult", "Accelerator", "https://eithealth.eu/programmes/catapult/", "EU", "High", "Annual", "Finalist/cohort page extraction", "European health startups.", "eit_health_catapult"),
    Source("Y Combinator Healthcare", "Accelerator", "https://www.ycombinator.com/companies", "US/global", "Medium", "Quarterly", "YC Algolia company directory query", "Broad startup source; useful for AI health, care delivery, diagnostics, and workflow software.", "yc_healthcare"),
    Source("Techstars Healthcare", "Accelerator", "https://www.techstars.com/accelerators", "US/global", "Medium", "Quarterly", "Program/cohort extraction", "Healthcare and AI-health accelerator cohorts.", "accelerator_page"),
    Source("StartX Med", "Accelerator", "https://startx.com/", "US", "Medium", "Quarterly", "Portfolio/company extraction", "Stanford-associated health and medtech startups.", "accelerator_page"),
    Source("MassChallenge HealthTech", "Accelerator", "https://masschallenge.org/programs-healthtech/", "US/global", "Medium", "Annual", "Cohort extraction", "Healthtech startups with provider/payer partnerships.", "accelerator_page"),
    Source("Plug and Play Health", "Accelerator", "https://www.plugandplaytechcenter.com/health/", "US/global", "Medium", "Quarterly", "Portfolio/cohort extraction", "Corporate innovation and health startup pipeline.", "accelerator_page"),
    Source("DigitalHealth.London Accelerator", "Accelerator", "https://digitalhealth.london/programmes/accelerator/", "UK", "High", "Annual", "Cohort/company page extraction", "NHS-facing digital health accelerator with strong UK provider access.", "digitalhealth_london"),
    Source("NHS Innovation Accelerator", "Accelerator", "https://nhsaccelerator.com/", "UK", "High", "Annual", "Fellow/innovation page extraction", "NHS-backed innovation accelerator for adopted healthcare innovations.", "accelerator_page"),
    Source("NHS Clinical Entrepreneur Programme", "Accelerator", "https://nhscep.com/", "UK", "High", "Annual", "Programme/company page extraction", "NHS clinical founder programme with healthtech and medtech ventures.", "accelerator_page"),
    Source("Health Innovation Hub Ireland", "Accelerator", "https://www.hih.ie/", "Ireland", "High", "Quarterly", "Innovation/company page extraction", "Irish healthcare innovation hub connected to hospitals and SMEs.", "accelerator_page"),
    Source("BioInnovate Ireland", "Accelerator", "https://www.bioinnovate.ie/", "Ireland", "High", "Annual", "Fellowship/company page extraction", "Irish needs-led medtech innovation programme.", "accelerator_page"),
    Source("NDRC Accelerator", "Accelerator", "https://www.ndrc.ie/", "Ireland", "Medium", "Quarterly", "Portfolio/company page extraction", "Irish national accelerator with occasional health and digital health companies.", "accelerator_page"),
    Source("Startupbootcamp Digital Health", "Accelerator", "https://www.startupbootcamp.org/accelerator/digital-health/", "EU/global", "High", "Annual", "Cohort/company page extraction", "Digital health-focused Startupbootcamp programme and alumni source.", "accelerator_page"),
    Source("NLC Health", "Accelerator", "https://nlc.health/", "EU", "High", "Quarterly", "Venture/company page extraction", "European healthtech venture builder focused on health innovations.", "accelerator_page"),
    Source("YES!Delft MedTech", "Accelerator", "https://www.yesdelft.com/", "EU", "Medium", "Quarterly", "Startup/company page extraction", "Dutch deeptech accelerator with medtech and health startups.", "accelerator_page"),
    Source("Bayer G4A", "Accelerator", "https://www.g4a.health/", "EU/global", "High", "Annual", "Startup/company page extraction", "Bayer digital health partnership and accelerator programme.", "accelerator_page"),
    Source("Roche Startup Creasphere", "Accelerator", "https://www.startupcreasphere.com/", "EU/global", "High", "Annual", "Startup/company page extraction", "Roche-backed digital health and diagnostics startup programme.", "accelerator_page"),
    Source("Novartis Biome", "Accelerator", "https://www.biome.novartis.com/", "EU/global", "Medium", "Quarterly", "Partner/company page extraction", "Novartis digital health innovation network and startup source.", "accelerator_page"),
    Source("Philips HealthWorks", "Accelerator", "https://www.philips.com/a-w/about/innovation/healthworks.html", "EU/global", "Medium", "Quarterly", "Startup/company page extraction", "Philips health innovation programme and venture collaboration source.", "accelerator_page"),
    Source("Health Venture Lab", "Accelerator", "https://www.healthventurelab.eu/", "EU", "Medium", "Annual", "Cohort/company page extraction", "European health venture builder and accelerator source.", "accelerator_page"),
    Source("EIT Health Wild Card", "Accelerator", "https://eithealth.eu/programmes/wild-card/", "EU", "Medium", "Annual", "Team/company page extraction", "EIT Health challenge-driven venture creation programme.", "accelerator_page"),
    Source("Cedars-Sinai Accelerator", "Accelerator", "https://www.cedars-sinai.edu/research/innovation/accelerator.html", "US", "High", "Annual", "Portfolio/company page extraction", "Provider-backed healthcare accelerator with clinical deployment relevance.", "accelerator_page"),
    Source("Dreamit Healthtech", "Accelerator", "https://www.dreamit.com/healthtech", "US", "High", "Annual", "Portfolio/company page extraction", "US healthtech accelerator and investor-backed startup source.", "accelerator_page"),
    Source("MATTER", "Accelerator", "https://matter.health/", "US/global", "High", "Quarterly", "Startup/member page extraction", "Healthcare startup hub and accelerator source.", "accelerator_page"),
    Source("Health Wildcatters", "Accelerator", "https://www.healthwildcatters.com/", "US", "High", "Annual", "Portfolio/company page extraction", "Healthcare and life-sciences accelerator with alumni companies.", "accelerator_page"),
    Source("Johnson & Johnson JLABS", "Accelerator", "https://jlabs.jnjinnovation.com/", "US/EU/Asia/global", "High", "Quarterly", "Company/resident page extraction", "Global life-sciences incubator and startup network.", "accelerator_page"),
    Source("Illumina Accelerator", "Accelerator", "https://www.illumina.com/science/accelerator.html", "US/UK", "High", "Annual", "Portfolio/company page extraction", "Genomics and diagnostics accelerator backed by Illumina.", "accelerator_page"),
    Source("IndieBio", "Accelerator", "https://indiebio.co/", "US/global", "Medium", "Annual", "Company page extraction", "SOSV biotech and life-sciences accelerator.", "accelerator_page"),
    Source("HAX", "Accelerator", "https://hax.co/", "US/Asia/global", "Medium", "Quarterly", "Portfolio/company page extraction", "SOSV hardtech accelerator relevant for connected devices and medtech hardware.", "accelerator_page"),
    Source("Fogarty Innovation", "Accelerator", "https://www.fogartyinnovation.org/", "US", "High", "Annual", "Company/program page extraction", "Medtech innovation incubator and company-building source.", "accelerator_page"),
    Source("UCSF Rosenman Institute", "Accelerator", "https://rosenmaninstitute.org/", "US", "High", "Annual", "Portfolio/company page extraction", "UCSF health technology accelerator and innovation network.", "accelerator_page"),
    Source("Berkeley SkyDeck", "Accelerator", "https://skydeck.berkeley.edu/", "US/global", "Medium", "Quarterly", "Portfolio/company page extraction", "University accelerator with healthtech, AI, and medical device startups.", "accelerator_page"),
    Source("MedTech Actuator", "Accelerator", "https://medtechactuator.com/", "Australia/Asia", "High", "Annual", "Portfolio/company page extraction", "Asia-Pacific medtech accelerator and venture source.", "accelerator_page"),
    Source("MedTech Innovator Asia Pacific", "Accelerator", "https://medtechinnovator.org/asia-pacific/", "Asia-Pacific", "High", "Annual", "Cohort/company page extraction", "Asia-Pacific medtech accelerator cohort source.", "accelerator_page"),
    Source("HealthTech Hub Africa", "Accelerator", "https://www.healthtechhubafrica.org/", "Africa/global", "Medium", "Annual", "Cohort/company page extraction", "Pan-African healthtech accelerator and innovation source.", "accelerator_page"),
    Source("Villgro Africa", "Accelerator", "https://www.villgroafrica.org/", "Africa/global", "Medium", "Annual", "Portfolio/company page extraction", "Healthcare and life-sciences impact accelerator in Africa.", "accelerator_page"),
    Source("SGInnovate", "Accelerator", "https://www.sginnovate.com/", "Asia", "Medium", "Quarterly", "Startup/company page extraction", "Singapore deeptech innovation source with medtech and health AI companies.", "accelerator_page"),
    Source("The MedTech Conference", "Conference", "https://themedtechconference.com/", "US/global", "High", "Annual", "Exhibitor/company directory extraction", "US/global medtech exhibitors, sponsors, startups, and presenting companies.", "conference_page"),
    Source("MEDICA exhibitor index", "Conference", "https://www.medica-tradefair.com/", "Global", "High", "Annual", "Exhibitor directory extraction", "Large global medical technology exhibitor universe.", "conference_page"),
    Source("MedTech Forum", "Conference", "https://www.themedtechforum.eu/", "EU", "High", "Annual", "Sponsor/exhibitor extraction", "European medical technology industry source.", "conference_page"),
    Source("RSNA exhibitors", "Conference", "https://www.rsna.org/annual-meeting", "Global", "High", "Annual", "Exhibitor/category extraction", "Radiology, imaging AI, and diagnostic software companies.", "conference_page"),
    Source("HLTH exhibitors/sponsors", "Conference", "https://www.hlth.com/", "US/global", "Medium", "Annual", "Sponsor/exhibitor extraction", "Digital health and AI health ecosystem.", "conference_page"),
    Source("ViVE exhibitors/sponsors", "Conference", "https://www.viveevent.com/", "US/global", "Medium", "Annual", "Sponsor/exhibitor extraction", "Digital health, provider, payer, and healthcare AI companies.", "conference_page"),
    Source("HIMSS exhibitors", "Conference", "https://www.himss.org/global-conference", "Global", "Medium", "Annual", "Exhibitor/category extraction", "Healthcare IT, clinical workflow, and digital health vendors.", "conference_page"),
    Source("DMEA exhibitors", "Conference", "https://www.dmea.de/en/", "EU", "Medium", "Annual", "Exhibitor/category extraction", "European digital health and healthcare IT companies.", "conference_page"),
    Source("HealthQuest Capital portfolio", "VC portfolio", "https://www.healthquestcapital.com/portfolio/", "US", "High", "Monthly", "Portfolio/news page extraction", "Medtech and healthcare growth companies.", "vc_portfolio_page"),
    Source("Lightstone Ventures portfolio", "VC portfolio", "https://www.lightstonevc.com/portfolio/", "US/EU", "High", "Monthly", "Portfolio/news page extraction", "Life sciences, medtech, and diagnostics companies.", "vc_portfolio_page"),
    Source("7wireVentures portfolio", "VC portfolio", "https://www.7wireventures.com/portfolio/", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Digital health and healthcare services/software companies.", "vc_portfolio_page"),
    Source("Rock Health portfolio", "VC portfolio", "https://rockhealth.com/portfolio/", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Digital health companies and market signals.", "vc_portfolio_page"),
    Source("Define Ventures portfolio", "VC portfolio", "https://www.definevc.com/portfolio", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare software and digital health startups.", "vc_portfolio_page"),
    Source("Flare Capital portfolio", "VC portfolio", "https://www.flarecapital.com/portfolio", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare technology and services companies.", "vc_portfolio_page"),
    Source("Lux Capital portfolio", "VC portfolio", "https://www.luxcapital.com/companies", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Deeptech, AI, bio, and health companies.", "vc_portfolio_page"),
    Source("Khosla Ventures healthcare portfolio", "VC portfolio", "https://www.khoslaventures.com/portfolio/", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "AI and healthcare venture-backed companies.", "vc_portfolio_page"),
    Source("Sofinnova Partners portfolio", "VC portfolio", "https://www.sofinnovapartners.com/portfolio/", "EU/US", "Medium", "Monthly", "Portfolio/news page extraction", "European life sciences, medtech, and digital medicine companies.", "vc_portfolio_page"),
    Source("Seroba portfolio", "VC portfolio", "https://seroba-lifesciences.com/portfolio/", "EU/Ireland", "High", "Monthly", "Portfolio/news page extraction", "Ireland/EU-relevant life sciences and medtech portfolio.", "vc_portfolio_page"),
    Source("General Catalyst healthcare portfolio", "VC portfolio", "https://www.generalcatalyst.com/companies", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare, AI, and care-delivery companies backed by a large multi-stage investor.", "vc_portfolio_page"),
    Source("Andreessen Horowitz Bio + Health portfolio", "VC portfolio", "https://a16z.com/portfolio/", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Bio, healthcare AI, and digital health portfolio companies.", "vc_portfolio_page"),
    Source("Bessemer healthcare portfolio", "VC portfolio", "https://www.bvp.com/portfolio", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare cloud, AI, and digital health companies.", "vc_portfolio_page"),
    Source("GV portfolio", "VC portfolio", "https://www.gv.com/portfolio/", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Life sciences, healthcare, AI, and diagnostics companies.", "vc_portfolio_page"),
    Source("F-Prime Capital portfolio", "VC portfolio", "https://fprimecapital.com/portfolio/", "US/global", "High", "Monthly", "Portfolio/news page extraction", "Healthcare, therapeutics, medtech, and digital health companies.", "vc_portfolio_page"),
    Source("Polaris Partners healthcare portfolio", "VC portfolio", "https://www.polarispartners.com/portfolio/", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare, life sciences, and technology companies.", "vc_portfolio_page"),
    Source("NEA healthcare portfolio", "VC portfolio", "https://www.nea.com/portfolio", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Large healthcare, life sciences, and digital health venture portfolio.", "vc_portfolio_page"),
    Source("Venrock healthcare portfolio", "VC portfolio", "https://www.venrock.com/portfolio/", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare, biotech, and technology companies.", "vc_portfolio_page"),
    Source("ARCH Venture Partners portfolio", "VC portfolio", "https://www.archventure.com/portfolio", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Life sciences and diagnostics companies with translational health relevance.", "vc_portfolio_page"),
    Source("Third Rock Ventures portfolio", "VC portfolio", "https://www.thirdrockventures.com/portfolio", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Life sciences and platform companies with clinical development signals.", "vc_portfolio_page"),
    Source("OrbiMed portfolio", "VC portfolio", "https://www.orbimed.com/en/portfolio", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Global healthcare and life-sciences companies.", "vc_portfolio_page"),
    Source("RA Capital portfolio", "VC portfolio", "https://www.racap.com/portfolio/", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare and life-sciences companies, including diagnostics and enabling technologies.", "vc_portfolio_page"),
    Source("Deerfield portfolio", "VC portfolio", "https://deerfield.com/portfolio", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare companies across therapeutics, devices, diagnostics, and services.", "vc_portfolio_page"),
    Source("Canaan healthcare portfolio", "VC portfolio", "https://www.canaan.com/portfolio", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare, digital health, and life-sciences venture-backed companies.", "vc_portfolio_page"),
    Source("Menlo Ventures healthcare portfolio", "VC portfolio", "https://menlovc.com/portfolio/", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare AI, digital health, and enterprise workflow companies.", "vc_portfolio_page"),
    Source("Norwest healthcare portfolio", "VC portfolio", "https://www.nvp.com/portfolio/", "US/global", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare, software, and medical technology companies.", "vc_portfolio_page"),
    Source("Mayfield healthcare portfolio", "VC portfolio", "https://www.mayfield.com/portfolio/", "US", "Medium", "Monthly", "Portfolio/news page extraction", "AI, health, and enterprise software companies.", "vc_portfolio_page"),
    Source("Transformation Capital portfolio", "VC portfolio", "https://transformation.capital/portfolio/", "US", "High", "Monthly", "Portfolio/news page extraction", "Digital health and healthcare technology companies.", "vc_portfolio_page"),
    Source("LRVHealth portfolio", "VC portfolio", "https://www.lrvhealth.com/portfolio/", "US", "High", "Monthly", "Portfolio/news page extraction", "Provider-connected healthcare technology and digital health companies.", "vc_portfolio_page"),
    Source("Echo Health Ventures portfolio", "VC portfolio", "https://echohealthventures.com/portfolio/", "US", "High", "Monthly", "Portfolio/news page extraction", "Healthcare technology companies backed by strategic health investors.", "vc_portfolio_page"),
    Source("Health Velocity Capital portfolio", "VC portfolio", "https://www.healthvelocitycapital.com/portfolio/", "US", "High", "Monthly", "Portfolio/news page extraction", "Healthcare software, services, and digital health companies.", "vc_portfolio_page"),
    Source("Town Hall Ventures portfolio", "VC portfolio", "https://townhallventures.com/portfolio/", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare companies focused on underserved populations and care delivery.", "vc_portfolio_page"),
    Source("Frist Cressey Ventures portfolio", "VC portfolio", "https://fcventures.com/portfolio/", "US", "Medium", "Monthly", "Portfolio/news page extraction", "Healthcare services, software, and technology companies.", "vc_portfolio_page"),
    Source("Wavemaker Three-Sixty Health portfolio", "VC portfolio", "https://wavemaker360.com/portfolio/", "US", "High", "Monthly", "Portfolio/news page extraction", "Healthcare seed-stage companies across digital health, medtech, and diagnostics.", "vc_portfolio_page"),
    Source("Nina Capital portfolio", "VC portfolio", "https://www.nina.capital/portfolio", "EU/global", "High", "Monthly", "Portfolio/news page extraction", "European health technology and digital health portfolio companies.", "vc_portfolio_page"),
    Source("Heal Capital portfolio", "VC portfolio", "https://www.healcapital.com/portfolio/", "EU", "High", "Monthly", "Portfolio/news page extraction", "European digital health and healthtech companies.", "vc_portfolio_page"),
    Source("Crista Galli Ventures portfolio", "VC portfolio", "https://www.cristagalli.com/portfolio", "EU/UK", "High", "Monthly", "Portfolio/news page extraction", "European healthtech, deeptech, and digital health companies.", "vc_portfolio_page"),
    Source("MTIP portfolio", "VC portfolio", "https://www.mtip.ch/portfolio/", "EU", "High", "Monthly", "Portfolio/news page extraction", "European healthtech growth companies.", "vc_portfolio_page"),
    Source("Ysios Capital portfolio", "VC portfolio", "https://ysioscapital.com/portfolio/", "EU", "Medium", "Monthly", "Portfolio/news page extraction", "European life-sciences, diagnostics, and healthcare companies.", "vc_portfolio_page"),
    Source("Kurma Partners portfolio", "VC portfolio", "https://www.kurmapartners.com/portfolio/", "EU", "Medium", "Monthly", "Portfolio/news page extraction", "European healthcare, biotech, and diagnostics companies.", "vc_portfolio_page"),
    Source("BioGeneration Ventures portfolio", "VC portfolio", "https://www.biogenerationventures.com/portfolio/", "EU", "Medium", "Monthly", "Portfolio/news page extraction", "European life-sciences and medtech-adjacent companies.", "vc_portfolio_page"),
    Source("EQT Life Sciences portfolio", "VC portfolio", "https://eqtgroup.com/current-portfolio/fund/eqt-life-sciences/", "EU/US", "Medium", "Monthly", "Portfolio/news page extraction", "Life-sciences and healthcare companies from the EQT Life Sciences platform.", "vc_portfolio_page"),
    Source("Forbion portfolio", "VC portfolio", "https://forbion.com/en/portfolio/", "EU/US", "Medium", "Monthly", "Portfolio/news page extraction", "Life-sciences, diagnostics, and therapeutic platform companies.", "vc_portfolio_page"),
    Source("Gilde Healthcare portfolio", "VC portfolio", "https://gildehealthcare.com/portfolio/", "EU/US", "High", "Monthly", "Portfolio/news page extraction", "Healthcare, medtech, diagnostics, and digital health companies.", "vc_portfolio_page"),
    Source("Endeavour Vision portfolio", "VC portfolio", "https://www.endeavourvision.com/portfolio/", "EU/US", "High", "Monthly", "Portfolio/news page extraction", "Medtech and digital health growth companies.", "vc_portfolio_page"),
    Source("SHS Capital portfolio", "VC portfolio", "https://www.shs-capital.eu/en/portfolio/", "EU", "High", "Monthly", "Portfolio/news page extraction", "European healthcare, medtech, diagnostics, and digital health companies.", "vc_portfolio_page"),
    Source("Panakes Partners portfolio", "VC portfolio", "https://www.panakes.it/portfolio/", "EU", "High", "Monthly", "Portfolio/news page extraction", "European medtech, diagnostics, and digital health companies.", "vc_portfolio_page"),
    Source("Vesalius Biocapital portfolio", "VC portfolio", "https://vesaliusbiocapital.com/portfolio/", "EU", "Medium", "Monthly", "Portfolio/news page extraction", "European life-sciences and medtech companies.", "vc_portfolio_page"),
    Source("Asabys Partners portfolio", "VC portfolio", "https://asabys.com/portfolio/", "EU", "High", "Monthly", "Portfolio/news page extraction", "European health innovation, medtech, and digital health companies.", "vc_portfolio_page"),
    Source("AlbionVC healthcare portfolio", "VC portfolio", "https://albion.vc/portfolio/", "UK/EU", "Medium", "Monthly", "Portfolio/news page extraction", "UK and European digital health, software, and life-sciences companies.", "vc_portfolio_page"),
    Source("CORDIS", "Grant/funding", "https://cordis.europa.eu/projects", "EU", "High", "Monthly", "Search/export", "EU-funded health, AI, and medtech projects.", "grant_funding_page"),
    Source("NIH RePORTER", "Grant/funding", "https://reporter.nih.gov/", "US", "High", "Monthly", "API/search export", "SBIR/STTR and translational research awards with abstracts.", "grant_funding_page"),
    Source("SBIR.gov awards", "Grant/funding", "https://www.sbir.gov/sbirsearch/award/all", "US", "High", "Monthly", "Award search/export", "US startup grant funding for translational health, AI, device, and diagnostics projects.", "grant_funding_page"),
    Source("ARPA-H portfolio", "Grant/funding", "https://arpa-h.gov/explore-funding/portfolio", "US", "High", "Monthly", "Portfolio review", "High-value translational health programs and awardees.", "grant_funding_page"),
    Source("EIC Accelerator", "Grant/funding", "https://eic.ec.europa.eu/eic-funding-opportunities/eic-accelerator_en", "EU", "High", "Quarterly", "Selected companies/results review", "European deep-tech startups with grant/equity funding.", "grant_funding_page"),
    Source("Innovate UK", "Grant/funding", "https://www.ukri.org/councils/innovate-uk/", "UK", "Medium", "Monthly", "Award/news search", "UK health innovation and medtech funding.", "grant_funding_page"),
    Source("SBRI Healthcare", "Grant/funding", "https://sbrihealthcare.co.uk/", "UK", "High", "Quarterly", "Competition/award review", "NHS-facing innovation awards.", "grant_funding_page"),
    Source("Enterprise Ireland", "Grant/funding", "https://www.enterprise-ireland.com/", "Ireland", "High", "Quarterly", "Portfolio/news review", "Irish startup and medtech funding.", "grant_funding_page"),
    Source("LinkedIn Jobs", "Jobs", "https://www.linkedin.com/jobs/", "US/EU/global", "High", "Weekly", "Manual search", "Hiring gaps for regulatory, QA, design assurance, V&V.", "jobs_page"),
    Source("Wellfound jobs", "Jobs", "https://wellfound.com/jobs", "US/EU/global", "Medium", "Weekly", "Manual search", "Startup hiring signals for regulatory, QA, clinical, and product roles.", "jobs_page"),
    Source("Greenhouse job boards", "Jobs", "https://www.greenhouse.com/", "US/EU/global", "Medium", "Weekly", "Company careers page search", "Direct company hiring gaps exposed through Greenhouse-hosted job pages.", "greenhouse_jobs"),
    Source("Lever job boards", "Jobs", "https://www.lever.co/", "US/EU/global", "Medium", "Weekly", "Company careers page search", "Direct company hiring gaps exposed through Lever-hosted job pages.", "lever_jobs"),
    Source("Indeed jobs", "Jobs", "https://www.indeed.com/", "US/EU/global", "Medium", "Weekly", "Manual search", "Broad hiring signal source for regulatory, quality, clinical, product, and engineering roles.", "jobs_page"),
    Source("Glassdoor jobs", "Jobs", "https://www.glassdoor.com/Job/", "US/EU/global", "Medium", "Weekly", "Manual search", "Broad job-board source with company hiring and role-context signals.", "jobs_page"),
    Source("ZipRecruiter jobs", "Jobs", "https://www.ziprecruiter.com/jobs-search", "US", "Medium", "Weekly", "Manual search", "US hiring signal source for regulatory, QA, clinical, and product roles.", "jobs_page"),
    Source("Google Jobs search", "Jobs", "https://www.google.com/search?q=medtech+regulatory+quality+jobs", "US/EU/global", "Medium", "Weekly", "Manual search", "Search-based discovery for company career pages and recent role postings.", "jobs_page"),
    Source("Built In jobs", "Jobs", "https://builtin.com/jobs", "US", "Medium", "Weekly", "Manual search", "Technology startup hiring signals, including healthtech and AI companies.", "jobs_page"),
    Source("Welcome to the Jungle jobs", "Jobs", "https://www.welcometothejungle.com/en/jobs", "EU/US", "Medium", "Weekly", "Manual search", "Startup and scaleup hiring signals across Europe and the US.", "jobs_page"),
    Source("Remote OK jobs", "Jobs", "https://remoteok.com/", "Global", "Low", "Weekly", "Manual search", "Remote startup hiring signals for product, engineering, and operations roles.", "jobs_page"),
    Source("Workable job boards", "Jobs", "https://www.workable.com/job-board", "US/EU/global", "Medium", "Weekly", "Company careers page search", "Direct company hiring gaps exposed through Workable-hosted job pages.", "workable_jobs"),
    Source("Ashby job boards", "Jobs", "https://www.ashbyhq.com/", "US/EU/global", "Medium", "Weekly", "Company careers page search", "Direct company hiring gaps exposed through Ashby-hosted job pages.", "ashby_jobs"),
    Source("SmartRecruiters job boards", "Jobs", "https://www.smartrecruiters.com/", "US/EU/global", "Medium", "Weekly", "Company careers page search", "Direct company hiring gaps exposed through SmartRecruiters-hosted job pages.", "smartrecruiters_jobs"),
    Source("Recruitee job boards", "Jobs", "https://recruitee.com/", "EU/global", "Medium", "Weekly", "Company careers page search", "Direct company hiring gaps exposed through Recruitee-hosted job pages.", "recruitee_jobs"),
    Source("Workday careers pages", "Jobs", "https://www.workday.com/", "US/EU/global", "Low", "Weekly", "Company careers page search", "Enterprise-hosted career pages for larger medtech and healthcare companies.", "jobs_page"),
    Source("BioSpace jobs", "Jobs", "https://www.biospace.com/jobs/", "US/global", "High", "Weekly", "Manual search", "Life-sciences hiring signals, including diagnostics, clinical, regulatory, and quality roles.", "jobs_page"),
    Source("MedReps jobs", "Jobs", "https://www.medreps.com/medical-sales-jobs", "US", "Medium", "Weekly", "Manual search", "Medical device commercial hiring signals and market-entry clues.", "jobs_page"),
    Source("NHS Jobs", "Jobs", "https://www.jobs.nhs.uk/", "UK", "Medium", "Weekly", "Manual search", "UK provider-side digital, clinical, and innovation hiring signals.", "jobs_page"),
    Source("HealthJobsUK", "Jobs", "https://www.healthjobsuk.com/", "UK", "Medium", "Weekly", "Manual search", "UK healthcare hiring signals for digital, clinical, implementation, and operational roles.", "jobs_page"),
]


# This registry is not a lead list. It is the vocabulary adapters look for on approved sources.
# A company is included only if an approved source page actually contains its configured names.
COMPANY_REGISTRY = {
    "Qure.ai": {"aliases": ["Qure AI", "Qure.ai"], "website": "https://qure.ai/", "geography": "India/US/EU", "product_type": "AI medical imaging", "job_boards": [{"platform": "greenhouse", "account": "qureai"}]},
    "Canary Speech": {"aliases": ["Canary Speech"], "website": "https://www.canaryspeech.com/", "geography": "US", "product_type": "Voice biomarkers / diagnostics"},
    "Fedo": {"aliases": ["Fedo"], "website": "https://fedo.ai/", "geography": "India/global", "product_type": "AI vitals / risk scoring"},
    "Oura": {"aliases": ["Oura"], "website": "https://ouraring.com/", "geography": "Finland/US", "product_type": "Wearable health device", "job_boards": [{"platform": "greenhouse", "account": "oura"}]},
    "Cera": {"aliases": ["Cera"], "website": "https://www.cera.care/", "geography": "UK", "product_type": "AI home care / risk prediction"},
    "Sword Health": {"aliases": ["Sword Health", "SWORD Health"], "website": "https://swordhealth.com/", "geography": "Portugal/US", "product_type": "Digital MSK / AI care", "job_boards": [{"platform": "greenhouse", "account": "swordhealth"}]},
    "Eko Health": {"aliases": ["Eko Health", "Eko"], "website": "https://www.ekohealth.com/", "geography": "US", "product_type": "AI-enabled digital stethoscope", "job_boards": [{"platform": "greenhouse", "account": "ekohealth"}]},
    "Aidoc": {"aliases": ["Aidoc", "AIDoc"], "website": "https://www.aidoc.com/", "geography": "Israel/US", "product_type": "Clinical AI / radiology triage", "job_boards": [{"platform": "greenhouse", "account": "aidoc"}]},
    "Quibim": {"aliases": ["Quibim"], "website": "https://www.quibim.com/", "geography": "Spain/US", "product_type": "AI imaging biomarkers"},
    "Outcomes4Me": {"aliases": ["Outcomes4Me"], "website": "https://outcomes4me.com/", "geography": "US/EU", "product_type": "Oncology digital health"},
    "Alimetry": {"aliases": ["Alimetry"], "website": "https://www.alimetry.com/", "geography": "New Zealand/US/EU", "product_type": "Gastric mapping medical device"},
}


TRIGGER_SOURCES = [
    {
        "name": "Axios Aidoc funding",
        "url": "https://www.axios.com/2026/04/29/exclusive-clinical-ai-provider-aidoc-raises-150m-series-e",
        "companies": ["Aidoc"],
        "trigger_type": "Funding",
        "trigger_event": "Raised $150M Series E for clinical AI expansion.",
    },
    {
        "name": "TechCrunch Eko funding",
        "url": "https://techcrunch.com/2024/06/05/eko-health-scores-41m-to-detect-heart-disease-earlier-and-more-accurately/",
        "companies": ["Eko Health"],
        "trigger_type": "Funding",
        "trigger_event": "Raised $41M Series D to expand AI stethoscope business.",
    },
    {
        "name": "TechCrunch Quibim funding",
        "url": "https://techcrunch.com/2025/01/28/quibim-raises-50m-to-develop-ai-models-for-medical-imaging/",
        "companies": ["Quibim"],
        "trigger_type": "Funding",
        "trigger_event": "Raised $50M Series A to develop AI models for medical imaging.",
    },
    {
        "name": "TechCrunch Sword funding",
        "url": "https://techcrunch.com/2024/06/04/sword-health-raises-130m-and-its-valuation-soars-to-3b/",
        "companies": ["Sword Health"],
        "trigger_type": "Funding",
        "trigger_event": "Raised $130M at a reported $3B valuation for AI-powered digital MSK care.",
    },
    {
        "name": "TIME HealthTech 2025",
        "url": "https://time.com/7318020/worlds-top-healthtech-companies-2025/",
        "companies": ["Qure.ai", "Canary Speech", "Fedo", "Oura", "Cera"],
        "trigger_type": "Market recognition",
        "trigger_event": "Recognized in TIME's 2025 HealthTech ranking.",
    },
]


SOURCE_PLAYBOOKS = [
    ["News/search", "Weekly searches: medtech AI funding; SaMD raises; digital health Series A; medical device seed funding; diagnostic AI funding.", "Result must name a company, date, funding/launch/regulatory event, and source URL.", "Company, trigger article, amount/event, date, geography, product clues", "Recent public trigger; discovery and trigger may be same URL.", "Advisory"],
    ["Public ranking", "Scan ranking page for configured medtech/AI-health company aliases.", "Company must appear on the approved ranking page.", "Company, source page, matched terms, rationale", "The source itself proves discovery; trigger may be separate.", "Advisory"],
    ["Accelerator", "Scan cohort/program pages for configured company aliases and category terms.", "Company must appear on the accelerator page.", "Company, cohort/source page, matched terms", "Curated early-stage medtech/digital health company.", "Advisory/design-dev"],
    ["Conference", "Extract exhibitors/sponsors/startups, then filter category text for AI, software, digital health, imaging, diagnostics, wearables, remote monitoring.", "Company must appear on event directory or sponsor/exhibitor page.", "Company, event, category, booth/profile URL, product summary", "Commercial presence in relevant market.", "Advisory/design-dev"],
    ["VC portfolio", "Extract portfolio pages and news pages; filter for medical device, diagnostics, digital health, AI healthcare, SaMD, clinical workflow.", "Company must appear on public portfolio/news page; paid databases excluded.", "Company, investor, portfolio URL, category, funding/news URL if present", "Funded or investor-backed company with likely budget/urgency.", "Advisory/design-dev"],
    ["Grant/funding", "Use API/export where available; search project abstract for SaMD, AI diagnostic, wearable, remote monitoring.", "Recipient must be company or startup-like organization.", "Company, award, abstract, amount, date, URL", "Funded translational program.", "Advisory"],
    ["Regulatory database", "Query FDA AI/ML list, 510(k), and De Novo databases for AI, software, imaging, diagnostics, monitoring, decision support.", "Company/product must have a regulatory record or listing.", "Company, product, decision/date, regulation number, source URL", "Regulatory pathway or market-entry signal.", "Advisory"],
    ["Jobs", "Search regulatory/QA/design assurance terms against public job pages.", "Role must be recent and company must be relevant to medtech/health AI.", "Company, role, job URL, gap hypothesis", "Capability gap or urgent workload.", "Embedded support/advisory"],
]


def fetch_raw_text(url: str) -> tuple[str, str | None]:
    try:
        req = Request(url, headers={"User-Agent": USER_AGENT})
        raw = urlopen(req, timeout=25).read()
    except (OSError, URLError) as exc:
        return "", str(exc)
    return raw.decode("utf-8", "ignore"), None


def fetch_text(url: str) -> tuple[str, str | None]:
    text, error = fetch_raw_text(url)
    if error:
        return "", error
    text = re.sub(r"<script\b.*?</script>", " ", text, flags=re.I | re.S)
    text = re.sub(r"<style\b.*?</style>", " ", text, flags=re.I | re.S)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text, None


def clean_text(value: str) -> str:
    value = re.sub(r"<[^>]+>", " ", value or "")
    value = html.unescape(value)
    return re.sub(r"\s+", " ", value).strip()


def excel_safe(value):
    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]", "", value)
    return value


def append_excel_row(ws, row):
    ws.append([excel_safe(value) for value in row])


class LinkTextParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href_stack: list[str | None] = []
        self._current_text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]):
        if tag.lower() != "a":
            return
        href = dict(attrs).get("href")
        self._href_stack.append(href)
        self._current_text.append("")

    def handle_data(self, data: str):
        if self._href_stack:
            self._current_text[-1] += data

    def handle_endtag(self, tag: str):
        if tag.lower() != "a" or not self._href_stack:
            return
        href = self._href_stack.pop()
        text = clean_text(self._current_text.pop())
        if href and text:
            self.links.append((text, href))


def extract_links(raw_html: str, base_url: str) -> list[tuple[str, str]]:
    parser = LinkTextParser()
    try:
        parser.feed(raw_html)
    except Exception:
        return []
    return [(text, urljoin(base_url, href)) for text, href in parser.links]


def text_from_html(raw_html: str) -> str:
    text = re.sub(r"<script\b.*?</script>", " ", raw_html, flags=re.I | re.S)
    text = re.sub(r"<style\b.*?</style>", " ", text, flags=re.I | re.S)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def google_news_rss_url(query: str) -> str:
    return "https://news.google.com/rss/search?" + urlencode({"q": query, "hl": "en-US", "gl": "US", "ceid": "US:en"})


def parse_google_news_rss(xml_text: str, query: str) -> list[SearchResult]:
    root = ET.fromstring(xml_text)
    results: list[SearchResult] = []
    for item in root.findall(".//item"):
        title = clean_text(item.findtext("title", ""))
        link = clean_text(item.findtext("link", ""))
        summary = clean_text(item.findtext("description", ""))
        publisher = clean_text(item.findtext("source", ""))
        published_at = clean_text(item.findtext("pubDate", ""))
        if title and link:
            results.append(SearchResult(query=query, title=title, link=link, summary=summary, publisher=publisher, published_at=published_at))
    return results


def article_title_without_publisher(title: str) -> str:
    if " - " in title:
        return title.rsplit(" - ", 1)[0].strip()
    return title.strip()


def known_company_from_text(text: str) -> str | None:
    lower_text = text.lower()
    for company, meta in COMPANY_REGISTRY.items():
        if any(alias.lower() in lower_text for alias in meta["aliases"]):
            return company
    return None


def tidy_company_name(candidate: str) -> str:
    candidate = re.sub(r"^[^A-Za-z0-9]+", "", candidate.strip())
    if re.search(r"['’]s\s+", candidate):
        candidate = re.split(r"['’]s\s+", candidate, maxsplit=1)[0]
    candidate = candidate.split(",", 1)[0]
    candidate = re.sub(
        r"\b(Exclusive|MedTech|Med-tech|Healthtech|Health-care|Digital health|Medical device|AI|Startup|Company|Maker|Developer|Platform|Provider|Firm|French|Estonian|Irish|Female-led|Europe-first|Vienna-based|SA-based|New)\b[:\s-]*",
        "",
        candidate,
        flags=re.I,
    )
    candidate = re.split(r"\s+(?:after|as|for|from|in|on|to|with)\s+", candidate, maxsplit=1, flags=re.I)[0]
    candidate = re.sub(r"['’]s$", "", candidate.strip(" :-,.;"))
    return re.sub(r"\s+", " ", candidate).strip()


def extract_company_from_search_result(result: SearchResult) -> str | None:
    combined = f"{result.title} {result.summary}"
    known = known_company_from_text(combined)
    if known:
        return known

    title = article_title_without_publisher(result.title)
    trigger_verbs = (
        r"raises?|raised|lands?|landed|secures?|secured|closes?|closed|bags?|bagged|nabs?|nabbed|"
        r"announces?|announced|launches?|launched|unveils?|unveiled|wins?|won|receives?|received|"
        r"gets?|got|earns?|earned|scores?|scored|obtains?|obtained|gains?|gained"
    )
    patterns = [
        rf"^(?P<company>[A-Z][A-Za-z0-9&.,'’\- ]{{1,80}}?)\s+(?:{trigger_verbs})\b",
        rf"^(?P<company>[A-Z][A-Za-z0-9&.,'’\- ]{{1,80}}?)\s+.*?\b(?:Series [A-Z]|seed|funding|FDA clearance|510\(k\)|De Novo|CE mark)\b",
        rf"\b(?:startup|company|maker|developer|provider)\s+(?P<company>[A-Z][A-Za-z0-9&.,'’\- ]{{1,60}}?)\s+(?:{trigger_verbs})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, title, flags=re.I)
        if not match:
            continue
        company = tidy_company_name(match.group("company"))
        if is_plausible_company_name(company):
            return company
    return None


def is_plausible_company_name(company: str) -> bool:
    if len(company) < 2 or len(company) > 70:
        return False
    if company[0].islower():
        return False
    words = company.split()
    if len(words) > 4:
        return False
    blocked = {
        "FDA", "AI", "The", "This", "What", "How", "Why", "Medical", "Digital", "Health",
        "Medtech", "More", "Big", "Paris", "Irish", "Diagnostics", "Assessing", "Quality",
        "Stroke", "Ultrasound", "Weekly", "Apple Watch", "X-Ray Software", "Powered MRI Software",
        "Next-Gen powered Coronary Imaging", "Six Aussie startups that",
    }
    blocked_terms = {
        "certified", "tool", "figure", "factors", "physiotherapist", "gpt-powered", "study",
        "trial", "programme", "program", "powered", "automating", "diagnostic", "next-gen",
        "x-ray", "software",
    }
    if company in blocked or company.lower() in blocked_terms:
        return False
    if any(word.lower() in blocked_terms for word in words):
        return False
    if words[-1].lower() in {"to", "for", "with", "and"}:
        return False
    return bool(re.search(r"[A-Za-z]", company))


def clean_page_candidate(text: str) -> str:
    candidate = clean_text(text)
    category_prefix = r"^(Acquired Core|Exited Core|IPO Core|MD Start|Capital|Core|Crossover|Digital Medicine|Industrial Biotech|Biovelocita)(?=[A-Z\s])\s*"
    previous = None
    while previous != candidate:
        previous = candidate
        candidate = re.sub(category_prefix, "", candidate)
    candidate = re.sub(r"\s+(Inc\.?|Ltd\.?|Limited|LLC|GmbH|AG|Corp\.?|Corporation|PLC)\b\.?", "", candidate, flags=re.I)
    candidate = candidate.strip(" :-,.;|")
    return re.sub(r"\s+", " ", candidate)


def is_plausible_page_candidate(candidate: str) -> bool:
    candidate = clean_page_candidate(candidate)
    if not is_plausible_company_name(candidate):
        return False
    lower = candidate.lower()
    blocked = {
        "about", "contact", "team", "news", "events", "portfolio", "companies", "apply", "program",
        "programs", "search", "login", "register", "privacy", "terms", "careers", "jobs", "blog",
        "resources", "learn more", "read more", "view all", "home", "menu", "investors", "partners",
        "speakers", "sponsors", "exhibitors", "agenda", "tickets", "subscribe", "download", "more",
        "skip to search", "skip to topics menu", "skip to common links", "follow fda", "en español",
        "food", "drugs", "medical devices", "radiation-emitting products", "vaccines, blood & biologics",
        "animal & veterinary", "cosmetics", "tobacco products", "fda home", "databases", "help",
        "product code", "quick search", "clear form", "de novo", "cfr title 21", "about us",
        "at a glance", "budget", "cms", "healthcare", "exhibiting at rsna", "eic portfolios",
        "linkedIn(opens in new window)", "linkedin-in", "donate",
    }
    if lower in blocked:
        return False
    if any(marker in lower for marker in ["@", "http", "cookie", "copyright", "javascript", "privacy policy", "linkedin", ", md", "sc.d", "mph"]):
        return False
    if re.search(r"\b(read|learn|view|see|contact|apply|register|download|submit)\b", lower):
        return False
    source_words = {
        "company", "companies", "cohort", "accelerator", "exhibitor", "exhibitors", "sponsor",
        "sponsors", "sponsorship", "portfolio", "dashboard", "application", "portal", "funding",
        "grant", "grants", "challenge", "challenges", "opportunities", "programme", "program",
        "resources", "directory", "profile", "presentation", "presentations", "videos", "news",
        "linkedin", "attending", "become", "interested", "planning", "partnership", "manager",
        "services", "material", "schedule", "agenda", "report", "committee", "council",
        "core", "acquired", "exited", "ipo", "bas", "budget", "cms", "host", "join",
        "log", "logistics", "manage", "investments", "international", "healthcare",
        "horizon", "advice", "contacts", "events", "dinner", "offer", "traffic",
    }
    if any(word.lower().strip("&") in source_words for word in candidate.split()):
        return False
    first_names = {
        "adaeze", "andrew", "anita", "brooke", "charlie", "dave", "david", "isaac",
        "jannine", "jeremy", "joe", "julius", "justin", "kate", "katherine", "matt",
        "matthew", "mike", "navin", "otto", "sam", "siva", "sylvia",
    }
    if len(candidate.split()) >= 2 and candidate.split()[0].lower().strip(",") in first_names:
        return False
    if lower.endswith((" us", " your award", " your offer")):
        return False
    return True


def is_relevant_candidate_link(source: Source, link_text: str, href: str) -> bool:
    lower_href = href.lower()
    if lower_href.startswith(("javascript:", "mailto:", "tel:")) or "#" in lower_href:
        return False
    negative_href_terms = [
        "team", "people", "staff", "advisor", "board", "speaker", "agenda", "event", "news",
        "blog", "login", "application", "apply", "contact", "about", "resource", "service",
    ]
    if any(term in lower_href for term in negative_href_terms):
        return False
    if source.source_type in {"Regulatory database", "Jobs"}:
        return False
    path_terms = {
        "Accelerator": ["compan", "startup", "cohort", "portfolio", "accelerator", "alumni"],
        "Conference": ["exhibitor", "sponsor", "startup", "compan", "showcase"],
        "VC portfolio": ["portfolio", "compan", "investment"],
        "Grant/funding": ["project", "award", "funding", "portfolio", "compan", "grant"],
    }
    terms = path_terms.get(source.source_type, [])
    if terms and not any(term in lower_href for term in terms):
        return False
    return is_plausible_page_candidate(link_text)


def infer_page_product_type(source: Source, context: str) -> str:
    text = f"{source.notes} {context}".lower()
    if "samd" in text:
        return "SaMD / health software"
    if "medical device" in text or "medtech" in text:
        return "Medical device / medtech"
    if "diagnostic" in text or "imaging" in text:
        return "Diagnostics / imaging"
    if "ai" in text:
        return "AI health / medtech"
    if "digital health" in text:
        return "Digital health"
    return source.source_type


def infer_yc_product_type(context: str) -> str:
    text = context.lower()
    if "medical device" in text or "medtech" in text:
        return "Medical device / medtech"
    if "diagnostic" in text or "imaging" in text or "radiology" in text:
        return "Diagnostics / imaging"
    if "clinic" in text or "provider" in text or "healthcare it" in text:
        return "Healthcare operations / IT"
    if "health insurance" in text or "insurance" in text:
        return "Health insurance / payer"
    if "therapeutic" in text or "biotech" in text or "gene therapy" in text or "drug" in text:
        return "Biotech / therapeutics"
    if "ai" in text or "artificial intelligence" in text:
        return "AI health"
    if "healthcare" in text or "digital health" in text:
        return "Digital health"
    return "Healthcare startup"


def source_type_trigger_event(source: Source, company: str) -> tuple[str, str] | None:
    trigger_type = SOURCE_TRIGGER_TYPES.get(source.source_type)
    if not trigger_type:
        return None
    if source.source_type == "Regulatory database":
        return trigger_type, f"{company} appeared on regulatory source '{source.name}'."
    if source.source_type == "Grant/funding":
        return trigger_type, f"{company} appeared on grant/funding source '{source.name}'."
    if source.source_type == "VC portfolio":
        return trigger_type, f"{company} appeared on investor portfolio source '{source.name}'."
    if source.source_type == "Jobs":
        return trigger_type, f"{company} appeared on jobs/careers signal source '{source.name}'."
    if source.source_type == "Accelerator":
        return trigger_type, f"{company} appeared on accelerator/cohort source '{source.name}'."
    if source.source_type == "Conference":
        return trigger_type, f"{company} appeared on conference/exhibitor source '{source.name}'."
    return None


def fetch_json(url: str, payload: dict) -> tuple[dict, str | None]:
    try:
        req = Request(
            url,
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "User-Agent": USER_AGENT,
                "Content-Type": "application/json",
                "X-Algolia-Application-Id": YC_ALGOLIA_APP_ID,
                "X-Algolia-API-Key": YC_ALGOLIA_API_KEY,
            },
        )
        raw = urlopen(req, timeout=30).read()
    except (OSError, URLError) as exc:
        return {}, str(exc)
    try:
        return json.loads(raw.decode("utf-8", "ignore")), None
    except json.JSONDecodeError as exc:
        return {}, f"JSON decode failed: {exc}"


def fetch_json_url(url: str) -> tuple[object, str | None]:
    try:
        req = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "application/json"})
        raw = urlopen(req, timeout=30).read()
    except (OSError, URLError) as exc:
        return {}, str(exc)
    try:
        return json.loads(raw.decode("utf-8", "ignore")), None
    except json.JSONDecodeError as exc:
        return {}, f"JSON decode failed: {exc}"


def configured_job_boards(platform: str) -> list[tuple[str, dict, str]]:
    boards: list[tuple[str, dict, str]] = []
    for company, meta in COMPANY_REGISTRY.items():
        for board in meta.get("job_boards", []):
            if board.get("platform") == platform and board.get("account"):
                boards.append((company, meta, board["account"]))
    return boards


def job_board_url(platform: str, account: str) -> str:
    if platform == "greenhouse":
        return f"https://boards-api.greenhouse.io/v1/boards/{account}/jobs?content=true"
    if platform == "lever":
        return f"https://api.lever.co/v0/postings/{account}?mode=json"
    if platform == "ashby":
        return f"https://api.ashbyhq.com/posting-api/job-board/{account}?includeCompensation=true"
    if platform == "workable":
        return f"https://apply.workable.com/api/v1/widget/accounts/{account}?details=true"
    if platform == "smartrecruiters":
        return f"https://api.smartrecruiters.com/v1/companies/{account}/postings"
    if platform == "recruitee":
        return f"https://{account}.recruitee.com/api/offers/"
    return ""


def html_to_text(value: str) -> str:
    return clean_text(re.sub(r"<[^>]+>", " ", html.unescape(value or "")))


def nested_text(value: object) -> str:
    if isinstance(value, str):
        return html_to_text(value)
    if isinstance(value, list):
        return " ".join(nested_text(item) for item in value)
    if isinstance(value, dict):
        return " ".join(nested_text(item) for item in value.values())
    return ""


def job_location(value: object) -> str:
    if isinstance(value, str):
        return clean_text(value)
    if isinstance(value, dict):
        return clean_text(value.get("name") or value.get("location") or value.get("city") or nested_text(value))
    if isinstance(value, list):
        return "; ".join(filter(None, [job_location(item) for item in value]))
    return ""


def parse_greenhouse_jobs(data: object) -> list[JobPosting]:
    jobs = data.get("jobs", []) if isinstance(data, dict) else []
    postings: list[JobPosting] = []
    for job in jobs:
        if not isinstance(job, dict):
            continue
        postings.append(
            JobPosting(
                title=clean_text(job.get("title", "")),
                url=job.get("absolute_url") or job.get("internal_job_id") or "",
                description=html_to_text(job.get("content", "")),
                location=job_location(job.get("location")),
                department=job_location(job.get("departments")),
            )
        )
    return postings


def parse_lever_jobs(data: object) -> list[JobPosting]:
    jobs = data if isinstance(data, list) else []
    postings: list[JobPosting] = []
    for job in jobs:
        if not isinstance(job, dict):
            continue
        postings.append(
            JobPosting(
                title=clean_text(job.get("text", "")),
                url=job.get("hostedUrl") or job.get("applyUrl") or "",
                description=html_to_text(job.get("description") or job.get("descriptionPlain") or nested_text(job.get("lists", []))),
                location=job_location(job.get("categories", {}).get("location") if isinstance(job.get("categories"), dict) else ""),
                department=job_location(job.get("categories", {}).get("team") if isinstance(job.get("categories"), dict) else ""),
            )
        )
    return postings


def parse_ashby_jobs(data: object) -> list[JobPosting]:
    jobs = data.get("jobs", []) if isinstance(data, dict) else []
    postings: list[JobPosting] = []
    for job in jobs:
        if not isinstance(job, dict):
            continue
        postings.append(
            JobPosting(
                title=clean_text(job.get("title", "")),
                url=job.get("jobUrl") or job.get("url") or "",
                description=html_to_text(job.get("descriptionHtml") or job.get("description") or ""),
                location=job_location(job.get("location") or job.get("locations")),
                department=job_location(job.get("department")),
            )
        )
    return postings


def parse_workable_jobs(data: object) -> list[JobPosting]:
    jobs = data.get("jobs") or data.get("results") or [] if isinstance(data, dict) else []
    postings: list[JobPosting] = []
    for job in jobs:
        if not isinstance(job, dict):
            continue
        url = job.get("url") or job.get("shortlink") or job.get("application_url") or ""
        shortcode = job.get("shortcode")
        if not url and shortcode:
            url = f"https://apply.workable.com/j/{shortcode}/"
        postings.append(
            JobPosting(
                title=clean_text(job.get("title", "")),
                url=url,
                description=html_to_text(job.get("description") or nested_text(job.get("requirements", ""))),
                location=job_location(job.get("location") or job.get("locations")),
                department=job_location(job.get("department")),
            )
        )
    return postings


def parse_smartrecruiters_jobs(data: object) -> list[JobPosting]:
    jobs = data.get("content") or data.get("jobs") or [] if isinstance(data, dict) else []
    postings: list[JobPosting] = []
    for job in jobs:
        if not isinstance(job, dict):
            continue
        ref = job.get("ref") if isinstance(job.get("ref"), str) else ""
        postings.append(
            JobPosting(
                title=clean_text(job.get("name") or job.get("title") or ""),
                url=job.get("url") or ref,
                description=html_to_text(job.get("jobAd", {}).get("sections", {}).get("jobDescription", {}).get("text", "") if isinstance(job.get("jobAd"), dict) else job.get("description", "")),
                location=job_location(job.get("location")),
                department=job_location(job.get("department")),
            )
        )
    return postings


def parse_recruitee_jobs(data: object) -> list[JobPosting]:
    jobs = data.get("offers") or data.get("jobs") or [] if isinstance(data, dict) else []
    postings: list[JobPosting] = []
    for job in jobs:
        if not isinstance(job, dict):
            continue
        postings.append(
            JobPosting(
                title=clean_text(job.get("title", "")),
                url=job.get("careers_url") or job.get("url") or "",
                description=html_to_text(job.get("description") or job.get("description_html") or ""),
                location=job_location(job.get("location") or job.get("locations")),
                department=job_location(job.get("department")),
            )
        )
    return postings


JOB_PARSERS = {
    "greenhouse": parse_greenhouse_jobs,
    "lever": parse_lever_jobs,
    "ashby": parse_ashby_jobs,
    "workable": parse_workable_jobs,
    "smartrecruiters": parse_smartrecruiters_jobs,
    "recruitee": parse_recruitee_jobs,
}


def relevant_job_terms(posting: JobPosting, company_meta: dict) -> list[str]:
    text = " ".join([posting.title, posting.description, posting.department, posting.location, company_meta.get("product_type", "")]).lower()
    role_terms = [
        "regulatory", "regulatory affairs", "quality", "quality engineer", "design assurance",
        "v&v", "verification", "validation", "clinical", "medical device", "digital health",
        "samd", "fda", "qa",
    ]
    context_terms = ["ai", "machine learning", "medical device", "medtech", "health", "clinical", "diagnostic", "imaging", "digital health", "samd", "fda"]
    matched = [term for term in role_terms if term in text]
    title_lower = posting.title.lower()
    if not matched and re.search(r"\b(engineer|engineering|product|software|ml|ai)\b", title_lower) and any(term in text for term in context_terms):
        matched = [term for term in context_terms if term in text]
    return matched


def run_job_board_adapter(source: Source, platform: str) -> tuple[list[DiscoveryHit], list[TriggerEvent], str]:
    boards = configured_job_boards(platform)
    if not boards:
        return [], [], f"No registry companies configured for {platform} job boards"
    discovery_hits: list[DiscoveryHit] = []
    trigger_events: list[TriggerEvent] = []
    errors: list[str] = []
    no_matches = 0
    fetched = 0
    for company, meta, account in boards:
        data, error = fetch_json_url(job_board_url(platform, account))
        if error:
            errors.append(f"{company}: {error}")
            continue
        fetched += 1
        postings = JOB_PARSERS[platform](data)
        matched_jobs = [posting for posting in postings if posting.title and relevant_job_terms(posting, meta)]
        if not matched_jobs:
            no_matches += 1
            continue
        selected = matched_jobs[:MAX_MATCHED_JOBS_PER_COMPANY]
        role_titles = "; ".join(posting.title for posting in selected)
        evidence_url = selected[0].url or meta.get("website") or source.url
        matched_terms = sorted({term for posting in selected for term in relevant_job_terms(posting, meta)})
        discovery_hits.append(
            DiscoveryHit(
                company=company,
                source_name=source.name,
                source_type=source.source_type,
                discovery_url=evidence_url,
                discovery_rationale=f"{company} has relevant open roles on {source.name}: {role_titles}",
                product_type=meta.get("product_type", ""),
                geography=meta.get("geography", source.geography),
                website=meta.get("website", ""),
                matched_terms=f"adapter: {source.adapter}; roles: {role_titles}; terms: {', '.join(matched_terms)}",
                company_description=" ".join([posting.description for posting in selected])[:1000],
            )
        )
        trigger_events.append(
            TriggerEvent(
                company=company,
                trigger_type="Hiring signal",
                trigger_event=f"Relevant hiring signal from {source.name}: {role_titles}",
                trigger_source=source.name,
                evidence_url=evidence_url,
            )
        )
    result = f"{fetched}/{len(boards)} configured boards fetched; {len(discovery_hits)} discovery hits; {len(trigger_events)} trigger events; {no_matches} boards with no matching jobs"
    if errors:
        result += "; errors: " + " | ".join(errors[:5])
    return discovery_hits, trigger_events, result


def yc_company_url(hit: dict) -> str:
    if hit.get("ycdc_company_url"):
        return hit["ycdc_company_url"]
    slug = hit.get("slug") or hit.get("objectID")
    return f"https://www.ycombinator.com/companies/{slug}" if slug else "https://www.ycombinator.com/companies"


def run_yc_healthcare(source: Source, max_hits: int = 1000) -> tuple[list[DiscoveryHit], list[TriggerEvent], str]:
    endpoint = f"https://{YC_ALGOLIA_APP_ID}-dsn.algolia.net/1/indexes/YCCompany_production/query"
    hits: list[dict] = []
    page = 0
    nb_hits = 0
    while len(hits) < max_hits:
        payload = {"query": YC_HEALTHCARE_QUERY, "hitsPerPage": 100, "page": page}
        data, error = fetch_json(endpoint, payload)
        if error:
            return [], [], f"YC Algolia fetch failed: {error}"
        nb_hits = int(data.get("nbHits") or 0)
        page_hits = data.get("hits") or []
        hits.extend(page_hits)
        if page >= int(data.get("nbPages") or 0) - 1 or not page_hits:
            break
        page += 1

    hits = sorted(hits[:max_hits], key=lambda item: item.get("launched_at") or 0, reverse=True)
    discovery_hits: list[DiscoveryHit] = []
    trigger_events: list[TriggerEvent] = []
    seen: set[str] = set()
    for hit in hits:
        company = clean_page_candidate(hit.get("name") or "")
        if not company or company.lower() in seen:
            continue
        seen.add(company.lower())
        url = yc_company_url(hit)
        tags = ", ".join(hit.get("tags") or [])
        batch = hit.get("batch") or hit.get("batch_name") or ""
        description = hit.get("one_liner") or hit.get("long_description") or ""
        matched = f"query: {YC_HEALTHCARE_QUERY}; batch: {batch}; tags: {tags}".strip("; ")
        discovery_hits.append(
            DiscoveryHit(
                company=company,
                source_name=f"{source.name}: {YC_HEALTHCARE_QUERY} search",
                source_type=source.source_type,
                discovery_url=url,
                discovery_rationale=f"YC company directory search for '{YC_HEALTHCARE_QUERY}', sorted by launch date, returned this company.",
                product_type=infer_yc_product_type(f"{description} {tags}"),
                geography=hit.get("all_locations") or hit.get("location") or source.geography,
                website=hit.get("website") or "",
                matched_terms=matched,
            )
        )
        trigger = source_type_trigger_event(source, company)
        if trigger:
            trigger_events.append(TriggerEvent(company, trigger[0], trigger[1], f"{source.name}: {YC_HEALTHCARE_QUERY} search", url))
    return discovery_hits, trigger_events, f"YC Algolia query '{YC_HEALTHCARE_QUERY}'; {nb_hits} matches; {len(discovery_hits)} discovery hits"


def infer_cohort_year(*values: str) -> str:
    for value in values:
        match = re.search(r"\b(20\d{2})\b", value or "")
        if match:
            return match.group(1)
    return ""


def infer_accelerator_product_type(context: str) -> str:
    text = context.lower()
    if any(term in text for term in ["medical device", "medtech", "device"]):
        return "Medical device / medtech"
    if any(term in text for term in ["diagnostic", "imaging", "radiology", "biomarker"]):
        return "Diagnostics / imaging"
    if any(term in text for term in ["ehr", "workflow", "provider", "hospital", "clinical workflow"]):
        return "Healthcare operations / IT"
    if any(term in text for term in ["ai", "machine learning", "artificial intelligence"]):
        return "AI health"
    if any(term in text for term in ["digital health", "virtual care", "remote monitoring", "platform"]):
        return "Digital health"
    if any(term in text for term in ["therapeutic", "biotech", "pharma", "drug"]):
        return "Biotech / therapeutics"
    return "Accelerator company"


def accelerator_trigger(source: Source, hit: DiscoveryHit) -> TriggerEvent | None:
    trigger = source_type_trigger_event(source, hit.company)
    if not trigger:
        return None
    detail = trigger[1]
    if hit.cohort_label:
        detail = f"{hit.company} appeared in {hit.cohort_label} for '{source.name}'."
    return TriggerEvent(hit.company, trigger[0], detail, hit.source_name, hit.discovery_url)


def make_accelerator_hit(
    source: Source,
    company: str,
    evidence_url: str,
    *,
    accelerator_program: str | None = None,
    cohort_label: str = "",
    cohort_year: str = "",
    category_or_track: str = "",
    company_description: str = "",
    website: str = "",
    geography: str = "",
    matched_terms: str = "",
) -> DiscoveryHit | None:
    company = clean_page_candidate(company)
    if not is_plausible_page_candidate(company):
        return None
    context = " ".join([category_or_track, company_description, source.notes])
    rationale = f"{source.name} adapter extracted this company from an accelerator cohort/directory source."
    if cohort_label:
        rationale += f" Cohort/source label: {cohort_label}."
    return DiscoveryHit(
        company=company,
        source_name=source.name,
        source_type=source.source_type,
        discovery_url=evidence_url,
        discovery_rationale=rationale,
        product_type=infer_accelerator_product_type(context),
        geography=geography or source.geography,
        website=website,
        matched_terms=matched_terms or f"adapter: {source.adapter}",
        accelerator_program=accelerator_program or source.name,
        cohort_label=cohort_label,
        cohort_year=cohort_year,
        category_or_track=category_or_track,
        company_description=company_description,
    )


def dedupe_hits_with_triggers(source: Source, hits: list[DiscoveryHit]) -> tuple[list[DiscoveryHit], list[TriggerEvent]]:
    deduped: list[DiscoveryHit] = []
    triggers: list[TriggerEvent] = []
    seen: set[tuple[str, str]] = set()
    for hit in hits:
        key = (hit.company.lower(), hit.discovery_url)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(hit)
        trigger = accelerator_trigger(source, hit)
        if trigger:
            triggers.append(trigger)
    return deduped, triggers


def extract_meta_description(raw_html: str) -> str:
    patterns = [
        r'<meta[^>]+property=["\']og:description["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+name=["\']description["\'][^>]+content=["\']([^"\']+)["\']',
    ]
    for pattern in patterns:
        match = re.search(pattern, raw_html, flags=re.I | re.S)
        if match:
            return clean_text(match.group(1))
    paragraphs = re.findall(r"<p\b[^>]*>(.*?)</p>", raw_html, flags=re.I | re.S)
    for paragraph in paragraphs:
        text = clean_text(paragraph)
        if len(text) >= 30:
            return text
    return ""


def context_after_link(raw_html: str, href: str, window: int = 1200) -> str:
    idx = raw_html.find(href)
    if idx < 0:
        return ""
    return text_from_html(raw_html[idx : idx + window])


def parse_digitalhealth_london_page(source: Source, raw_html: str, page_url: str, profile_html_by_url: dict[str, str] | None = None) -> list[DiscoveryHit]:
    hits: list[DiscoveryHit] = []
    profile_html_by_url = profile_html_by_url or {}
    seen_urls: set[str] = set()
    for link_text, href in extract_links(raw_html, page_url):
        if "/innovation-directory/profile/" not in href or href in seen_urls:
            continue
        seen_urls.add(href)
        company = re.split(r"\s+Company\s+", link_text, maxsplit=1, flags=re.I)[0]
        card_description = ""
        if re.search(r"\s+Company\s+", link_text, flags=re.I):
            card_description = re.split(r"\s+Company\s+", link_text, maxsplit=1, flags=re.I)[1]
        profile_html = profile_html_by_url.get(href, "")
        profile_text = text_from_html(profile_html) if profile_html else context_after_link(raw_html, href)
        description = extract_meta_description(profile_html) if profile_html else card_description or profile_text
        cohort = ""
        cohort_match = re.search(r"\b(?:Accelerator|Launchpad|Cohort)\s*(?:cohort)?\s*(20\d{2})\b", profile_text, flags=re.I)
        if cohort_match:
            cohort = f"DigitalHealth.London {cohort_match.group(1)}"
        track_parts = []
        for label in ["Sector", "Technology", "Area innovation", "Area of innovation"]:
            match = re.search(rf"{label}\s*:?\s*([A-Za-z0-9 /,&+-]{{3,80}})", profile_text, flags=re.I)
            if match:
                track_parts.append(clean_text(match.group(1)))
        hit = make_accelerator_hit(
            source,
            company,
            href,
            cohort_label=cohort,
            cohort_year=infer_cohort_year(cohort, profile_text),
            category_or_track="; ".join(dict.fromkeys(track_parts)),
            company_description=description,
            matched_terms="adapter: digitalhealth_london; profile link",
        )
        if hit:
            hits.append(hit)
    return hits


def run_digitalhealth_london(source: Source, max_pages: int = 50) -> tuple[list[DiscoveryHit], list[TriggerEvent], str]:
    base_url = ACCELERATOR_SOURCE_PAGES[source.name][0]
    all_hits: list[DiscoveryHit] = []
    errors: list[str] = []
    page = 1
    while page <= max_pages:
        url = base_url if page == 1 else f"{base_url}/page/{page}"
        raw_html, error = fetch_raw_text(url)
        if error:
            if page == 1:
                errors.append(f"{url}: {error}")
            break
        page_hits = parse_digitalhealth_london_page(source, raw_html, url)
        if not page_hits:
            break
        all_hits.extend(page_hits)
        if f"/companies/page/{page + 1}" not in raw_html:
            break
        page += 1
    hits, triggers = dedupe_hits_with_triggers(source, all_hits)
    result = f"{page} directory pages scanned; {len(hits)} discovery hits; {len(triggers)} trigger events"
    if errors:
        result += "; errors: " + " | ".join(errors)
    return hits, triggers, result


def parse_medtech_innovator_showcase(source: Source, raw_html: str, page_url: str, cohort_label: str) -> list[DiscoveryHit]:
    hits: list[DiscoveryHit] = []
    current_track = ""
    parts = re.split(r"(<h[1-4]\b.*?</h[1-4]>|<a\b.*?</a>)", raw_html, flags=re.I | re.S)
    for part in parts:
        if re.match(r"<h[1-4]\b", part, flags=re.I):
            heading = clean_text(part)
            if heading and len(heading) <= 80:
                current_track = heading
            continue
        if not re.match(r"<a\b", part, flags=re.I):
            continue
        href_match = re.search(r"href=[\"']([^\"']+)[\"']", part, flags=re.I)
        text = clean_text(part)
        href = urljoin(page_url, href_match.group(1)) if href_match else page_url
        if not text or text.lower() in {"website", "learn more", "view profile"}:
            continue
        if not current_track:
            continue
        if not is_relevant_candidate_link(Source(source.name, source.source_type, page_url, source.geography, source.priority, source.update_cadence, source.extraction_method, source.notes, "accelerator_page"), text, href):
            host = urlparse(href).netloc.lower()
            if any(domain in host for domain in ["medtechinnovator.org", "medtechinnovator.asia", "biotoolsinnovator.org", "pro.innovator.org"]):
                continue
        hit = make_accelerator_hit(
            source,
            text,
            href,
            cohort_label=cohort_label,
            cohort_year=infer_cohort_year(cohort_label, page_url),
            category_or_track=current_track,
            company_description=context_after_link(raw_html, href),
            matched_terms="adapter: medtech_innovator; showcase/company link",
        )
        if hit:
            hits.append(hit)
    return hits


def pory_value(fields: dict, *names: str) -> str:
    for name in names:
        value = fields.get(name)
        if value is None:
            continue
        if isinstance(value, list):
            return ", ".join(clean_text(str(item)) for item in value if clean_text(str(item)))
        return clean_text(str(value))
    return ""


def parse_medtech_innovator_pory_records(source: Source, records: list[dict]) -> list[DiscoveryHit]:
    hits: list[DiscoveryHit] = []
    for record in records:
        fields = record.get("fields") or {}
        company = pory_value(fields, "Company", "fld4BXQlB6TZgMSJa")
        if not company:
            continue
        year = pory_value(fields, "Year.", "Year", "flddSysOAMNrvWgLE")
        program = pory_value(fields, "Program.", "Program", "fldtA2KHphCLDbQXy") or "MedTech Innovator portfolio"
        if program in {"US", "APAC", "BTI"}:
            program = f"MedTech Innovator {program}"
        category = pory_value(
            fields,
            "Clinical Categories",
            "Device Categories",
            "Digital Categories",
            "Diagnostic Categories",
            "Thematic Categories",
            "Primary Industry Group",
        )
        description = pory_value(fields, "Product Short Description", "Description", "Long Description", "fld3ZsMoNwuiP15vS")
        website = pory_value(fields, "Website", "fldSEQbd6GpIJBF6J")
        geography = pory_value(fields, "Company Country/Territory", "Company Country/Territory (Old Field)", "Country")
        hit = make_accelerator_hit(
            source,
            company,
            MEDTECH_INNOVATOR_PORY_APP_URL,
            accelerator_program=program,
            cohort_label=f"{program} {year}".strip(),
            cohort_year=year,
            category_or_track=category,
            company_description=description,
            website=website,
            geography=geography,
            matched_terms="adapter: medtech_innovator; pory portfolio records",
        )
        if hit:
            hits.append(hit)
    return hits


def fetch_medtech_innovator_pory_records(max_pages: int = 100) -> tuple[list[dict], list[str]]:
    records: list[dict] = []
    errors: list[str] = []
    offset = ""
    for _ in range(max_pages):
        params = {"pageSize": 100}
        if offset:
            params["offset"] = offset
        url = MEDTECH_INNOVATOR_PORY_RECORDS_URL + "?" + urlencode(params)
        data, error = fetch_json_url(url)
        if error:
            errors.append(f"{url}: {error}")
            break
        page_records = data.get("records") if isinstance(data, dict) else []
        if not page_records:
            break
        records.extend(page_records)
        offset = data.get("offset") if isinstance(data, dict) else ""
        if not offset:
            break
    return records, errors


def run_medtech_innovator(source: Source) -> tuple[list[DiscoveryHit], list[TriggerEvent], str]:
    cohort_url = ACCELERATOR_SOURCE_PAGES[source.name][0]
    cohort_html, error = fetch_raw_text(cohort_url)
    errors: list[str] = []
    if error:
        errors.append(f"{cohort_url}: {error}")
        cohort_html = ""
    expected = 65
    expected_match = re.search(r"\b(\d{2,3})\s+(?:companies|startups)\b", text_from_html(cohort_html), flags=re.I)
    if expected_match:
        expected = int(expected_match.group(1))
    cohort_label = f"MedTech Innovator {infer_cohort_year(cohort_url, cohort_html) or TODAY[:4]} cohort"
    showcase_urls = []
    for _, href in extract_links(cohort_html, cohort_url):
        if "pro.innovator.org/showcase" in href or "flippingbook.com" in href:
            showcase_urls.append(href)
    hits: list[DiscoveryHit] = []
    current_hits: list[DiscoveryHit] = []
    for url in dict.fromkeys(showcase_urls):
        showcase_html, showcase_error = fetch_raw_text(url)
        if showcase_error:
            errors.append(f"{url}: {showcase_error}")
            continue
        current_hits.extend(parse_medtech_innovator_showcase(source, showcase_html, url, cohort_label))
    if not current_hits and cohort_html:
        current_hits.extend(parse_medtech_innovator_showcase(source, cohort_html, cohort_url, cohort_label))
    hits.extend(current_hits)

    pory_records, pory_errors = fetch_medtech_innovator_pory_records()
    errors.extend(pory_errors)
    hits.extend(parse_medtech_innovator_pory_records(source, pory_records))

    hits, triggers = dedupe_hits_with_triggers(source, hits)
    result = f"{len(showcase_urls) or 1} cohort/showcase pages scanned; {len(pory_records)} Pory portfolio records; {len(hits)} discovery hits; {len(triggers)} trigger events"
    if len(current_hits) < expected:
        result = f"INCOMPLETE current-cohort extraction: expected about {expected} current-cohort companies; found {len(current_hits)} in HTML/showcase. " + result
    if errors:
        result += "; errors: " + " | ".join(errors)
    return hits, triggers, result


def parse_mayo_accelerate_page(source: Source, raw_html: str, page_url: str) -> list[DiscoveryHit]:
    hits: list[DiscoveryHit] = []
    cohort_year = infer_cohort_year(raw_html, page_url)
    chunks = re.split(r"(<h[2-4]\b.*?</h[2-4]>)", raw_html, flags=re.I | re.S)
    for idx, chunk in enumerate(chunks):
        if not re.match(r"<h[2-4]\b", chunk, flags=re.I):
            continue
        company = clean_text(chunk)
        body = text_from_html(" ".join(chunks[idx + 1 : idx + 3]))
        if not body or len(body) < 20:
            continue
        hit = make_accelerator_hit(
            source,
            company,
            page_url,
            cohort_label=f"Mayo Clinic Platform Accelerate {cohort_year}".strip(),
            cohort_year=cohort_year,
            company_description=body,
            matched_terms="adapter: mayo_accelerate; cohort heading",
        )
        if hit:
            hits.append(hit)
    if not hits and "Markdown Content:" in raw_html:
        hits.extend(parse_mayo_accelerate_markdown(source, raw_html, page_url))
    return hits


def markdown_to_text(value: str) -> str:
    value = re.sub(r"\[!\[[^\]]+\]\([^)]+\)\]\([^)]+\)", " ", value)
    value = re.sub(r"!\[[^\]]+\]\([^)]+\)", " ", value)
    value = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", value)
    value = value.replace("**", "")
    return clean_text(value)


def parse_mayo_accelerate_markdown(source: Source, markdown: str, page_url: str) -> list[DiscoveryHit]:
    hits: list[DiscoveryHit] = []
    body = markdown.split("Markdown Content:", 1)[-1]
    body = body.split("### Interested in learning more?", 1)[0]
    body = body.split("## Meet the Newest Cohort", 1)[-1]
    cohort_year = infer_cohort_year(markdown, page_url)
    sections = re.split(r"\n(?=\s*(?:\[!\[Image|\!\[Image))", body)
    for section in sections:
        if not re.search(r"(?:\[!\[Image|\!\[Image)", section):
            continue
        website = ""
        linked_image = re.search(r"\[!\[[^\]]+\]\([^)]+\)\]\((https?://[^)]+)\)", section)
        if linked_image:
            website = linked_image.group(1)
        text = markdown_to_text(section)
        if len(text) < 30:
            continue
        quoted_name = re.match(r"[\"“]([^\"”]+)[\"”]\s+is\b", text)
        bold_names = [clean_text(match) for match in re.findall(r"\*\*([^*]+)\*\*", section)]
        company = quoted_name.group(1) if quoted_name else (bold_names[0] if bold_names else "")
        company = re.sub(r"[’']s$", "", company)
        if not company:
            continue
        hit = make_accelerator_hit(
            source,
            company,
            page_url,
            cohort_label=f"Mayo Clinic Platform Accelerate {cohort_year}".strip(),
            cohort_year=cohort_year,
            company_description=text,
            website=website,
            matched_terms="adapter: mayo_accelerate; live reader page",
        )
        if hit:
            hits.append(hit)
    return hits


def mayo_reader_url(url: str) -> str:
    return MAYO_READER_PREFIX + url


def run_mayo_accelerate(source: Source) -> tuple[list[DiscoveryHit], list[TriggerEvent], str]:
    urls = ACCELERATOR_SOURCE_PAGES[source.name]
    hits: list[DiscoveryHit] = []
    errors: list[str] = []
    scanned = 0
    for url in urls:
        raw_html, error = fetch_raw_text(url)
        scanned += 1
        if error:
            errors.append(f"{url}: {error}")
        else:
            hits.extend(parse_mayo_accelerate_page(source, raw_html, url))
        if hits:
            continue
        reader_url = mayo_reader_url(url)
        reader_text, reader_error = fetch_raw_text(reader_url)
        scanned += 1
        if reader_error:
            errors.append(f"{reader_url}: {reader_error}")
            continue
        hits.extend(parse_mayo_accelerate_page(source, reader_text, url))
    hits, triggers = dedupe_hits_with_triggers(source, hits)
    result = f"{scanned} cohort/live-reader pages scanned; {len(hits)} discovery hits; {len(triggers)} trigger events"
    if not hits:
        result = "INCOMPLETE Mayo extraction: live Mayo page and reader fetch returned no companies. " + result
    if errors:
        result += "; errors: " + " | ".join(errors)
    return hits, triggers, result


def parse_eit_health_catapult_page(source: Source, raw_html: str, page_url: str) -> list[DiscoveryHit]:
    hits: list[DiscoveryHit] = []
    previous_match = re.search(r"previous\s+winners", raw_html, flags=re.I)
    if previous_match:
        raw_html = raw_html[previous_match.start() :]
    page_text = text_from_html(raw_html)
    default_year = infer_cohort_year(page_text, page_url)
    current_track = ""
    current_award = ""
    parts = re.split(r"(<h[2-4]\b.*?</h[2-4]>|<img\b[^>]*>|<a\b.*?</a>)", raw_html, flags=re.I | re.S)
    for part in parts:
        if re.match(r"<h[2-4]\b", part, flags=re.I):
            heading = clean_text(part)
            if re.search(r"\b(BioTech|MedTech|Digital Health|Winner|Award|Edition|Final)\b", heading, flags=re.I):
                if "winner" in heading.lower() or "award" in heading.lower():
                    current_award = heading
                else:
                    current_track = heading
            continue
        company = ""
        href = page_url
        if re.match(r"<img\b", part, flags=re.I):
            alt_match = re.search(r"alt=[\"']([^\"']+)[\"']", part, flags=re.I)
            company = clean_text(alt_match.group(1)) if alt_match else ""
        elif re.match(r"<a\b", part, flags=re.I):
            href_match = re.search(r"href=[\"']([^\"']+)[\"']", part, flags=re.I)
            href = urljoin(page_url, href_match.group(1)) if href_match else page_url
            company = clean_text(part)
        if not company:
            continue
        if not current_track and not current_award:
            continue
        track = "; ".join([value for value in [current_track, current_award] if value])
        hit = make_accelerator_hit(
            source,
            company,
            href,
            cohort_label=f"EIT Health Catapult {default_year}".strip(),
            cohort_year=default_year,
            category_or_track=track,
            company_description=context_after_link(raw_html, href),
            matched_terms="adapter: eit_health_catapult; previous winner/finalist",
        )
        if hit:
            hits.append(hit)
    return hits


def run_eit_health_catapult(source: Source) -> tuple[list[DiscoveryHit], list[TriggerEvent], str]:
    url = ACCELERATOR_SOURCE_PAGES[source.name][0]
    raw_html, error = fetch_raw_text(url)
    if error:
        return [], [], f"{url}: {error}"
    hits, triggers = dedupe_hits_with_triggers(source, parse_eit_health_catapult_page(source, raw_html, url))
    return hits, triggers, f"Previous winners/finalists page scanned; {len(hits)} discovery hits; {len(triggers)} trigger events"


def build_source_page_evidence(source: Source, raw_html: str, max_candidates: int = 25) -> tuple[list[DiscoveryHit], list[TriggerEvent]]:
    text = text_from_html(raw_html)
    discovery_hits = find_companies_on_source(source, text)
    trigger_events: list[TriggerEvent] = []
    seen_companies = {hit.company.lower() for hit in discovery_hits}

    for hit in discovery_hits:
        trigger = source_type_trigger_event(source, hit.company)
        if trigger:
            trigger_events.append(TriggerEvent(hit.company, trigger[0], trigger[1], source.name, hit.discovery_url))

    for link_text, href in extract_links(raw_html, source.url):
        if not is_relevant_candidate_link(source, link_text, href):
            continue
        company = clean_page_candidate(link_text)
        if company.lower() in seen_companies:
            continue
        seen_companies.add(company.lower())
        matched_terms = f"adapter: {source.adapter}; link text"
        discovery_hits.append(
            DiscoveryHit(
                company=company,
                source_name=source.name,
                source_type=source.source_type,
                discovery_url=href,
                discovery_rationale=f"{source.source_type} adapter found company-like link text on '{source.name}'.",
                product_type=infer_page_product_type(source, link_text),
                geography=source.geography,
                website="",
                matched_terms=matched_terms,
            )
        )
        trigger = source_type_trigger_event(source, company)
        if trigger:
            trigger_events.append(TriggerEvent(company, trigger[0], trigger[1], source.name, href))
        if len(discovery_hits) >= max_candidates:
            break
    return discovery_hits, trigger_events


def run_accelerator_pages(source: Source) -> tuple[list[DiscoveryHit], list[TriggerEvent], str]:
    urls = ACCELERATOR_SOURCE_PAGES.get(source.name, [source.url])
    all_hits: list[DiscoveryHit] = []
    all_triggers: list[TriggerEvent] = []
    errors: list[str] = []
    seen: set[tuple[str, str]] = set()
    for url in urls:
        page_source = Source(source.name, source.source_type, url, source.geography, source.priority, source.update_cadence, source.extraction_method, source.notes, source.adapter)
        raw_html, error = fetch_raw_text(url)
        if error:
            errors.append(f"{url}: {error}")
            continue
        hits, triggers = build_source_page_evidence(page_source, raw_html)
        for hit in hits:
            key = (hit.company.lower(), hit.discovery_url)
            if key in seen:
                continue
            seen.add(key)
            all_hits.append(hit)
        all_triggers.extend(triggers)
    result = f"{len(urls)} accelerator pages; {len(all_hits)} discovery hits; {len(all_triggers)} trigger events"
    if errors:
        result += "; errors: " + " | ".join(errors)
    return all_hits, all_triggers, result


def classify_search_trigger(result: SearchResult) -> tuple[str, str] | None:
    text = f"{result.title} {result.summary}".lower()
    title = article_title_without_publisher(result.title)
    if re.search(r"\b(series [a-z]|seed|pre-seed|funding|raises?|raised|lands?|landed|secures?|secured|closes?|closed|\$\d)", text):
        return "Funding", f"Funding signal from Google News result: {title}"
    if re.search(r"\b(fda clearance|fda clears|510\(k\)|de novo|ce mark|regulatory clearance|clearance|cleared)\b", text):
        return "Regulatory clearance", f"Regulatory clearance signal from Google News result: {title}"
    if re.search(r"\b(launches?|launched|unveils?|unveiled|commercial launch)\b", text):
        return "Launch", f"Launch signal from Google News result: {title}"
    if re.search(r"\b(approval|approved|authori[sz]ed)\b", text):
        return "Approval", f"Approval signal from Google News result: {title}"
    return None


def infer_product_type(result: SearchResult, company: str) -> str:
    registry_meta = COMPANY_REGISTRY.get(company)
    if registry_meta:
        return registry_meta["product_type"]
    text = f"{result.query} {result.title} {result.summary}".lower()
    if "samd" in text:
        return "SaMD / health software"
    if "medical device" in text or "device" in text:
        return "Medical device"
    if "fda" in text or "clearance" in text or "regulatory" in text:
        return "Regulated digital health / device"
    if "ai" in text:
        return "AI health / medtech"
    return "Digital health / medtech"


def infer_geography(company: str) -> str:
    registry_meta = COMPANY_REGISTRY.get(company)
    return registry_meta["geography"] if registry_meta else "Unknown"


def infer_website(company: str) -> str:
    registry_meta = COMPANY_REGISTRY.get(company)
    return registry_meta["website"] if registry_meta else ""


def search_result_matched_terms(result: SearchResult, trigger_type: str | None) -> str:
    terms = [f"query: {result.query}"]
    if trigger_type:
        terms.append(f"trigger: {trigger_type}")
    return "; ".join(terms)


def build_google_news_evidence(source: Source, results: list[SearchResult]) -> tuple[list[DiscoveryHit], list[TriggerEvent]]:
    discovery_hits: list[DiscoveryHit] = []
    trigger_events: list[TriggerEvent] = []
    seen_discovery: set[tuple[str, str]] = set()
    seen_trigger: set[tuple[str, str, str]] = set()

    for result in results:
        company = extract_company_from_search_result(result)
        if not company:
            continue
        trigger = classify_search_trigger(result)
        discovery_key = (company.lower(), result.link)
        if discovery_key not in seen_discovery:
            seen_discovery.add(discovery_key)
            discovery_hits.append(
                DiscoveryHit(
                    company=company,
                    source_name=f"{source.name}: {result.query}",
                    source_type=source.source_type,
                    discovery_url=result.link,
                    discovery_rationale=f"Google News result for query '{result.query}' named the company in the title/snippet.",
                    product_type=infer_product_type(result, company),
                    geography=infer_geography(company),
                    website=infer_website(company),
                    matched_terms=search_result_matched_terms(result, trigger[0] if trigger else None),
                )
            )
        if trigger:
            trigger_key = (company.lower(), trigger[0], result.link)
            if trigger_key in seen_trigger:
                continue
            seen_trigger.add(trigger_key)
            trigger_events.append(
                TriggerEvent(
                    company=company,
                    trigger_type=trigger[0],
                    trigger_event=trigger[1],
                    trigger_source=f"{source.name}: {result.query}",
                    evidence_url=result.link,
                )
            )
    return discovery_hits, trigger_events


def run_google_news_search(source: Source) -> tuple[list[DiscoveryHit], list[TriggerEvent], str]:
    all_results: list[SearchResult] = []
    errors: list[str] = []
    for query in SEARCH_QUERIES:
        rss_url = google_news_rss_url(query)
        xml_text, error = fetch_raw_text(rss_url)
        if error:
            errors.append(f"{query}: {error}")
            continue
        try:
            all_results.extend(parse_google_news_rss(xml_text, query))
        except ET.ParseError as exc:
            errors.append(f"{query}: RSS parse failed: {exc}")
    discovery_hits, trigger_events = build_google_news_evidence(source, all_results)
    result = f"{len(all_results)} RSS items; {len(discovery_hits)} discovery hits; {len(trigger_events)} trigger events"
    if errors:
        result += "; errors: " + " | ".join(errors)
    return discovery_hits, trigger_events, result


def find_companies_on_source(source: Source, text: str) -> list[DiscoveryHit]:
    hits: list[DiscoveryHit] = []
    lower_text = text.lower()
    for company, meta in COMPANY_REGISTRY.items():
        matched = [alias for alias in meta["aliases"] if alias.lower() in lower_text]
        if not matched:
            continue
        rationale = f"Company name appeared on approved discovery source '{source.name}'."
        if source.name == "TIME HealthTech 2025":
            rationale = "Company appeared in TIME HealthTech 2025 ranking/article."
        hits.append(
            DiscoveryHit(
                company=company,
                source_name=source.name,
                source_type=source.source_type,
                discovery_url=source.url,
                discovery_rationale=rationale,
                product_type=meta["product_type"],
                geography=meta["geography"],
                website=meta["website"],
                matched_terms=", ".join(matched),
            )
        )
    return hits


def run_discovery(sources: list[Source]) -> tuple[list[DiscoveryHit], list[TriggerEvent], list[list[str]]]:
    discovery_hits: list[DiscoveryHit] = []
    trigger_events: list[TriggerEvent] = []
    run_log: list[list[str]] = []
    for source in sources:
        if not source.adapter:
            run_log.append([source.name, source.source_type, source.url, "Skipped", "No automated adapter yet"])
            continue
        if source.source_type == "Accelerator" and source.adapter == "accelerator_page":
            run_log.append([source.name, source.source_type, source.url, "Skipped", "No source-specific accelerator adapter yet"])
            continue
        if source.adapter == "google_news_search":
            hits, triggers, result = run_google_news_search(source)
            discovery_hits.extend(hits)
            trigger_events.extend(triggers)
            run_log.append([source.name, source.source_type, source.url, "Fetched", result])
            continue
        if source.adapter == "yc_healthcare":
            hits, triggers, result = run_yc_healthcare(source)
            discovery_hits.extend(hits)
            trigger_events.extend(triggers)
            run_log.append([source.name, source.source_type, source.url, "YC Healthcare directory adapter", result])
            continue
        if source.adapter == "medtech_innovator":
            hits, triggers, result = run_medtech_innovator(source)
            discovery_hits.extend(hits)
            trigger_events.extend(triggers)
            run_log.append([source.name, source.source_type, source.url, ADAPTER_STATUS_NAMES[source.adapter], result])
            continue
        if source.adapter == "digitalhealth_london":
            hits, triggers, result = run_digitalhealth_london(source)
            discovery_hits.extend(hits)
            trigger_events.extend(triggers)
            run_log.append([source.name, source.source_type, source.url, ADAPTER_STATUS_NAMES[source.adapter], result])
            continue
        if source.adapter == "mayo_accelerate":
            hits, triggers, result = run_mayo_accelerate(source)
            discovery_hits.extend(hits)
            trigger_events.extend(triggers)
            run_log.append([source.name, source.source_type, source.url, ADAPTER_STATUS_NAMES[source.adapter], result])
            continue
        if source.adapter == "eit_health_catapult":
            hits, triggers, result = run_eit_health_catapult(source)
            discovery_hits.extend(hits)
            trigger_events.extend(triggers)
            run_log.append([source.name, source.source_type, source.url, ADAPTER_STATUS_NAMES[source.adapter], result])
            continue
        if source.adapter in JOB_BOARD_ADAPTERS:
            hits, triggers, result = run_job_board_adapter(source, JOB_BOARD_ADAPTERS[source.adapter])
            discovery_hits.extend(hits)
            trigger_events.extend(triggers)
            run_log.append([source.name, source.source_type, source.url, ADAPTER_STATUS_NAMES[source.adapter], result])
            continue
        raw_html, error = fetch_raw_text(source.url)
        if error:
            run_log.append([source.name, source.source_type, source.url, "Fetch failed", error])
            continue
        source_triggers: list[TriggerEvent] = []
        if source.adapter in ADAPTER_STATUS_NAMES:
            hits, source_triggers = build_source_page_evidence(source, raw_html)
            trigger_events.extend(source_triggers)
        else:
            hits = find_companies_on_source(source, text_from_html(raw_html))
        discovery_hits.extend(hits)
        status = ADAPTER_STATUS_NAMES.get(source.adapter or "", "Fetched")
        run_log.append([source.name, source.source_type, source.url, status, f"{len(hits)} discovery hits; {len(source_triggers)} trigger events"])
    return discovery_hits, trigger_events, run_log


def adapter_inventory_label(source: Source) -> str:
    if source.source_type == "Accelerator" and source.adapter == "accelerator_page":
        return "Manual/not implemented"
    if source.adapter:
        return ADAPTER_STATUS_NAMES.get(source.adapter, source.adapter)
    return "Manual/not implemented"


def normalize_companies(discovery_hits: list[DiscoveryHit]) -> dict[str, CompanyRecord]:
    companies: dict[str, CompanyRecord] = {}
    for hit in discovery_hits:
        record = companies.setdefault(hit.company, CompanyRecord(company=hit.company))
        record.discovery_hits.append(hit)
        record.website = record.website or hit.website
        record.geography = record.geography or hit.geography
        record.product_type = record.product_type or hit.product_type
    return companies


def attach_trigger_events(companies: dict[str, CompanyRecord], trigger_events: list[TriggerEvent]) -> list[TriggerEvent]:
    attached: list[TriggerEvent] = []
    seen: set[tuple[str, str, str]] = set()
    for event in trigger_events:
        record = companies.get(event.company)
        if not record:
            continue
        key = (event.company.lower(), event.trigger_type, event.evidence_url)
        if key in seen:
            continue
        seen.add(key)
        record.triggers.append(event)
        attached.append(event)
    return attached


def mark_primary_triggers(companies: dict[str, CompanyRecord]) -> None:
    for record in companies.values():
        for idx, trigger in enumerate(record.triggers):
            trigger.trigger_role = "Primary" if idx == 0 else "Secondary"


def run_trigger_research(companies: dict[str, CompanyRecord]) -> list[TriggerEvent]:
    trigger_events: list[TriggerEvent] = []
    discovered = set(companies)
    for source in TRIGGER_SOURCES:
        text, error = fetch_text(source["url"])
        if error:
            continue
        lower_text = text.lower()
        for company in source["companies"]:
            if company not in discovered:
                continue
            aliases = COMPANY_REGISTRY[company]["aliases"]
            if any(alias.lower() in lower_text for alias in aliases):
                event = TriggerEvent(
                    company=company,
                    trigger_type=source["trigger_type"],
                    trigger_event=source["trigger_event"],
                    trigger_source=source["name"],
                    evidence_url=source["url"],
                )
                trigger_events.append(event)
    return attach_trigger_events(companies, trigger_events)


def score_company(record: CompanyRecord) -> tuple[dict[str, int], int, str, str, str, str]:
    text = " ".join(
        [record.product_type, record.company]
        + [h.discovery_rationale for h in record.discovery_hits]
        + [h.company_description for h in record.discovery_hits]
        + [h.category_or_track for h in record.discovery_hits]
        + [t.trigger_event for t in record.triggers]
    ).lower()
    flags = {
        "Recently funded +3": int(any(t.trigger_type == "Funding" for t in record.triggers)),
        "AI/SaMD/device +3": int(any(term in text for term in ["ai", "samd", "medical device", "diagnostic", "imaging", "wearable", "stethoscope"])),
        "Hiring QA/reg/V&V +3": int(any(term in text for term in ["regulatory affairs", "quality engineer", "design assurance", "v&v"])),
        "Clinical validation +2": int(any(term in text for term in ["clinical", "diagnostic", "screening", "validation"])),
        "FDA/CE/reg language +2": int(any(term in text for term in ["fda", "ce ", "samd", "regulated"])),
        "Grant/public funding +2": int(any(h.source_type == "Grant/funding" for h in record.discovery_hits)),
        "University/grant origin +2": int(any(h.source_type in ["University/spinout", "Grant/funding"] for h in record.discovery_hits)),
        "No obvious reg team +2": 0,
        "Pre-commercial +1": int(any(h.source_type == "Accelerator" for h in record.discovery_hits)),
        "Large company -1": 0,
        "Wellness/non-medical -2": -1 if "wellness" in text else 0,
        "Pharma-only -2": -1 if "pharma-only" in text else 0,
    }
    score = (
        flags["Recently funded +3"] * 3
        + flags["AI/SaMD/device +3"] * 3
        + flags["Hiring QA/reg/V&V +3"] * 3
        + flags["Clinical validation +2"] * 2
        + flags["FDA/CE/reg language +2"] * 2
        + flags["Grant/public funding +2"] * 2
        + flags["University/grant origin +2"] * 2
        + flags["No obvious reg team +2"] * 2
        + flags["Pre-commercial +1"]
        + flags["Large company -1"] * -1
        + flags["Wellness/non-medical -2"] * 2
        + flags["Pharma-only -2"] * 2
    )
    band = "Strong" if score >= 10 else "Good" if score >= 7 else "Maybe" if score >= 4 else "Low"
    persona = "AI/SaMD or healthtech company from approved source"
    quadrant = "Advisory"
    secondary = "Design/dev" if "imaging" in text or "diagnostic" in text else "Embedded support"
    pain = "Likely needs a defensible regulatory, validation, claims, or productisation story before broader commercial expansion."
    return flags, score, band, persona, quadrant, secondary, pain


def primary_discovery(record: CompanyRecord) -> DiscoveryHit:
    return sorted(record.discovery_hits, key=lambda h: h.source_name)[0]


def primary_trigger(record: CompanyRecord) -> TriggerEvent | None:
    primary = [t for t in record.triggers if t.trigger_role == "Primary"]
    return primary[0] if primary else None


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(name="Aptos", size=10, color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Aptos", size=10)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False


def size_columns(ws, widths=None):
    widths = widths or {}
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        if letter in widths:
            ws.column_dimensions[letter].width = widths[letter]
            continue
        max_len = max((len(str(cell.value)) for cell in ws[letter] if cell.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)


def add_hyperlinks(ws, cols: list[int]):
    for row in range(2, ws.max_row + 1):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def write_workbook(companies: dict[str, CompanyRecord], discovery_hits: list[DiscoveryHit], trigger_events: list[TriggerEvent], run_log: list[list[str]]):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Pipeline Summary")
    append_excel_row(ws, ["Metric", "Value"])
    append_excel_row(ws, ["Run date", TODAY])
    append_excel_row(ws, ["Approved sources", len(SOURCES)])
    append_excel_row(ws, ["Fetched/skipped sources", len(run_log)])
    append_excel_row(ws, ["Discovery hits", len(discovery_hits)])
    append_excel_row(ws, ["Companies", len(companies)])
    append_excel_row(ws, ["Trigger events", len(trigger_events)])
    style_sheet(ws)
    size_columns(ws, {"A": 28, "B": 24})

    ws = wb.create_sheet("Pipeline Run Log")
    append_excel_row(ws, ["Source", "Source type", "URL", "Status", "Result"])
    for row in run_log:
        append_excel_row(ws, row)
    style_sheet(ws)
    add_hyperlinks(ws, [3])
    size_columns(ws, {"A": 26, "B": 18, "C": 52, "D": 16, "E": 42})

    ws = wb.create_sheet("Source Inventory")
    append_excel_row(ws, ["Source name", "Type", "URL", "Geography", "Priority", "Update cadence", "Extraction method", "Signal / notes", "Automated adapter"])
    for s in SOURCES:
        append_excel_row(ws, [s.name, s.source_type, s.url, s.geography, s.priority, s.update_cadence, s.extraction_method, s.notes, adapter_inventory_label(s)])
    style_sheet(ws)
    add_hyperlinks(ws, [3])
    size_columns(ws, {"A": 30, "B": 20, "C": 54, "H": 58, "I": 22})

    ws = wb.create_sheet("Source Playbooks")
    append_excel_row(ws, ["Source type", "Search/query terms", "Filters", "Output fields", "Expected evidence", "Best-fit BBT wedge"])
    for row in SOURCE_PLAYBOOKS:
        append_excel_row(ws, row)
    style_sheet(ws)
    size_columns(ws, {"A": 22, "B": 56, "C": 44, "D": 44, "E": 44, "F": 22})

    ws = wb.create_sheet("Discovery Hits")
    append_excel_row(ws, ["Company", "Discovery source", "Source type", "Discovery evidence URL", "Discovery rationale", "Matched terms", "Website", "Geography", "Product type", "Accelerator program", "Cohort label", "Cohort year", "Category / track", "Company description", "Captured at"])
    for hit in discovery_hits:
        append_excel_row(ws, [
            hit.company,
            hit.source_name,
            hit.source_type,
            hit.discovery_url,
            hit.discovery_rationale,
            hit.matched_terms,
            hit.website,
            hit.geography,
            hit.product_type,
            hit.accelerator_program,
            hit.cohort_label,
            hit.cohort_year,
            hit.category_or_track,
            hit.company_description,
            hit.captured_at,
        ])
    style_sheet(ws)
    add_hyperlinks(ws, [4, 7])
    size_columns(ws, {"A": 24, "B": 28, "C": 18, "D": 54, "E": 56, "F": 24, "G": 36, "I": 28, "J": 26, "K": 24, "M": 28, "N": 58})

    ws = wb.create_sheet("Lead Intake")
    append_excel_row(ws, ["Company", "Website", "Geography", "Product type", "Accelerator program", "Cohort label", "Cohort year", "Category / track", "Company description", "Discovery source", "Discovery source type", "Discovery evidence URL", "Discovery rationale", "Primary trigger event", "Trigger source", "Trigger evidence URL", "Evidence status", "Date captured"])
    for record in sorted(companies.values(), key=lambda r: r.company):
        hit = primary_discovery(record)
        trigger = primary_trigger(record)
        append_excel_row(ws, [
            record.company,
            record.website,
            record.geography,
            record.product_type,
            hit.accelerator_program,
            hit.cohort_label,
            hit.cohort_year,
            hit.category_or_track,
            hit.company_description,
            hit.source_name,
            hit.source_type,
            hit.discovery_url,
            hit.discovery_rationale,
            trigger.trigger_event if trigger else "",
            trigger.trigger_source if trigger else "",
            trigger.evidence_url if trigger else "",
            "Verified trigger" if trigger else "Discovered only",
            TODAY,
        ])
    style_sheet(ws)
    add_hyperlinks(ws, [2, 12, 16])
    size_columns(ws, {"A": 24, "B": 36, "C": 16, "D": 28, "E": 26, "F": 24, "H": 28, "I": 58, "J": 28, "L": 54, "M": 58, "N": 58, "O": 28, "P": 54})

    ws = wb.create_sheet("Trigger Log")
    append_excel_row(ws, ["Company", "Trigger type", "Trigger event", "Trigger source", "Evidence URL", "Trigger role", "Captured at"])
    for event in trigger_events:
        append_excel_row(ws, [event.company, event.trigger_type, event.trigger_event, event.trigger_source, event.evidence_url, event.trigger_role, event.captured_at])
    style_sheet(ws)
    add_hyperlinks(ws, [5])
    size_columns(ws, {"A": 24, "B": 20, "C": 64, "D": 30, "E": 54, "F": 16})

    ws = wb.create_sheet("Lead Scoring")
    headers = [
        "Company", "Persona", "Primary BBT quadrant", "Secondary tag", "Pain hypothesis",
        "Recently funded +3", "AI/SaMD/device +3", "Hiring QA/reg/V&V +3", "Clinical validation +2",
        "FDA/CE/reg language +2", "Grant/public funding +2", "University/grant origin +2",
        "No obvious reg team +2", "Pre-commercial +1", "Large company -1", "Wellness/non-medical -2",
        "Pharma-only -2", "Score", "Priority band", "Evidence status", "Primary evidence URL",
    ]
    append_excel_row(ws, headers)
    scoring_rows = []
    for record in sorted(companies.values(), key=lambda r: r.company):
        flags, score, band, persona, quadrant, secondary, pain = score_company(record)
        trigger = primary_trigger(record)
        hit = primary_discovery(record)
        evidence_url = trigger.evidence_url if trigger else hit.discovery_url
        status = "Verified trigger" if trigger else "Discovered only"
        append_excel_row(ws, [
            record.company, persona, quadrant, secondary, pain,
            *flags.values(), score, band, status, evidence_url,
        ])
        scoring_rows.append((score, band, status, record, pain, evidence_url))
    style_sheet(ws)
    add_hyperlinks(ws, [21])
    size_columns(ws, {"A": 24, "B": 32, "C": 22, "D": 22, "E": 64, "R": 10, "S": 14, "T": 18, "U": 54})

    scoring_rows.sort(key=lambda row: (row[2] == "Verified trigger", row[0]), reverse=True)
    ws = wb.create_sheet("Weekly Review")
    append_excel_row(ws, ["Rank", "Company", "Score", "Priority band", "Evidence status", "Why shortlisted", "Next action", "Owner", "Status", "Review notes", "Evidence URL"])
    for idx, (score, band, status, record, pain, evidence_url) in enumerate(scoring_rows[:15], start=1):
        append_excel_row(ws, [
            idx,
            record.company,
            score,
            band,
            status,
            pain,
            "Contact research" if status == "Verified trigger" else "Run trigger research before outreach",
            "",
            "Validate",
            "",
            evidence_url,
        ])
    style_sheet(ws)
    add_hyperlinks(ws, [11])
    size_columns(ws, {"A": 8, "B": 24, "C": 10, "D": 14, "E": 18, "F": 62, "G": 36, "K": 54})

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    try:
        wb.save(OUT)
        return OUT
    except PermissionError:
        fallback = OUT.with_name("BlueBridge_TOFU_BizDev_V1_pipeline.xlsx")
        wb.save(fallback)
        return fallback


def main():
    discovery_hits, search_trigger_events, run_log = run_discovery(SOURCES)
    companies = normalize_companies(discovery_hits)
    trigger_events = attach_trigger_events(companies, search_trigger_events)
    trigger_events.extend(run_trigger_research(companies))
    mark_primary_triggers(companies)
    output = write_workbook(companies, discovery_hits, trigger_events, run_log)
    print(output.resolve())
    print(f"discovery_hits={len(discovery_hits)} companies={len(companies)} trigger_events={len(trigger_events)}")


if __name__ == "__main__":
    main()
